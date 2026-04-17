"""
Reference Job Finder — Search the Quote Index for similar past jobs.

Usage:
    python reffinder.py --job J15874
    python reffinder.py --job J15874 --jobs-root "//NPSVR05/FOXFAB_REDIRECT$/lbadong/Desktop/AGENT ARMY/Testing"

Output: prints a JSON document to stdout with matched reference jobs.
Errors: UNKNOWN_JOB:, NO_PRF:, NO_MATCHES: written to stderr.
"""

import argparse
import json
import os
import re
import sys
from pathlib import Path

import openpyxl

# Module-level caches
_JOB_FOLDER_INDEX: dict[str, str] | None = None  # upper(J####) -> folder path
_CAD_CACHE: dict[str, bool] = {}                  # J#### -> has CAD
_CAD_CACHE_DIRTY = False

# ── Configuration ─────────────────────────────────────────────────────
JOBS_ROOT    = Path(r"Z:\FOXFAB_DATA\ENGINEERING\2 JOBS")  # Production path — read-only, never modify
QUOTE_INDEX  = Path(r"\\NPSVR05\FOXFAB_REDIRECT$\lbadong\Desktop\AGENT ARMY\Quote Index - LAMBERT.xlsx")
PRF_SUBPATH  = Path("300 Inputs") / "302 Production Release Form"
BOM_SUBPATH  = Path("100 Elec") / "101 Bill of Materials"
CAD_SUBPATH  = Path("200 Mech") / "201 CAD"
INDEX_SHEET  = "Index"
EXCEL_EXTS   = {".xlsx", ".xlsm", ".xls"}

# Matching tolerances
AMP_TOLERANCE  = 200   # +/- amps
SIZE_TOLERANCE = 2     # +/- inches per dimension


# ── Errors ────────────────────────────────────────────────────────────
def die(code: str, msg: str = ""):
    sys.stderr.write(f"{code}: {msg}\n")
    sys.exit(1)


# ── Job folder resolution ─────────────────────────────────────────────
def find_job_folder(job_number: str) -> Path:
    if not JOBS_ROOT.is_dir():
        die("UNKNOWN_JOB", f"Jobs root not found: {JOBS_ROOT}")
    job_up = job_number.upper()
    matches = [d for d in JOBS_ROOT.iterdir()
               if d.is_dir() and d.name.upper().startswith(job_up)]
    if not matches:
        die("UNKNOWN_JOB", f"No folder matches '{job_number}' in {JOBS_ROOT}")
    if len(matches) > 1:
        # Pick the exact match or first alphabetically
        exact = [m for m in matches if m.name.upper().split()[0] == job_up]
        if exact:
            return exact[0]
    return matches[0]


# ── PRF reading ───────────────────────────────────────────────────────
def find_prf(job_folder: Path) -> Path:
    prf_folder = job_folder / PRF_SUBPATH
    if not prf_folder.is_dir():
        die("NO_PRF", f"PRF folder missing: {prf_folder}")
    prfs = [f for f in prf_folder.iterdir()
            if f.is_file()
            and f.suffix.lower() in EXCEL_EXTS
            and "prf" in f.name.lower()]
    if not prfs:
        die("NO_PRF", f"No PRF file found in {prf_folder}")
    return sorted(prfs, key=lambda p: p.name.lower())[0]


def read_prf(prf_path: Path) -> dict:
    """Extract Model No (G9 — Drawing Number), Enclosure Size (G18), Current (C11) from PRF."""
    wb = openpyxl.load_workbook(str(prf_path), read_only=True, data_only=True)
    ws = wb["Form"]
    result = {
        "model_no": str(ws["G9"].value or "").strip(),
        "enclosure_size": str(ws["G18"].value or "").strip(),
        "current": str(ws["C11"].value or "").strip(),
        "job_name": str(ws["C8"].value or "").strip(),
    }
    wb.close()
    return result


# ── Parsing helpers ───────────────────────────────────────────────────
def parse_model_number(model_no: str) -> dict:
    """Parse FFCC-P2-500-FT-V4-... into family, config, amperage.
    Delegates amperage to parse_amperage_from_description so compound codes
    (e.g. '1600LSI', '400LSI80') are handled and numeric material grades
    (e.g. '304', '316' stainless steel) are skipped."""
    parts = model_no.split("-")
    family = parts[0] if parts else ""
    config = parts[1] if len(parts) > 1 else ""
    amperage = parse_amperage_from_description(model_no)
    return {"family": family, "config": config, "amperage": amperage}


def parse_size(size_str: str) -> list[int] | None:
    """Parse 'H x W x D' or 'HxWxD' into [H, W, D] integers."""
    if not size_str:
        return None
    # Remove all spaces, then split on 'x' or 'X'
    nums = re.findall(r"\d+", size_str)
    if len(nums) >= 3:
        return [int(n) for n in nums[:3]]
    return None


def parse_amperage_from_current(current_str: str) -> int | None:
    """Parse '500A RMS' or '500' into integer amperage."""
    m = re.search(r"(\d+)", current_str)
    return int(m.group(1)) if m else None


def parse_amperage_from_description(desc: str) -> int | None:
    """Extract the first amperage-like segment from a description code.
    Handles:
      'FFCC-P2-1200-LSI100-V4'      -> 1200  (pure digit)
      'FFCC-C3-750LSI-750LSI-...'   -> 750   (digits fused to breaker type)
      'FFCC-C3-1000TM-1000TM-...'   -> 1000
      'FFCC-C3-1200LSI100-...'      -> 1200  (ignores trailing 100)
    """
    parts = desc.split("-")
    for p in parts[2:]:  # skip family and config
        cleaned = p.strip()
        # Pure numeric segment: "1200"
        if cleaned.isdigit() and int(cleaned) >= 50:
            return int(cleaned)
        # Leading digits followed by breaker code: "750LSI", "800TM", "1200LSI100"
        m = re.match(r"^(\d{2,4})(?:[A-Z]|$)", cleaned)
        if m:
            val = int(m.group(1))
            if val >= 50:
                return val
    return None


_COLOR_CODES    = {"V3", "V4", "V5", "U3", "U4", "U5", "UC4", "UC5",
                   "C4", "C5", "C6", "U6"}
# Enclosure metal only. Note: G / GL / L are NOT materials — they are
# functional accessories:
#   G  = ground bar
#   L  = load bank
#   GL = ground + load bank
# These stay as strong matchable segments (see STRONG_SEGS below).
_MATERIAL_CODES = {"ALU", "316", "316SS", "304", "304N4", "3044X",
                   "STL", "STEEL"}
# Low-signal accessory codes — too common to distinguish candidates.
_IGNORE_CODES   = {"GENAP1"}

# Regex identifying a breaker segment — pure digits or digits+breaker code
# (LSI, TM, LSIG, KS, etc.), optionally with trailing interrupt-rating digits
# (LSI100, TM80, LSIG100, etc.). Matches "1000LSI100", "500TM", "400", "600KS".
_BREAKER_RE = re.compile(r"^\d{2,4}([A-Z]{1,5}\d{0,3})?$")


def count_breakers(model: str) -> int:
    """Count the number of breaker segments in a model code.
    A single-breaker config has 1, a dual-breaker has 2, triple has 3, etc.
    Only counts the leading contiguous run of breaker-shaped segments so that
    later numeric accessory codes (e.g. rating spec) don't inflate the count.
    """
    # Take just the first description (before any comma) and split on '-'
    first = model.split(",")[0]
    parts = [p.strip().upper() for p in first.split("-")]
    # Skip family/config, then start scanning. Switch types (T, FT, ATS, MTS)
    # break the breaker run — they separate the breaker block from accessories.
    count = 0
    started = False
    for p in parts[2:]:
        if p in {"T", "FT", "ATS", "MTS"}:
            break
        if _BREAKER_RE.match(p):
            count += 1
            started = True
        elif started:
            # Left the contiguous breaker block
            break
    return count


def has_switch_type(model: str) -> str | None:
    """Return the first switch-type designator (T/FT/ATS/MTS) found, or None."""
    parts = [p.strip().upper() for p in model.replace(",", "-").split("-")]
    for p in parts[2:]:
        if p in {"T", "FT", "ATS", "MTS"}:
            return p
    return None


def model_segments(model: str) -> set[str]:
    """Return lug-config + accessory segments only (colors/materials ignored).
    e.g. 'FFCC-P2-500-FT-V4-ALU-LA-GENAP1' -> {'FT','LA','GENAP1'}
    Used for rank-boosting when many candidates match basic criteria."""
    segs: set[str] = set()
    # Split on both '-' and ',' so comma-separated descriptions contribute all tokens.
    for chunk in re.split(r"[,\s]+", model):
        for p in chunk.split("-")[3:]:  # skip family, config, amps
            s = p.strip().upper()
            if not s:
                continue
            # Skip pure numerics (amp ratings already scored separately)
            if s.isdigit():
                continue
            # Skip color, material, and low-signal accessory codes.
            if s in _COLOR_CODES or s in _MATERIAL_CODES or s in _IGNORE_CODES:
                continue
            segs.add(s)
    return segs


def parse_family_config_from_description(desc: str) -> tuple[str, str]:
    """Extract family and config from description code."""
    parts = desc.split("-")
    family = parts[0].strip() if parts else ""
    config = parts[1].strip() if len(parts) > 1 else ""
    # Handle descriptions with commas (multiple items)
    family = family.split(",")[-1].strip()  # take last if comma-separated
    return family, config


# ── Quote Index cache ─────────────────────────────────────────────────
def _index_cache_path() -> Path:
    return Path(__file__).resolve().parent.parent / "index_cache.json"


def load_index_rows() -> list[dict]:
    """Return the pre-filtered Shipped/Won rows from the Quote Index.

    Uses a local JSON snapshot keyed on the source xlsx mtime so that repeat
    runs don't re-parse the 10MB network file. Only the 6 columns we actually
    use are stored, keeping the cache well under 1MB.
    """
    cache = _index_cache_path()
    src_mtime = QUOTE_INDEX.stat().st_mtime if QUOTE_INDEX.is_file() else 0.0

    if cache.is_file():
        try:
            payload = json.loads(cache.read_text(encoding="utf-8"))
            if payload.get("mtime") == src_mtime and isinstance(payload.get("rows"), list):
                return payload["rows"]
        except Exception:
            pass  # fall through and rebuild

    # Rebuild cache from source xlsx
    rows: list[dict] = []
    wb = openpyxl.load_workbook(str(QUOTE_INDEX), read_only=True, data_only=True)
    ws = wb[INDEX_SHEET]
    for row in ws.iter_rows(min_row=2, values_only=True):
        status = str(row[16] or "").strip()
        if not status:
            continue
        status_lc = status.lower()
        if "shipped" not in status_lc and "won" not in status_lc:
            continue
        description = str(row[5] or "").strip()
        enc_size_str = str(row[6] or "").strip()
        if not description and not enc_size_str:
            continue
        rows.append({
            "quote_number": str(row[2] or "").strip(),      # C
            "job_number":   str(row[3] or "").strip(),      # D
            "job_name":     str(row[4] or "").strip(),      # E
            "description":  description,                    # F
            "enc_size":     enc_size_str,                   # G
            "status":       status,                         # Q
            "job_number_alt": str(row[18] or "").strip(),   # S
        })
    wb.close()

    try:
        cache.write_text(
            json.dumps({"mtime": src_mtime, "rows": rows}),
            encoding="utf-8",
        )
    except Exception:
        pass  # caching is best-effort
    return rows


# ── Quote Index search ────────────────────────────────────────────────
def search_quote_index(prf_data: dict, parsed_model: dict, source_job: str) -> list[dict]:
    """Search Quote Index for matching shipped/won jobs."""
    index_rows = load_index_rows()

    source_job_upper = source_job.upper().strip()
    source_size = parse_size(prf_data["enclosure_size"])
    source_breakers = count_breakers(prf_data["model_no"])
    source_switch = has_switch_type(prf_data["model_no"])
    source_amp = parsed_model["amperage"]
    if source_amp is None:
        source_amp = parse_amperage_from_current(prf_data["current"])

    candidates = []
    for r in index_rows:
        # Pre-filtered for Shipped/Won in load_index_rows
        job_num = r["job_number"]
        job_num_alt = r["job_number_alt"]
        job_name = r["job_name"]
        description = r["description"]
        enc_size_str = r["enc_size"]
        quote_num = r["quote_number"]
        status_str = r["status"]

        # Use whichever job number is available
        effective_job = job_num if job_num and job_num != "0" else job_num_alt

        # Skip self-match (the source job)
        if effective_job and effective_job.upper().strip() == source_job_upper:
            continue

        # ── Match product family (hard filter) ──
        # Family is a hard filter — FFCC vs FFTS vs FFMDC is a fundamentally
        # different product line. Config (C1/C2/C3/P2/S1/B1) is soft — layouts
        # often translate across configs when size + amperage + architecture line up.
        desc_family, desc_config = parse_family_config_from_description(description)
        if parsed_model["family"] and desc_family.upper() != parsed_model["family"].upper():
            continue
        config_match = bool(
            parsed_model["config"]
            and desc_config.upper() == parsed_model["config"].upper()
        )

        # ── Amperage (soft match — ranking only, not a hard filter) ──
        desc_amp = parse_amperage_from_description(description)
        amp_diff = None
        if source_amp is not None and desc_amp is not None:
            amp_diff = abs(source_amp - desc_amp)

        # ── Match enclosure size (+-2" per dimension) ──
        cand_size = parse_size(enc_size_str)
        size_match = False
        size_diff = None
        if source_size and cand_size:
            diffs = [abs(s - c) for s, c in zip(source_size, cand_size)]
            size_match = all(d <= SIZE_TOLERANCE for d in diffs)
            size_diff = sum(diffs)
        elif source_size is None and cand_size is None:
            size_match = True

        if not size_match:
            continue

        # ── Score: prefer closer matches ──
        score = 100
        # Config match is a mild preference, not a requirement.
        if config_match:
            score += 15
        # Amperage distance is a strong signal — closer = much better.
        #   Exact (0A):            +30
        #   Within 50A:            +20
        #   Within 150A:           +10
        #   Within 300A:            0
        #   Beyond 300A:  linear penalty (-0.1 per amp over 300)
        if amp_diff is not None:
            if amp_diff == 0:
                score += 30
            elif amp_diff <= 50:
                score += 20
            elif amp_diff <= 150:
                score += 10
            elif amp_diff > 300:
                score -= (amp_diff - 300) * 0.1
        if size_diff is not None:
            score -= size_diff * 2    # penalty for size distance

        # ── Breaker count (single vs dual vs triple) ──
        # Dual-breaker configs are architecturally very different from
        # single-breaker + transfer-switch configs. Exact match = big bonus.
        cand_breakers = count_breakers(description)
        if source_breakers > 0 and cand_breakers > 0:
            if source_breakers == cand_breakers:
                score += 30   # same breaker layout
            else:
                score -= 10 * abs(source_breakers - cand_breakers)

        # ── Switch-type mismatch penalty ──
        # If the source is NOT a transfer-switch job but the candidate IS
        # (or vice-versa), the internal layout is fundamentally different.
        cand_switch = has_switch_type(description)
        if bool(source_switch) != bool(cand_switch):
            score -= 20

        # Soft recency preference: newer (higher) job numbers get a tiny bonus.
        # (Job_num - 10000) * 0.001 yields a ~2-point range across the fleet,
        # enough to break ties without overriding structural match signals.
        m = re.search(r"(\d{3,})", effective_job)
        if m:
            score += max(0, (int(m.group(1)) - 10000)) * 0.001

        # Capture segment overlap for optional tie-breaking / rank-boost
        cand_segs = model_segments(description)

        candidates.append({
            "job_number": effective_job,
            "job_name": job_name,
            "description": description,
            "enclosure_size": enc_size_str,
            "amperage": desc_amp,
            "breaker_count": cand_breakers,
            "switch_type": cand_switch,
            "quote_number": quote_num,
            "status": status_str,
            "score": score,
            "_segments": cand_segs,
        })

    # Rank-boost candidates that share model segments with the source.
    # Some segments carry extra weight because they drive the physical layout.
    # - Pole count (3P, 4P): +25 (major enclosure / switch architecture signal)
    # - Switch / breaker type (T, FT, ATS, MTS): +15
    # - Fused configuration (FUS): +15
    # - Everything else (lug config, accessories): +5
    # Applied to any pool size >20; smaller pools rely solely on amp/size ranking.
    POLE_SEGS        = {"3P", "4P"}
    SWITCH_TYPE_SEGS = {"T", "FT", "ATS", "MTS"}
    STRONG_SEGS      = {"FUS", "G", "GL", "L"}  # FUS=fused, G=ground, L=loadbank, GL=both

    if len(candidates) > 20:
        source_segs = model_segments(prf_data["model_no"])
        for c in candidates:
            shared = source_segs & c["_segments"]
            c["shared_segments"] = sorted(shared)
            bonus = 0
            for s in shared:
                if s in POLE_SEGS:
                    bonus += 25
                elif s in SWITCH_TYPE_SEGS or s in STRONG_SEGS:
                    bonus += 15
                else:
                    bonus += 5
            c["score"] += bonus

            # Penalise candidates that clearly have a *different* pole count.
            # (Only when both source and candidate advertise a pole explicitly.)
            src_poles = source_segs & POLE_SEGS
            cand_poles = c["_segments"] & POLE_SEGS
            if src_poles and cand_poles and not (src_poles & cand_poles):
                c["score"] -= 25

    # Strip internal segment set (not JSON-serialisable) before returning
    for c in candidates:
        c.pop("_segments", None)

    # Sort by score descending
    candidates.sort(key=lambda x: x["score"], reverse=True)
    return candidates


# ── CAD folder validation + caching ───────────────────────────────────
def _cache_path() -> Path:
    return Path(__file__).resolve().parent.parent / "cad_cache.json"


def load_cad_cache() -> None:
    """Load the CAD-status cache from disk into module state."""
    global _CAD_CACHE
    p = _cache_path()
    if p.is_file():
        try:
            _CAD_CACHE = json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            _CAD_CACHE = {}
    else:
        _CAD_CACHE = {}


def save_cad_cache() -> None:
    """Write the CAD-status cache to disk if dirty."""
    global _CAD_CACHE_DIRTY
    if not _CAD_CACHE_DIRTY:
        return
    p = _cache_path()
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(_CAD_CACHE, indent=2, sort_keys=True), encoding="utf-8")
    _CAD_CACHE_DIRTY = False


def build_job_folder_index() -> dict[str, str]:
    """One-time scan of JOBS_ROOT; builds {J_prefix_upper: folder_path}.
    The J_prefix is the first whitespace-separated token of the folder name."""
    global _JOB_FOLDER_INDEX
    if _JOB_FOLDER_INDEX is not None:
        return _JOB_FOLDER_INDEX
    idx: dict[str, str] = {}
    if JOBS_ROOT.is_dir():
        try:
            for d in JOBS_ROOT.iterdir():
                if d.is_dir():
                    token = d.name.split()[0].upper() if d.name.split() else d.name.upper()
                    # First match wins (stable alphabetical from iterdir on Windows)
                    if token not in idx:
                        idx[token] = str(d)
        except (PermissionError, OSError):
            pass
    _JOB_FOLDER_INDEX = idx
    return idx


def find_candidate_folder(job_number: str) -> Path | None:
    """Look up a job folder via the pre-built index."""
    if not job_number or job_number == "0":
        return None
    idx = build_job_folder_index()
    job_up = job_number.upper().strip()
    hit = idx.get(job_up)
    if hit:
        return Path(hit)
    # Fallback: prefix match (handles weird job naming edge cases)
    for key, path in idx.items():
        if key.startswith(job_up):
            return Path(path)
    return None


def has_cad_content_cached(job_number: str, job_folder: Path) -> bool:
    """Check 201 CAD for real files, with persistent caching keyed by job number."""
    global _CAD_CACHE_DIRTY
    key = job_number.upper().strip()
    if key in _CAD_CACHE:
        return _CAD_CACHE[key]

    cad = job_folder / CAD_SUBPATH
    result = False
    if cad.is_dir():
        try:
            for f in cad.iterdir():
                if f.is_file() and not f.name.startswith("~$") and f.name != "Thumbs.db":
                    result = True
                    break
        except (PermissionError, OSError):
            result = False

    _CAD_CACHE[key] = result
    _CAD_CACHE_DIRTY = True
    return result


# ── DWG-embedded xlsx detection + EMF preview extraction ──────────────
def xlsx_has_embedded_dwg(path: Path) -> bool:
    """True if the xlsx's sheet1 is empty but it contains an OLE embedding."""
    import zipfile, re as _re
    try:
        with zipfile.ZipFile(path) as z:
            names = z.namelist()
            sheet = "xl/worksheets/sheet1.xml"
            if sheet not in names:
                return False
            xml = z.read(sheet).decode("utf-8", errors="replace")
            if _re.search(r"<row[\s>]", xml):
                return False  # real data present
            embeds = [n for n in names if n.startswith("xl/embeddings/") and n.endswith(".bin")]
            return bool(embeds)
    except Exception:
        return False


def extract_emf_as_png(xlsx_path: Path, out_png: Path, dpi: int = 300) -> bool:
    """Pull xl/media/image1.emf from the xlsx and render it to PNG."""
    import zipfile
    try:
        with zipfile.ZipFile(xlsx_path) as z:
            emfs = [n for n in z.namelist()
                    if n.startswith("xl/media/") and n.lower().endswith(".emf")]
            if not emfs:
                return False
            emf_bytes = z.read(emfs[0])
    except Exception:
        return False

    out_png.parent.mkdir(parents=True, exist_ok=True)
    tmp_emf = out_png.with_suffix(".emf")
    tmp_emf.write_bytes(emf_bytes)

    # Try Pillow first
    try:
        from PIL import Image
        img = Image.open(str(tmp_emf))
        try:
            img.load(dpi=dpi)
        except TypeError:
            img.load()
        img.save(str(out_png), "PNG")
        tmp_emf.unlink(missing_ok=True)
        return True
    except Exception:
        pass

    # Fallback: pywin32 GDI+
    try:
        import win32ui, win32con, win32gui
        from ctypes import windll
        hemf = windll.gdi32.GetEnhMetaFileW(str(tmp_emf))
        if not hemf:
            return False
        w = int(11 * dpi)
        h = int(8.5 * dpi)
        hdc_screen = win32gui.GetDC(0)
        mem_dc = win32ui.CreateDCFromHandle(hdc_screen).CreateCompatibleDC()
        bmp = win32ui.CreateBitmap()
        bmp.CreateCompatibleBitmap(win32ui.CreateDCFromHandle(hdc_screen), w, h)
        mem_dc.SelectObject(bmp)
        mem_dc.FillSolidRect((0, 0, w, h), 0xFFFFFF)
        rect = (0, 0, w, h)
        windll.gdi32.PlayEnhMetaFile(mem_dc.GetHandleOutput(), hemf, rect)
        bmp.SaveBitmapFile(mem_dc, str(out_png.with_suffix(".bmp")))
        from PIL import Image
        Image.open(str(out_png.with_suffix(".bmp"))).save(str(out_png), "PNG")
        out_png.with_suffix(".bmp").unlink(missing_ok=True)
        tmp_emf.unlink(missing_ok=True)
        windll.gdi32.DeleteEnhMetaFile(hemf)
        return True
    except Exception:
        return False


# ── BOM reading ───────────────────────────────────────────────────────
def find_bom_file(job_folder: Path) -> Path | None:
    """Find the BOM xlsx in the job's electrical BOM folder."""
    bom_folder = job_folder / BOM_SUBPATH
    if not bom_folder.is_dir():
        return None
    bom_files = [f for f in bom_folder.iterdir()
                 if f.is_file() and f.suffix.lower() in EXCEL_EXTS
                 and "bom" in f.name.lower() and not f.name.startswith("~")]
    return bom_files[0] if bom_files else None


def read_bom_descriptions(job_folder: Path) -> list[str]:
    """Read DESCRIPTION column from the electrical BOM. Returns list of description strings.
    For DWG-embedded BOMs, returns empty list (needs visual extraction by Claude)."""
    bf = find_bom_file(job_folder)
    if bf is None:
        return []

    # Check if DWG-embedded — cannot read programmatically
    if xlsx_has_embedded_dwg(bf):
        return []  # will be handled via EMF_IMAGE output

    try:
        wb = openpyxl.load_workbook(str(bf), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        # Find the DESCRIPTION column (usually E)
        desc_col = None
        for cell in next(ws.iter_rows(min_row=1, max_row=3, values_only=False)):
            val = str(cell.value or "").strip().upper()
            if val in ("DESCRIPTION", "DESC", "ITEM DESCRIPTION", "PART DESCRIPTION"):
                desc_col = cell.column - 1
                break

        descriptions = []
        if desc_col is not None:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if desc_col < len(row) and row[desc_col]:
                    desc = str(row[desc_col]).strip()
                    if desc and desc.upper() not in ("DESCRIPTION", "DESC"):
                        descriptions.append(desc)
        wb.close()
        return descriptions
    except Exception:
        return []


def _bom_cache_dir() -> Path:
    """Shared BOM image cache — one PNG per job, reused across source jobs."""
    return Path(__file__).resolve().parent.parent / "bom_cache"


def extract_bom_images(job_folder: Path, job_number: str, output_dir: Path) -> str | None:
    """For DWG-embedded BOMs, extract EMF → PNG with a persistent shared cache.

    Returns the PNG path or None. The PNG is stored in the skill-wide cache
    (`bom_cache/<JOB>_BOM.png`) and is invalidated when the source BOM xlsx
    mtime changes. `output_dir` is accepted for backwards compatibility but
    no longer used — cached PNGs are shared across source jobs.
    """
    bf = find_bom_file(job_folder)
    if bf is None:
        return None
    if not xlsx_has_embedded_dwg(bf):
        return None

    cache_dir = _bom_cache_dir()
    cache_dir.mkdir(parents=True, exist_ok=True)
    cached_png = cache_dir / f"{job_number}_BOM.png"
    stamp_file = cache_dir / f"{job_number}_BOM.mtime"
    try:
        src_mtime = str(bf.stat().st_mtime)
    except OSError:
        src_mtime = ""

    # Reuse cached PNG if BOM xlsx hasn't changed since last extraction
    if cached_png.is_file() and stamp_file.is_file():
        try:
            if stamp_file.read_text(encoding="utf-8").strip() == src_mtime:
                return str(cached_png)
        except Exception:
            pass

    if extract_emf_as_png(bf, cached_png):
        try:
            stamp_file.write_text(src_mtime, encoding="utf-8")
        except Exception:
            pass
        return str(cached_png)
    return None


def categorize_component(desc: str) -> str:
    """Categorize a BOM description into a component type."""
    d = desc.upper()
    if any(k in d for k in ("BREAKER", "MCCB", "MCB", "CB ")):
        return "breakers"
    if any(k in d for k in ("TERMINAL", "TERM BLOCK", "TB ")):
        return "terminal_blocks"
    if any(k in d for k in ("LUG", "CONNECTOR")):
        return "lugs"
    if any(k in d for k in ("HEATER", "STAT", "THERMOSTAT")):
        return "heaters"
    if any(k in d for k in ("METER", "POWER METER")):
        return "meters"
    if any(k in d for k in ("RELAY", "PSR", "PHASE SEQ")):
        return "relays"
    if any(k in d for k in ("TRANSFER", "MTS", "ATS")):
        return "transfer_switches"
    if any(k in d for k in ("DISCONNECT", "SWITCH")):
        return "disconnect_switches"
    if any(k in d for k in ("TRANSFORMER", "XFMR")):
        return "transformers"
    if any(k in d for k in ("MOTOR",)):
        return "motors"
    if any(k in d for k in ("SHUNT",)):
        return "shunts"
    if any(k in d for k in ("FUSE",)):
        return "fuses"
    if any(k in d for k in ("INTERLOCK", "MICRO", "KIRK")):
        return "interlocks"
    if any(k in d for k in ("CAMLOCK", "CAM LOCK")):
        return "camlocks"
    if any(k in d for k in ("DOOR", "PANEL")):
        return "enclosure_parts"
    if any(k in d for k in ("SURGE", "SPD", "TVS")):
        return "surge_protection"
    if any(k in d for k in ("NEUTRAL", "NCT", "CT ")):
        return "CTs"
    if any(k in d for k in ("EKIP", "SIGNALLING", "MODULE")):
        return "accessories"
    return "other"


def compare_boms(source_descs: list[str], candidate_descs: list[str]) -> dict:
    """Compare two BOM description lists by component type overlap."""
    if not source_descs or not candidate_descs:
        return {"overlap_pct": None, "matching_types": [], "source_types": [], "candidate_types": []}

    source_types = set(categorize_component(d) for d in source_descs)
    cand_types = set(categorize_component(d) for d in candidate_descs)

    # Remove 'other' from overlap calculation since it's too generic
    source_meaningful = source_types - {"other"}
    cand_meaningful = cand_types - {"other"}

    if not source_meaningful:
        return {"overlap_pct": None, "matching_types": [], "source_types": list(source_types), "candidate_types": list(cand_types)}

    overlap = source_meaningful & cand_meaningful
    overlap_pct = round(len(overlap) / len(source_meaningful) * 100)

    return {
        "overlap_pct": overlap_pct,
        "matching_types": sorted(overlap),
        "source_types": sorted(source_meaningful),
        "candidate_types": sorted(cand_meaningful),
    }


# ── Main ──────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Find reference jobs")
    parser.add_argument("--job", required=True, help="Job number (e.g. J15874)")
    parser.add_argument("--jobs-root", help="Override jobs root folder")
    parser.add_argument("--quote-index", help="Override Quote Index path")
    parser.add_argument("--max-results", type=int, default=5,
                        help="Max validated candidates to return (BOM images extracted for each)")
    parser.add_argument("--refresh-cache", action="store_true",
                        help="Ignore the CAD cache and re-scan the filesystem for every candidate")
    args = parser.parse_args()

    global JOBS_ROOT, QUOTE_INDEX, _CAD_CACHE
    if args.jobs_root:
        JOBS_ROOT = Path(args.jobs_root)
    if args.quote_index:
        QUOTE_INDEX = Path(args.quote_index)

    # Load CAD cache (empty if --refresh-cache)
    if args.refresh_cache:
        _CAD_CACHE = {}
    else:
        load_cad_cache()

    # Step 1: Find job folder and read PRF
    job_folder = find_job_folder(args.job)
    prf_path = find_prf(job_folder)
    prf_data = read_prf(prf_path)
    parsed_model = parse_model_number(prf_data["model_no"])

    # Step 2: Search Quote Index
    candidates = search_quote_index(prf_data, parsed_model, args.job)

    if not candidates:
        die("NO_MATCHES", f"No matching jobs found for {args.job} "
            f"(family={parsed_model['family']}, config={parsed_model['config']}, "
            f"amp={parsed_model['amperage']}, size={prf_data['enclosure_size']})")

    # Output directory for BOM images
    output_dir = Path(__file__).resolve().parent.parent / "output" / args.job
    output_dir.mkdir(parents=True, exist_ok=True)

    # Extract source BOM image if DWG-embedded
    source_bom_image = extract_bom_images(job_folder, args.job, output_dir)
    source_descs = read_bom_descriptions(job_folder)

    # Step 3: Validate candidates — check CAD folder has content (early exit at max_results)
    validated = []
    for cand in candidates:
        if len(validated) >= args.max_results:
            break  # early exit — we have enough validated candidates

        cand_folder = find_candidate_folder(cand["job_number"])
        if cand_folder is None:
            cand["has_folder"] = False
            cand["has_cad"] = False
            cand["bom_comparison"] = {"overlap_pct": None, "matching_types": []}
            continue

        cand["has_folder"] = True
        cand["has_cad"] = has_cad_content_cached(cand["job_number"], cand_folder)
        cand["folder_path"] = str(cand_folder)

        if not cand["has_cad"]:
            continue

        # Step 4: BOM comparison
        cand_descs = read_bom_descriptions(cand_folder)
        cand["bom_comparison"] = compare_boms(source_descs, cand_descs)

        # Extract candidate BOM image if DWG-embedded — shared cache handles reuse
        cand_bom_image = extract_bom_images(cand_folder, cand["job_number"], output_dir)
        if cand_bom_image:
            cand["bom_image"] = cand_bom_image

        validated.append(cand)

    # Only show candidates with populated CAD folders
    if not validated:
        pass  # no fallback — empty list means no usable references

    # Limit results
    validated = validated[:args.max_results]

    # Build output
    output = {
        "source_job": {
            "job_number": args.job,
            "job_name": prf_data["job_name"],
            "model_no": prf_data["model_no"],
            "enclosure_size": prf_data["enclosure_size"],
            "current": prf_data["current"],
            "parsed": parsed_model,
            "bom_image": source_bom_image,
        },
        "candidates": validated,
        "total_index_matches": len(candidates),
    }

    # Persist CAD cache
    save_cad_cache()

    json.dump(output, sys.stdout, indent=2, default=str)


if __name__ == "__main__":
    main()
