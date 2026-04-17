"""
Doc Prep — Standalone headless script for FoxFab manufacturing packet preparation.

Usage (called by Claude or run directly):
    python docprep.py --job J16204
    python docprep.py --job J16204 --print
    python docprep.py --job J16204 --ref J15302          # Vikram workflow
    python docprep.py --job J16204 --ref J15302 --print

Without --print, generates PDFs in simulation mode (timestamped output folder).
With --print, sends to the preferred printer after generating.
"""

import argparse
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
import traceback
from datetime import datetime
from pathlib import Path

# ── Configuration ──────────────────────────────────────────────────────
JOBS_ROOT          = Path(r"Z:\FOXFAB_DATA\ENGINEERING\2 JOBS")
PREFERRED_PRINTER  = r"\\NPSVR05\FoxFab (Konica Bizhub C360i)"
BOM_SHEET_NAME     = "FFMPL"
EXCEL_EXTENSIONS   = {".xlsx", ".xls", ".xlsm"}
PDF_EXTENSIONS     = {".pdf", ".PDF"}

ACROBAT_SEARCH_PATHS = [
    r"C:\Program Files (x86)\Adobe\Acrobat 2017\Acrobat\Acrobat.exe",
    r"C:\Program Files\Adobe\Acrobat 2017\Acrobat\Acrobat.exe",
    r"C:\Program Files (x86)\Adobe\Acrobat DC\Acrobat\Acrobat.exe",
    r"C:\Program Files\Adobe\Acrobat DC\Acrobat\Acrobat.exe",
    r"C:\Program Files (x86)\Adobe\Acrobat 2020\Acrobat\Acrobat.exe",
    r"C:\Program Files\Adobe\Acrobat 2020\Acrobat\Acrobat.exe",
]

# FWO overlay positions (PDF points)
FWO_JOB_NO_X    = 165
FWO_JOB_NO_Y    = 145
FWO_JOB_NAME_X  = 165
FWO_JOB_NAME_Y  = 165
FWO_DATE_X      = 165
FWO_DATE_Y      = 182
FWO_ENCLOSURE_X = 165
FWO_ENCLOSURE_Y = 210
FWO_UNITS_X     = 165
FWO_UNITS_Y     = 245
FWO_FONT_SIZE   = 11

# ── Dependency checks ─────────────────────────────────────────────────
_missing = []

try:
    import win32print
    import win32com.client
except ImportError:
    win32print = None
    win32com = None
    _missing.append("pywin32  (pip install pywin32)")

try:
    from pypdf import PdfReader, PdfWriter
except ImportError:
    PdfReader = PdfWriter = None
    _missing.append("pypdf  (pip install pypdf)")

try:
    import fitz as _fitz
except ImportError:
    _fitz = None
    _missing.append("PyMuPDF  (pip install pymupdf)")

try:
    import openpyxl as _openpyxl
except ImportError:
    _openpyxl = None
    _missing.append("openpyxl  (pip install openpyxl)")

try:
    import xlwings as xw
except ImportError:
    xw = None
    _missing.append("xlwings  (pip install xlwings)")

# ── Stock check config ───────────────────────────────────────────────
STOCK_PARTS_FOLDER = r"Z:\FOXFAB_DATA\ENGINEERING\0 PRODUCTS\300 Stock Parts\PDFs & Flats"
NA_PREFIXES = ("245-", "295-")   # flexibars, fishpaper — mark N/A if no CNC match


def _find_es_exe() -> str:
    """Locate es.exe from known project paths or system PATH."""
    script_dir = Path(__file__).resolve().parent
    project_root = script_dir.parents[3]
    candidates = [
        project_root / "tools" / "BomFiller" / "es.exe",
        project_root / "dist" / "_internal" / "es.exe",
        script_dir / "es.exe",
    ]
    if getattr(sys, "frozen", False):
        candidates.insert(0, Path(sys._MEIPASS) / "es.exe")
    for p in candidates:
        if p.is_file():
            return str(p)
    return "es.exe"


_ES_EXE = _find_es_exe()


def _run_es(args: list[str]) -> str | None:
    """Run es.exe with the given args. Returns stdout or None on timeout."""
    try:
        result = subprocess.run(
            [_ES_EXE] + args, capture_output=True, text=True, timeout=10)
        return result.stdout.strip()
    except FileNotFoundError:
        warn(f"es.exe not found at '{_ES_EXE}' — stock check unavailable.")
        return None
    except subprocess.TimeoutExpired:
        return None


_ES_AVAILABLE = None
_STOCK_FILE_CACHE: list[str] | None = None


def _stock_fallback_scan(part_number: str) -> bool:
    global _STOCK_FILE_CACHE
    if _STOCK_FILE_CACHE is None:
        _STOCK_FILE_CACHE = []
        try:
            root = Path(STOCK_PARTS_FOLDER)
            if root.is_dir():
                for p in root.iterdir():
                    if p.is_file():
                        _STOCK_FILE_CACHE.append(p.name.lower())
                    elif p.is_dir():
                        try:
                            for c in p.iterdir():
                                if c.is_file():
                                    _STOCK_FILE_CACHE.append(c.name.lower())
                        except Exception:
                            pass
        except Exception:
            pass
    pn_lc = part_number.lower()
    if any(pn_lc in name for name in _STOCK_FILE_CACHE):
        return True
    base = re.sub(r'_\d+$', '', part_number).lower()
    if base != pn_lc and any(base in name for name in _STOCK_FILE_CACHE):
        return True
    return False


def check_stock(part_number: str) -> bool:
    """Check if a part number exists in the stock parts folder via Everything Search."""
    global _ES_AVAILABLE
    if _ES_AVAILABLE is None:
        try:
            subprocess.run([_ES_EXE, "-h"], capture_output=True, text=True, timeout=5)
            _ES_AVAILABLE = True
        except (FileNotFoundError, OSError):
            _ES_AVAILABLE = False
            warn("Everything Search unavailable — using slower direct folder scan for stock check.")
        except subprocess.TimeoutExpired:
            _ES_AVAILABLE = True
        except Exception:
            _ES_AVAILABLE = False
            warn("Everything Search unavailable — using slower direct folder scan for stock check.")

    if not _ES_AVAILABLE:
        return _stock_fallback_scan(part_number)

    output = _run_es(["-path", STOCK_PARTS_FOLDER, part_number])
    if output is None:
        return False
    if any(line.strip() for line in output.splitlines()):
        return True
    # Try without config suffix (_digits)
    base = re.sub(r'_\d+$', '', part_number)
    if base != part_number:
        output = _run_es(["-path", STOCK_PARTS_FOLDER, base])
        if output and any(line.strip() for line in output.splitlines()):
            return True
    return False


# ── CNC parsing regexes ───────────────────────────────────────────────
_CNC_REV_RE          = re.compile(r"\s*r[A-Za-z]$", re.IGNORECASE)
_CNC_DIRECT_PREFIXES = re.compile(r"^(240|250|200|295|210|220|230|260|270|280|290)-(\d+)")
_CNC_USCORE_PREFIXES = re.compile(r"^(240|250|200|295|210|220|230|260|270|280|290)_(\d+)")
_CNC_BARE_DIGITS     = re.compile(r"^(\d{4,5})")
_CNC_J_PREFIX        = re.compile(r"^J\d+", re.IGNORECASE)
_CNC_DRAWING_NO_RE   = re.compile(
    r"DRAWING\s+NUMBER\s*:\s*([0-9]{3}-[0-9]+(?:\s*r[A-Za-z])?)", re.IGNORECASE)


# ── Logging ────────────────────────────────────────────────────────────
_log_lines: list[str] = []


def log(msg: str):
    stamp = datetime.now().strftime("%H:%M:%S")
    line = f"[{stamp}] {msg}"
    _log_lines.append(line)
    print(line)


def warn(msg: str):
    log(f"[WARN] {msg}")


def error(msg: str):
    log(f"[ERROR] {msg}")


# ── Helpers ────────────────────────────────────────────────────────────
def safe_name(text: str) -> str:
    if not text:
        return "UnknownJob"
    # Normalize unicode, strip accents, remove illegal path chars
    import unicodedata
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    return re.sub(r'[\\/:*?"<>|]+', "_", text).strip() or "UnknownJob"


def list_files(folder: Path) -> list[Path]:
    return sorted([p for p in folder.iterdir() if p.is_file()],
                  key=lambda p: p.name.lower())


def find_job_folder(job_number: str) -> Path:
    """Find a job folder in JOBS_ROOT by matching the job number prefix."""
    if not JOBS_ROOT.is_dir():
        raise RuntimeError(f"Jobs root not found: {JOBS_ROOT}")

    if not re.match(r"^J\d{5}$", job_number, re.IGNORECASE):
        raise RuntimeError("INVALID_JOB: must be J##### with 5 digits")

    job_up = job_number.upper()
    _job_re = re.compile(rf"^{re.escape(job_up)}(?:[ \-]|$)")
    matches = [d for d in JOBS_ROOT.iterdir()
               if d.is_dir() and _job_re.match(d.name.upper())]

    if len(matches) == 0:
        raise RuntimeError(f"No job folder found matching '{job_number}' in {JOBS_ROOT}")
    if len(matches) == 1:
        log(f"Found job folder: {matches[0].name}")
        return matches[0]

    # Multiple matches — print them and let caller decide
    raise RuntimeError(
        f"Multiple folders match '{job_number}':\n"
        + "\n".join(f"  {i+1}. {m.name}" for i, m in enumerate(matches))
        + "\nPlease specify which one to use.")


# ── Folder structure resolution ────────────────────────────────────────
def _has_mech_subs(p: Path) -> bool:
    return all((p / s).is_dir() for s in
               ["204 BOM", "205 CNC", "202 PDFs_Flats", "203 Assemblies"])


def _is_numbered_variant(p: Path) -> bool:
    return p.is_dir() and re.search(r"-\d{2}$", p.name) is not None


def _resolve_internal(folder: Path) -> Path | None:
    """Check if folder contains named subfolders; if so, return Internal."""
    internal = folder / "Internal"
    if internal.is_dir() and _has_mech_subs(internal):
        log(f"Found Internal subfolder in {folder.name}, using it")
        return internal
    # Check other named subfolders that have mech subs
    for child in sorted(folder.iterdir(), key=lambda p: p.name.lower()):
        if child.is_dir() and _has_mech_subs(child):
            # Not Internal but has the right structure — flag for user decision
            return None
    return None


def _build_dir_tree(folder: Path, depth: int = 3, prefix: str = "") -> str:
    """Build an indented directory tree string for display."""
    if depth <= 0 or not folder.is_dir():
        return ""
    lines = []
    try:
        children = sorted(folder.iterdir(), key=lambda p: p.name.lower())
    except PermissionError:
        return f"{prefix}(permission denied)\n"
    dirs = [c for c in children if c.is_dir()]
    for i, child in enumerate(dirs):
        is_last = (i == len(dirs) - 1)
        connector = "└── " if is_last else "├── "
        tag = ""
        if _has_mech_subs(child):
            tag = "  [has 204 BOM, 205 CNC, etc.]"
        lines.append(f"{prefix}{connector}{child.name}{tag}")
        extension = "    " if is_last else "│   "
        subtree = _build_dir_tree(child, depth - 1, prefix + extension)
        if subtree:
            lines.append(subtree.rstrip("\n"))
    return "\n".join(lines)


def get_context(job_folder: Path) -> dict:
    """Resolve folder structure and return job_root, mech_roots, variant_only."""

    # Pattern 1: user selected a numbered variant directly
    if _is_numbered_variant(job_folder) and _has_mech_subs(job_folder):
        return {"job_root": job_folder.parent.parent,
                "mech_roots": [job_folder], "variant_only": True}

    # Also check if numbered variant has Internal subfolder
    if _is_numbered_variant(job_folder):
        internal = _resolve_internal(job_folder)
        if internal:
            return {"job_root": job_folder.parent.parent,
                    "mech_roots": [internal], "variant_only": True}

    mech = job_folder / "200 Mech"
    if not mech.is_dir():
        raise RuntimeError("Missing folder: 200 Mech")

    # Pattern 2: standard direct structure
    if _has_mech_subs(mech):
        return {"job_root": job_folder, "mech_roots": [mech], "variant_only": False}

    # Pattern 3: numbered variants inside 200 Mech
    variants = []
    for child in sorted(mech.iterdir(), key=lambda p: p.name.lower()):
        if not child.is_dir():
            continue
        if _is_numbered_variant(child):
            # Check if variant itself has mech subs
            if _has_mech_subs(child):
                variants.append(child)
            else:
                # Pattern 4b: Internal inside a numbered variant
                internal = _resolve_internal(child)
                if internal:
                    variants.append(internal)

    if variants:
        return {"job_root": job_folder, "mech_roots": variants, "variant_only": False}

    # Pattern 4a: named subfolders directly in 200 Mech (no numbered variants)
    internal = _resolve_internal(mech)
    if internal:
        return {"job_root": job_folder, "mech_roots": [internal], "variant_only": False}

    # Check for any subfolder with mech subs (unknown pattern)
    named_subs = []
    for child in sorted(mech.iterdir(), key=lambda p: p.name.lower()):
        if child.is_dir() and _has_mech_subs(child):
            named_subs.append(child)
    if named_subs:
        tree = _build_dir_tree(mech, depth=3)
        raise RuntimeError(
            f"UNKNOWN_STRUCTURE:Found subfolders with mechanical structure but no 'Internal' folder.\n"
            f"Available subfolders with BOM/CNC/Flats/Assemblies:\n"
            + "\n".join(f"  - {s.name}" for s in named_subs)
            + f"\n\nDirectory tree:\n{tree}")

    tree = _build_dir_tree(mech, depth=3)
    raise RuntimeError(
        f"UNKNOWN_STRUCTURE:Could not find a usable mechanical folder structure.\n"
        f"Directory tree:\n{tree}")


# ── File matching ──────────────────────────────────────────────────────
def match_fwo(folder: Path) -> Path | None:
    for f in list_files(folder):
        if f.suffix in PDF_EXTENSIONS and f.stem == "Fabrication Work Order - Standard v1.0":
            return f
    if folder.is_dir():
        names = [f.name for f in list_files(folder)
                 if f.suffix in PDF_EXTENSIONS and f.stem.startswith("Fabrication Work Order")]
        if names:
            warn(f"FWO not found at exact 'Fabrication Work Order - Standard v1.0'. Similar files exist: {names}. Update the script if the template version changed.")
    warn("Fabrication Work Order PDF not found in 300 Inputs.")
    return None


def _revision_key(p: Path) -> tuple:
    """Parse r<Letter><digits?> (e.g. rA, rA1, rA2, rB1) -> (letter_ord, sub_int)."""
    m = re.search(r'r([A-Za-z])(\d*)', p.stem)
    if not m:
        return (0, 0)
    letter = m.group(1).upper()
    sub = int(m.group(2)) if m.group(2) else 0
    return (ord(letter), sub)


def match_excel(folder: Path, token: str, title: str) -> Path | None:
    if not folder.is_dir():
        warn(f"{title} folder not found: {folder}")
        return None
    matches = [f for f in list_files(folder)
               if f.suffix.lower() in {e.lower() for e in EXCEL_EXTENSIONS}
               and token.lower() in f.name.lower()]
    if not matches:
        # BOM fallback: scan .xlsx files for an "FFMPL" sheet
        if title == "BOM" or token.upper() == "BOM":
            try:
                from openpyxl import load_workbook
            except Exception as e:
                warn(f"openpyxl unavailable for BOM fallback: {e}")
                load_workbook = None
            if load_workbook is not None:
                xlsx_files = [f for f in list_files(folder)
                              if f.suffix.lower() == ".xlsx"]
                ffmpl_matches = []
                for xf in xlsx_files:
                    try:
                        wb = load_workbook(xf, read_only=True, data_only=True)
                        try:
                            if "FFMPL" in wb.sheetnames:
                                ffmpl_matches.append(xf)
                        finally:
                            try:
                                wb.close()
                            except Exception:
                                pass
                    except Exception as e:
                        warn(f"Could not inspect {xf.name} for FFMPL sheet: {e}")
                        continue
                if ffmpl_matches:
                    if len(ffmpl_matches) > 1:
                        selected = max(ffmpl_matches, key=_revision_key)
                    else:
                        selected = ffmpl_matches[0]
                    log(f"BOM filename did not contain 'BOM' — fell back to FFMPL sheet detection: {selected.name}")
                    return selected
        warn(f"No Excel file containing '{token}' found in {title}.")
        return None
    if len(matches) == 1:
        return matches[0]
    # Auto-select latest revision
    with_rev = [(f, _revision_key(f)) for f in matches]
    if any(rev != (0, 0) for _, rev in with_rev):
        selected = max(with_rev, key=lambda x: x[1])[0]
        log(f"Auto-selected {title} by revision: {selected.name}")
        return selected
    # Multiple matches, no revision info — pick first and warn
    warn(f"Multiple {title} files found, using first: {matches[0].name}")
    return matches[0]


def _pack_revision_key(p: Path) -> tuple:
    """Extract a sortable revision key from a pack PDF filename.
    Handles patterns like 'R1.3', 'R1.4', 'Rev A', 'rB', etc."""
    name = p.stem.upper()
    # Try R<major>.<minor> pattern (e.g., R1.3, R1.4)
    m = re.search(r"R(\d+)\.(\d+)", name)
    if m:
        return (int(m.group(1)), int(m.group(2)))
    # Try R<number> pattern (e.g., R3)
    m = re.search(r"R(\d+)", name)
    if m:
        return (int(m.group(1)), 0)
    # Try rA, rB letter revisions
    m = re.search(r"R([A-Z])", name)
    if m:
        return (0, ord(m.group(1)))
    # No revision found — lowest priority
    return (0, 0)


def match_pack(folder: Path, model_no: str = "", pack_override: str | None = None) -> Path | None:
    if pack_override:
        p = Path(pack_override)
        if p.is_file():
            log(f"Using specified pack: {p.name}")
            return p
        warn(f"Specified pack not found: {pack_override}")
        return None
    if not folder.is_dir():
        warn(f"Electrical drawings folder not found: {folder}")
        return None
    # Only search direct files in the folder (no subfolders)
    all_packs = [f for f in list_files(folder)
                 if f.suffix in PDF_EXTENSIONS and "PACK" in f.name.upper()]
    if not all_packs:
        warn("No PDF with 'PACK' in name found in Electrical Drawings.")
        return None
    if model_no:
        model_up = model_no.upper()
        model_matches = [f for f in all_packs if model_up in f.name.upper()]
        if len(model_matches) == 1:
            return model_matches[0]
        if model_matches:
            # Multiple matches — pick highest revision
            best = max(model_matches, key=_pack_revision_key)
            log(f"Multiple PACK PDFs match model '{model_no}', using highest revision: {best.name}")
            return best
    if len(all_packs) == 1:
        return all_packs[0]
    # Multiple packs, no model filter — pick highest revision
    best = max(all_packs, key=_pack_revision_key)
    log(f"Multiple PACK PDFs found, using highest revision: {best.name}")
    return best


def match_cnc(folder: Path) -> list[Path]:
    if not folder.is_dir():
        warn(f"CNC folder not found: {folder}")
        return []
    pdfs = [f for f in list_files(folder) if f.suffix in PDF_EXTENSIONS]
    if not pdfs:
        warn("No PDFs found in CNC folder.")
    return sorted(pdfs, key=lambda p: p.name.lower())


def match_flats(folder: Path) -> list[Path]:
    if not folder.is_dir():
        warn(f"PDFs_Flats folder not found: {folder}")
        return []
    pdfs = [f for f in list_files(folder) if f.suffix in PDF_EXTENSIONS]
    if not pdfs:
        warn("No PDFs found in PDFs_Flats folder.")
    return sorted(pdfs, key=lambda p: p.name.lower())


def match_assemblies(folder: Path) -> tuple[list[Path], list[Path]]:
    if not folder.is_dir():
        warn(f"Assemblies folder not found: {folder}")
        return [], []
    pdfs, lay = [], []
    for f in list_files(folder):
        if f.suffix not in PDF_EXTENSIONS:
            continue
        if f.stem.upper().endswith(("-LAY", "- LAY")):
            lay.append(f)
        else:
            pdfs.append(f)
    if not pdfs:
        warn("No printable PDFs found in Assemblies folder.")
    return (sorted(pdfs, key=lambda p: p.name.lower()),
            sorted(lay, key=lambda p: p.name.lower()))


def find_variant_prf(prf_folder: Path, mechs: list[Path]) -> Path | None:
    if not prf_folder.is_dir():
        warn(f"PRF folder not found: {prf_folder}")
        return None
    all_prfs = [f for f in list_files(prf_folder)
                if f.suffix.lower() in {e.lower() for e in EXCEL_EXTENSIONS}
                and "prf" in f.name.lower()]
    if not all_prfs:
        warn(f"No PRF file found in {prf_folder}")
        return None

    variant_suffix = ""
    if len(mechs) == 1:
        m = re.search(r"(-\d{2})$", mechs[0].name)
        if m:
            variant_suffix = m.group(1)

    if variant_suffix:
        variant_prfs = [f for f in all_prfs if variant_suffix in f.name]
        if len(variant_prfs) == 1:
            return variant_prfs[0]
        if variant_prfs:
            if len(variant_prfs) > 1:
                warn(f"Multiple PRFs match variant {variant_suffix}: {[f.name for f in variant_prfs]} — using {variant_prfs[0].name}")
            return variant_prfs[0]

    if len(all_prfs) == 1:
        return all_prfs[0]
    warn(f"Multiple PRF files found, using first: {all_prfs[0].name}")
    return all_prfs[0]


def classify_cnc(p: Path) -> str:
    base = p.stem
    if re.match(r"^[Jj]", base) or re.match(r"^\d{3}-", base):
        return "duplex"
    return "simplex"


# ── PRF reading ────────────────────────────────────────────────────────
def read_prf(prf_path: Path) -> dict | None:
    if _openpyxl is None:
        warn("openpyxl not installed — cannot read PRF data.")
        return None
    try:
        wb = _openpyxl.load_workbook(str(prf_path), data_only=True)
        ws = None
        for _sn in ("Form", "PRF", "PRF Form", "Production Release"):
            if _sn in wb.sheetnames:
                ws = wb[_sn]
                break
        if ws is None:
            ws = wb.worksheets[0]
            warn(f"PRF sheet 'Form' not found, falling back to '{ws.title}' — verify FWO data!")

        def cell(ref: str) -> str:
            v = ws[ref].value
            return str(v).strip() if v is not None else ""

        job_no   = cell("C4")
        model_no = cell("G9")
        job_name = cell("C8")
        size     = cell("G18")
        material = cell("G19")
        rating   = cell("G20")
        qty      = cell("G22")

        MATERIAL_MAP = {
            "aluminum": "ALU", "aluminium": "ALU",
            "stainless": "SS",
            "carbon steel": "CS",
            "galvanized": "GALV", "galvanised": "GALV",
            "mild steel": "MS",
            "cold rolled": "CRS",
            "hot rolled": "HRS",
        }
        RATING_MAP = {
            "type 3r": "N3R", "type3r": "N3R",
            "type 4x": "N4X", "type4x": "N4X",
            "type 4": "N4",
            "type 12": "N12",
            "type 1": "N1",
            "nema 3r": "N3R",
            "nema 4x": "N4X",
            "nema 4": "N4",
            "nema 12": "N12",
        }

        mat_lc = material.lower()
        mat_abbr = None
        for _k, _v in MATERIAL_MAP.items():
            if _k in mat_lc:
                mat_abbr = _v
                break
        if mat_abbr is None:
            mat_abbr = material

        rat_lc = rating.lower().strip()
        rat_abbr = None
        for _k, _v in RATING_MAP.items():
            if _k in rat_lc:
                rat_abbr = _v
                break
        if rat_abbr is None:
            rat_abbr = rating

        if (material and mat_abbr == material) or (rating and rat_abbr == rating):
            warn(f"Unknown material/rating: '{material}' / '{rating}' — using raw text. Add to MATERIAL_MAP/RATING_MAP if needed.")

        enclosure = " ".join(p for p in [size, mat_abbr, rat_abbr] if p)

        return {
            "job_no": job_no, "model_no": model_no, "job_name": job_name,
            "enclosure": enclosure, "qty": qty,
            "size": size, "material": material, "rating": rating,
        }
    except Exception as e:
        warn(f"Failed to read PRF: {e}")
        return None


# ── FWO auto-fill ──────────────────────────────────────────────────────
def fill_fwo(fwo_path: Path, prf_data: dict) -> Path | None:
    if _fitz is None:
        warn("PyMuPDF not installed — cannot auto-fill FWO.")
        return None
    try:
        today = datetime.now().strftime("%B %d, %Y")
        fields = [
            (FWO_JOB_NO_X,    FWO_JOB_NO_Y,    prf_data["job_no"]),
            (FWO_JOB_NAME_X,  FWO_JOB_NAME_Y,  prf_data["job_name"]),
            (FWO_DATE_X,      FWO_DATE_Y,       today),
            (FWO_ENCLOSURE_X, FWO_ENCLOSURE_Y,  prf_data["enclosure"]),
            (FWO_UNITS_X,     FWO_UNITS_Y,      prf_data["qty"]),
        ]
        doc = _fitz.open(str(fwo_path))
        page = doc[0]
        for x, y, text in fields:
            if text:
                page.insert_text(
                    _fitz.Point(x, y), text,
                    fontsize=FWO_FONT_SIZE, fontname="helv", color=(0, 0, 0))
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix="_FWO_filled.pdf")
        tmp_path = Path(tmp.name)
        tmp.close()
        doc.save(str(tmp_path))
        doc.close()
        log(f"FWO filled: {prf_data['job_no']}  {prf_data['enclosure']}")
        return tmp_path
    except Exception as e:
        warn(f"FWO fill failed: {e}")
        return None


# ── CNC column marking ────────────────────────────────────────────────
def _cnc_strip_rev(s: str) -> str:
    return _CNC_REV_RE.sub("", s).strip()


def _cnc_parts_from_rest(prefix: str, rest: str) -> list[str]:
    parts = []
    for seg in rest.split("_"):
        seg = seg.strip()
        if not seg:
            continue
        m = _CNC_BARE_DIGITS.match(seg)
        if m:
            parts.append(f"{prefix}-{m.group(1)}")
        else:
            break
    return parts


def _cnc_extract_galv_parts(pdf_path: Path) -> list[str]:
    if _fitz is None:
        return []
    parts = []
    try:
        doc = _fitz.open(str(pdf_path))
        for page in doc:
            text = page.get_text()
            for m in _CNC_DRAWING_NO_RE.finditer(text):
                raw = m.group(1).strip()
                pn = _cnc_strip_rev(raw)
                if pn and pn not in parts:
                    parts.append(pn)
        doc.close()
    except Exception:
        pass
    return parts


def _cnc_parse_filename(pdf_path: Path) -> list[str]:
    stem = pdf_path.stem
    stem_clean = _CNC_REV_RE.sub("", stem).strip()

    if _CNC_J_PREFIX.match(stem_clean):
        parts = _cnc_extract_galv_parts(pdf_path)
        if parts:
            log(f"  {pdf_path.name}  ->  GALV  {parts}")
        else:
            warn(f"  {pdf_path.name}  ->  GALV  (no DRAWING NUMBER lines found)")
        return parts

    m = _CNC_DIRECT_PREFIXES.match(stem_clean)
    if m:
        prefix = m.group(1)
        first_num = m.group(2)
        first_pn = f"{prefix}-{first_num}"
        rest = stem_clean[m.end():]
        extra = _cnc_parts_from_rest(prefix, rest)
        parts = [first_pn] + extra
        log(f"  {pdf_path.name}  ->  {parts}")
        return parts

    # Underscore-separated prefix (e.g., 295_70067_70070_LEXAN.pdf)
    m = _CNC_USCORE_PREFIXES.match(stem_clean)
    if m:
        prefix = m.group(1)
        first_num = m.group(2)
        first_pn = f"{prefix}-{first_num}"
        rest = stem_clean[m.end():]
        extra = _cnc_parts_from_rest(prefix, rest)
        parts = [first_pn] + extra
        log(f"  {pdf_path.name}  ->  {parts}")
        return parts

    bare_m = _CNC_BARE_DIGITS.match(stem_clean)
    if bare_m:
        digits = bare_m.group(1)
        pn = f"240-{digits}"
        log(f"  {pdf_path.name}  ->  {[pn]}")
        return [pn]

    warn(f"  {pdf_path.name}  ->  unrecognised pattern — skipped")
    return []


def cnc_mark_bom(bom_path: Path, cnc_folder: Path) -> tuple[int, list[tuple[int, str]], int, int, list[tuple[int, str]], list[tuple[int, str]]]:
    """Mark CNC column (H) in BOM workbook.
    Returns (matched_count, unmatched_parts, total_parts, stock_count,
             na_marked_parts, stock_found_parts).
    Each part list contains (row_number, part_number) tuples."""
    if xw is None:
        warn("xlwings not installed — skipping CNC column marking.")
        return 0, [], 0, 0, [], []

    DATA_START = 6

    # Fix #10: Abort if Excel is already running
    try:
        import pywintypes
        try:
            win32com.client.GetActiveObject("Excel.Application")
            raise RuntimeError("EXCEL_RUNNING: Please close all Excel windows before running doc prep — we don't want to risk closing your unsaved work.")
        except pywintypes.com_error:
            pass
    except ImportError:
        pass

    # Fix #1: Lock check + backup
    try:
        with open(bom_path, 'a+b'):
            pass
    except PermissionError:
        raise RuntimeError(f"BOM_LOCKED: {bom_path}")
    log(f"BOM lock check passed: {bom_path}")
    try:
        bak_path = bom_path.with_suffix(bom_path.suffix + ".bak")
        shutil.copy2(str(bom_path), str(bak_path))
        log(f"BOM backup created: {bak_path}")
    except Exception as _e:
        warn(f"BOM backup failed: {_e}")

    xw_app = xw.App(visible=False, add_book=False)
    xw_app.display_alerts = False
    xw_app.screen_updating = False
    try:
        wb = xw_app.books.open(str(bom_path))
        sheet_names = [s.name for s in wb.sheets]
        if BOM_SHEET_NAME not in sheet_names:
            warn(f"Sheet '{BOM_SHEET_NAME}' not found. Available: {sheet_names}")
            wb.close()
            return 0, [], 0, 0, [], []

        ws = wb.sheets[BOM_SHEET_NAME]
        last_row = ws.range(f"A{DATA_START}").end("down").row
        if last_row > 1_000_000:
            last_row = DATA_START

        # Fix #15: Warn on mid-list blank rows
        try:
            used_last = ws.used_range.last_cell.row
            if used_last > last_row + 2:
                first_nonempty = None
                for _r in range(last_row + 1, used_last + 1):
                    _v = ws.range(f"A{_r}").value
                    if _v and str(_v).strip():
                        first_nonempty = _r
                        break
                if first_nonempty:
                    warn(f"BOM has data below row {last_row} (blank row truncated scan). First non-empty: row {first_nonempty}. Check the BOM for missing parts.")
        except Exception:
            pass

        col_a = ws.range(f"A{DATA_START}:A{last_row}").value
        col_h = ws.range(f"H{DATA_START}:H{last_row}").value
        if not isinstance(col_a, list):
            col_a = [col_a]
        if not isinstance(col_h, list):
            col_h = [col_h]

        cnc_pdfs = [f for f in cnc_folder.iterdir()
                    if f.is_file() and f.suffix.lower() == ".pdf"
                    and "merged" not in f.name.lower()]
        log(f"Found {len(cnc_pdfs)} CNC PDFs in {cnc_folder.name}")

        cnc_parts: dict[str, Path] = {}
        for pdf in cnc_pdfs:
            for pn in _cnc_parse_filename(pdf):
                cnc_parts[pn] = pdf

        log(f"Resolved {len(cnc_parts)} unique part numbers from CNC folder")

        matched, unmatched = [], []
        na_marked, stock_found = [], []
        stock_existing, cnc_existing, na_existing = 0, 0, 0
        total_parts = 0
        # Track unmatched rows for second pass (N/A + stock)
        unmatched_rows: list[tuple[int, str, str]] = []  # (row_idx, pn_str, pn_clean)

        for i, pn_raw in enumerate(col_a):
            if not pn_raw or not str(pn_raw).strip():
                continue
            pn_str = str(pn_raw).strip()
            total_parts += 1
            row_idx = DATA_START + i
            h_val = col_h[i] if i < len(col_h) else None
            h_upper = str(h_val).strip().upper() if h_val else ""
            if h_upper == "S":
                stock_existing += 1
                continue
            if h_upper == "X":
                cnc_existing += 1
                continue
            if h_upper == "N/A":
                na_existing += 1
                continue
            pn_clean = _cnc_strip_rev(pn_str)
            if pn_clean in cnc_parts:
                ws.range(f"H{row_idx}").value = "X"
                matched.append(row_idx)
                log(f"  row {row_idx}  {pn_str:<28}  -> X  ({cnc_parts[pn_clean].name})")
            else:
                unmatched_rows.append((row_idx, pn_str, pn_clean))

        # Second pass: N/A for 245-/295- prefixes, stock check for the rest
        truly_unmatched = []
        if unmatched_rows:
            log("Smart scan: checking unmatched parts...")
        for row_idx, pn_str, pn_clean in unmatched_rows:
            # Check N/A prefixes (flexibars, fishpaper)
            if any(pn_clean.startswith(pfx) for pfx in NA_PREFIXES):
                ws.range(f"H{row_idx}").value = "N/A"
                na_marked.append((row_idx, pn_str))
                log(f"  row {row_idx}  {pn_str:<28}  -> N/A  (no CNC needed)")
                continue
            # Stock check via Everything Search
            if check_stock(pn_clean):
                ws.range(f"H{row_idx}").value = "S"
                stock_found.append((row_idx, pn_str))
                log(f"  row {row_idx}  {pn_str:<28}  -> S  (stock part)")
                continue
            truly_unmatched.append((row_idx, pn_str))

        wb.save()
        wb.close()
        total_cnc = cnc_existing + len(matched)
        total_stock = stock_existing + len(stock_found)
        total_na = na_existing + len(na_marked)
        return (total_cnc, truly_unmatched, total_parts, total_stock,
                na_marked, stock_found)
    finally:
        try:
            xw_app.quit()
        except Exception:
            pass


# ── BOM Check (read-only) ──────────────────────────────────────────────
def bom_check(bom_path: Path, cnc_folder: Path) -> dict:
    """Read-only BOM vs CNC check. Returns a dict with matched, unmatched BOM parts,
    and orphaned CNC parts (in CNC but not in BOM). Does NOT modify the BOM."""
    if _openpyxl is None:
        raise RuntimeError("openpyxl not installed. pip install openpyxl")

    DATA_START = 6

    # Read BOM part numbers from column A, skip stock (column H = "S")
    wb = _openpyxl.load_workbook(str(bom_path), data_only=True)
    sheet_names = wb.sheetnames
    if BOM_SHEET_NAME not in sheet_names:
        raise RuntimeError(f"Sheet '{BOM_SHEET_NAME}' not found. Available: {sheet_names}")

    ws = wb[BOM_SHEET_NAME]
    bom_parts: list[tuple[int, str]] = []  # (row, part_number)
    for row in range(DATA_START, ws.max_row + 1):
        pn_raw = ws[f"A{row}"].value
        if not pn_raw or not str(pn_raw).strip():
            continue
        h_val = ws[f"H{row}"].value
        if h_val and str(h_val).strip().upper() == "S":
            continue
        bom_parts.append((row, _cnc_strip_rev(str(pn_raw).strip())))
    wb.close()

    # Parse CNC folder for part numbers
    cnc_pdfs = [f for f in cnc_folder.iterdir()
                if f.is_file() and f.suffix.lower() == ".pdf"
                and "merged" not in f.name.lower()]
    log(f"Found {len(cnc_pdfs)} CNC PDFs in {cnc_folder.name}")

    cnc_parts: dict[str, Path] = {}
    for pdf in cnc_pdfs:
        for pn in _cnc_parse_filename(pdf):
            cnc_parts[pn] = pdf

    log(f"Resolved {len(cnc_parts)} unique part numbers from CNC folder")

    # Two-way comparison
    matched = []
    unmatched_bom = []
    bom_pn_set = set()

    for row, pn in bom_parts:
        bom_pn_set.add(pn)
        if pn in cnc_parts:
            matched.append((row, pn, cnc_parts[pn].name))
        else:
            unmatched_bom.append((row, pn))

    orphaned_cnc = []
    for pn, pdf in sorted(cnc_parts.items(), key=lambda x: x[0]):
        if pn not in bom_pn_set:
            orphaned_cnc.append((pn, pdf.name))

    return {
        "bom_path": str(bom_path),
        "cnc_folder": str(cnc_folder),
        "total_bom_parts": len(bom_parts),
        "total_cnc_parts": len(cnc_parts),
        "matched": matched,
        "unmatched_bom": unmatched_bom,
        "orphaned_cnc": orphaned_cnc,
    }


def format_bom_check_report(result: dict) -> str:
    lines = [
        "=" * 60,
        "BOM CHECK REPORT",
        "=" * 60,
        "",
        f"BOM: {Path(result['bom_path']).name}",
        f"CNC Folder: {Path(result['cnc_folder']).name}",
        "",
        f"BOM parts (non-stock): {result['total_bom_parts']}",
        f"CNC part numbers:      {result['total_cnc_parts']}",
        f"Matched:               {len(result['matched'])}",
        f"BOM missing CNC:       {len(result['unmatched_bom'])}",
        f"CNC not in BOM:        {len(result['orphaned_cnc'])}",
    ]

    if result["matched"]:
        lines += ["", "-" * 60, "MATCHED (BOM row -> CNC file):", "-" * 60]
        for row, pn, cnc_file in result["matched"]:
            lines.append(f"  row {row:<4}  {pn:<28}  ->  {cnc_file}")

    if result["unmatched_bom"]:
        lines += ["", "-" * 60, "BOM PARTS MISSING CNC FILE:", "-" * 60]
        for row, pn in result["unmatched_bom"]:
            lines.append(f"  row {row:<4}  {pn}")

    if result["orphaned_cnc"]:
        lines += ["", "-" * 60, "CNC FILES NOT IN BOM:", "-" * 60]
        for pn, cnc_file in result["orphaned_cnc"]:
            lines.append(f"  {pn:<28}  ({cnc_file})")

    lines += ["", "=" * 60]
    return "\n".join(lines)


def run_bom_check(job_number: str) -> tuple[str, Path | None]:
    """Run a read-only BOM check for a job. Returns (report_text, report_file_path)."""
    if _missing:
        print("\n[WARNING] Missing dependencies:")
        for dep in _missing:
            print(f"  - {dep}")
        print()

    job_folder = find_job_folder(job_number)
    ctx = get_context(job_folder)
    mechs = ctx["mech_roots"]

    bom_path = match_excel(mechs[0] / "204 BOM", "BOM", "BOM")
    if not bom_path:
        raise RuntimeError("BOM file not found.")

    cnc_folder = mechs[0] / "205 CNC"
    if not cnc_folder.is_dir():
        raise RuntimeError(f"CNC folder not found: {cnc_folder}")

    log(f"BOM Check: {bom_path.name} vs {cnc_folder.name}")
    result = bom_check(bom_path, cnc_folder)
    report = format_bom_check_report(result)
    print("\n" + report)

    # Save report
    script_dir = Path(__file__).resolve().parent
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    job_name = safe_name(job_folder.name)
    report_path = script_dir / f"BOM_Check_{job_name}_{ts}.txt"
    report_path.write_text(report, encoding="utf-8")
    log(f"Report saved: {report_path}")

    return report, report_path


# ── PDF operations ─────────────────────────────────────────────────────
def merge_pdfs(files: list[Path], label: str) -> Path:
    if PdfReader is None:
        raise RuntimeError("pypdf not installed. pip install pypdf")
    writer = PdfWriter()
    for pdf in files:
        reader = PdfReader(str(pdf))
        for page in reader.pages:
            writer.add_page(page)
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=f"_{label}.pdf")
    tmp_path = Path(tmp.name)
    tmp.close()
    with open(tmp_path, "wb") as f:
        writer.write(f)
    return tmp_path


def save_pdf(src: Path, dest: Path, pages=None):
    if pages is not None:
        if PdfReader is None:
            raise RuntimeError("pypdf not installed. pip install pypdf")
        reader = PdfReader(str(src))
        writer = PdfWriter()
        for i in range(pages[0], pages[1] + 1):
            writer.add_page(reader.pages[i])
        with open(dest, "wb") as f:
            writer.write(f)
    else:
        shutil.copy2(str(src), str(dest))


def excel_to_pdf(file: Path, dest: Path, first_sheet_only: bool, excel=None,
                 orientation: int = 2, fit_to_one_page: bool = False):
    if win32com is None:
        raise RuntimeError("pywin32 not installed. pip install pywin32")
    created = excel is None
    if created:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
    wb = None
    try:
        wb = excel.Workbooks.Open(str(file))
        ws = wb.Worksheets(1) if first_sheet_only else wb.ActiveSheet
        # Apply Letter / Landscape / Fit-all-columns-to-1-page to every sheet
        # that will be exported so the BOM (and any multi-sheet workbook)
        # always prints consistently.
        targets = [ws] if first_sheet_only else list(wb.Worksheets)
        for t in targets:
            try:
                ps = t.PageSetup
                ps.PaperSize = 1          # xlPaperLetter
                ps.Orientation = orientation  # 1=portrait, 2=landscape
                ps.Zoom = False
                ps.FitToPagesWide = 1
                # Fit entire sheet on one page (both dirs) for PRF-style prints
                ps.FitToPagesTall = 1 if fit_to_one_page else False
            except Exception:
                pass
        ws.ExportAsFixedFormat(0, str(dest))
    finally:
        if wb is not None:
            try:
                wb.Close(False)
            except Exception:
                pass
        if created and excel:
            try:
                excel.Quit()
            except Exception:
                pass


def save_merged(files: list[Path], dest: Path):
    if not files:
        warn(f"No files to merge for {dest.name}")
        return
    merged = merge_pdfs(files, dest.stem)
    try:
        shutil.copy2(str(merged), str(dest))
    finally:
        if merged.exists():
            try:
                merged.unlink()
            except Exception:
                pass


# ── Plan building ──────────────────────────────────────────────────────
def _filter_variant(mechs: list[Path], variant: str, label: str = "") -> list[Path]:
    """Filter mech_roots to a single variant by numeric suffix.
    Raises RuntimeError if the specified variant is not found (never silently falls back)."""
    v_match = re.search(r"(\d{2})$", variant)
    v_num = v_match.group(1) if v_match else variant.lstrip("-")
    suffix = f"-{v_num}"
    filtered = [m for m in mechs if m.name.lower().endswith(suffix.lower()) or
                (m.name.lower() == "internal" and m.parent.name.lower().endswith(suffix.lower()))]
    if filtered:
        log(f"Filtered {label}to variant {suffix}: {[m.name for m in filtered]}")
        return filtered
    variant_names = [m.name for m in mechs]
    raise RuntimeError(
        f"Variant '{suffix}' not found in {label}mech_roots: {variant_names}. "
        f"Available variants: {', '.join(variant_names)}")


def _require_variant(mechs: list[Path], variant: str | None, label: str = "") -> list[Path]:
    """Check for multiple variants; filter if specified, raise if ambiguous."""
    if len(mechs) > 1 and not variant:
        variant_names = [m.name for m in mechs]
        raise RuntimeError(
            f"CHOOSE_VARIANT:{label}{','.join(variant_names)}")
    if variant:
        mechs = _filter_variant(mechs, variant, label)
    if len(mechs) != 1:
        warn(f"Expected exactly 1 mech_root after variant filtering, got {len(mechs)}: "
             f"{[m.name for m in mechs]}. Using first.")
        mechs = mechs[:1]
    return mechs


def build_plan(job_folder: Path, ref_folder: Path | None = None,
               variant: str | None = None,
               ref_variant: str | None = None,
               pack_override: str | None = None) -> dict:
    """
    Build the document plan.
    - Normal mode: all docs come from job_folder
    - Vikram mode (ref_folder set): FWO, electrical, PRF from job_folder;
      BOM, CNC, flats, assemblies from ref_folder
    - variant: filter the main job's variants (or the only job in normal mode)
    - ref_variant: filter the reference job's variants (Vikram mode only)
    """
    # In Vikram mode, variant applies to main job, ref_variant to ref job
    # In normal mode, variant applies to the only job
    if ref_folder:
        # Reference job — provides BOM, CNC, flats, assemblies
        ref_ctx = get_context(ref_folder)
        base = Path(ref_ctx["job_root"])
        mechs = _require_variant(ref_ctx["mech_roots"], ref_variant, "ref job: ")

        # Main job — provides FWO, PRF, electrical
        main_ctx = get_context(job_folder)
        main_base = Path(main_ctx["job_root"])
        main_mechs = _require_variant(main_ctx["mech_roots"], variant, "main job: ")
    else:
        ctx = get_context(job_folder)
        base = Path(ctx["job_root"])
        mechs = _require_variant(ctx["mech_roots"], variant)
        main_base = base
        main_mechs = mechs

    plan = {
        "job_folder": str(job_folder),
        "ref_folder": str(ref_folder) if ref_folder else None,
        "base": str(base),
        "main_base": str(main_base),
        "mech_roots": [str(m) for m in mechs],
        "variant_only": ref_ctx["variant_only"] if ref_folder else ctx["variant_only"],
        "sources": {},  # track which job each doc comes from
    }

    # FWO — from main job
    inputs_folder = main_base / "300 Inputs"
    plan["fwo"] = match_fwo(inputs_folder)
    plan["sources"]["fwo"] = str(main_base)

    # BOM — from reference job (or main if no ref)
    plan["bom"] = match_excel(mechs[0] / "204 BOM", "BOM", "BOM")
    plan["sources"]["bom"] = str(base)

    # PRF — from main job
    prf_folder = main_base / "300 Inputs" / "302 Production Release Form"
    plan["prf"] = find_variant_prf(prf_folder, main_mechs)
    plan["sources"]["prf"] = str(main_base)

    # Read PRF data
    prf_data = None
    model_no = ""
    if plan["prf"]:
        prf_data = read_prf(plan["prf"])
        if prf_data:
            model_no = prf_data.get("model_no", "")
    plan["prf_data"] = prf_data

    # Electrical pack — from main job
    plan["pack"] = match_pack(main_base / "100 Elec" / "102 Drawings", model_no, pack_override)
    plan["sources"]["pack"] = str(main_base)

    # CNC, flats, assemblies — from reference job (or main if no ref)
    cnc, flats, assemblies, excluded_lay = [], [], [], []
    for m in mechs:
        cnc.extend(match_cnc(m / "205 CNC"))
        flats.extend(match_flats(m / "202 PDFs_Flats"))
        a, ex = match_assemblies(m / "203 Assemblies")
        assemblies.extend(a)
        excluded_lay.extend(ex)

    plan["cnc"]        = sorted(cnc, key=lambda p: p.name.lower())
    plan["flats"]      = sorted(flats, key=lambda p: p.name.lower())
    plan["assemblies"] = sorted(assemblies, key=lambda p: p.name.lower())
    plan["excluded_lay"] = sorted(excluded_lay, key=lambda p: p.name.lower())
    plan["sources"]["cnc"] = str(base)
    plan["sources"]["flats"] = str(base)
    plan["sources"]["assemblies"] = str(base)

    # CNC mark folder (for marking BOM before export)
    if mechs:
        cnc_folder = mechs[0] / "205 CNC"
        plan["cnc_mark_folder"] = cnc_folder if cnc_folder.is_dir() else None
    else:
        plan["cnc_mark_folder"] = None

    return plan


# ── Document generation ────────────────────────────────────────────────
def generate_documents(plan: dict, out_dir: Path) -> list[tuple[str, Path]]:
    """Generate all PDFs into out_dir. Returns list of (title, path) for printing."""
    out_dir.mkdir(parents=True, exist_ok=True)
    generated = []
    n = [0]

    def dest(label: str) -> Path:
        n[0] += 1
        return out_dir / f"{n[0]:02d}_{safe_name(label)}.pdf"

    # Fill FWO
    fwo_filled = None
    prf_data = plan.get("prf_data")
    if prf_data and plan.get("fwo"):
        fwo_filled = fill_fwo(plan["fwo"], prf_data)

    # CNC column marking
    cnc_mark_folder = plan.get("cnc_mark_folder")
    if cnc_mark_folder and plan.get("bom"):
        log("CNC Column Marker: scanning...")
        try:
            matched, unmatched, total_parts, stock_count, na_marked, stock_found = \
                cnc_mark_bom(plan["bom"], cnc_mark_folder)
            plan["cnc_marked"] = matched       # total X marks (existing + new)
            plan["cnc_unmatched"] = unmatched
            plan["bom_total_parts"] = total_parts
            plan["bom_stock_count"] = stock_count  # total S marks (existing + new)
            plan["bom_na_marked"] = na_marked      # newly marked N/A parts
            plan["bom_stock_found"] = stock_found   # newly found stock parts
            log(f"CNC: marked {matched} row(s).")
            if unmatched:
                warn(f"CNC: {len(unmatched)} row(s) with no CNC file:")
                for item in unmatched:
                    if isinstance(item, tuple):
                        warn(f"  - row {item[0]}: {item[1]}")
                    else:
                        warn(f"  - {item}")
        except Exception as e:
            warn(f"CNC Column Marker failed ({e}) — continuing without marks.")

    # Excel COM session for BOM and PRF export
    excel = None
    if win32com:
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
        except Exception as e:
            warn(f"Could not start Excel COM: {e}")

    try:
        # 1. FWO
        fwo_src = fwo_filled or plan.get("fwo")
        if fwo_src:
            d = dest("Fabrication_Work_Order")
            try:
                save_pdf(fwo_src, d)
                generated.append(("Fabrication Work Order", d))
                log(f"  OK  Fabrication Work Order")
            except Exception as e:
                warn(f"  SKIP  Fabrication Work Order: {e}")

        # 2. BOM
        if plan.get("bom"):
            d = dest("BOM")
            try:
                excel_to_pdf(plan["bom"], d, first_sheet_only=False, excel=excel)
                generated.append(("BOM", d))
                log(f"  OK  BOM")
            except Exception as e:
                warn(f"  SKIP  BOM: {e}")

        # 3. CNC duplex
        for pdf in plan.get("cnc", []):
            if classify_cnc(pdf) == "duplex":
                d = dest(f"CNC_{pdf.stem}")
                try:
                    save_pdf(pdf, d)
                    generated.append((f"CNC (duplex): {pdf.name}", d))
                    log(f"  OK  CNC duplex: {pdf.name}")
                except Exception as e:
                    warn(f"  SKIP  CNC duplex {pdf.name}: {e}")

        # 4. CNC simplex merged
        cnc_simplex = [p for p in plan.get("cnc", []) if classify_cnc(p) == "simplex"]
        if cnc_simplex:
            d = dest("CNC_Simplex_Merged")
            try:
                save_merged(cnc_simplex, d)
                generated.append(("CNC Simplex (merged)", d))
                log(f"  OK  CNC Simplex merged ({len(cnc_simplex)} files)")
            except Exception as e:
                warn(f"  SKIP  CNC Simplex merged: {e}")

        # 5. PDFs_Flats merged
        if plan.get("flats"):
            d = dest("PDFs_Flats_Merged")
            try:
                save_merged(plan["flats"], d)
                generated.append(("PDFs_Flats (merged)", d))
                log(f"  OK  PDFs_Flats merged ({len(plan['flats'])} files)")
            except Exception as e:
                warn(f"  SKIP  PDFs_Flats merged: {e}")

        # 6. PRF
        if plan.get("prf"):
            d = dest("Production_Release_Form")
            try:
                # PRF: Letter, Portrait, one-sided, fit sheet on one page
                excel_to_pdf(plan["prf"], d, first_sheet_only=True, excel=excel,
                             orientation=1, fit_to_one_page=True)
                generated.append(("Production Release Form", d))
                log(f"  OK  Production Release Form")
            except Exception as e:
                warn(f"  SKIP  Production Release Form: {e}")

        # 7. Electrical Pack (pages 1-2)
        if plan.get("pack"):
            d = dest("Electrical_Pack_Pages_1_2")
            try:
                save_pdf(plan["pack"], d, pages=(0, 1))
                generated.append(("Electrical Pack (pages 1-2)", d))
                log(f"  OK  Electrical Pack")
            except Exception as e:
                warn(f"  SKIP  Electrical Pack: {e}")

        # 8. Assemblies merged
        if plan.get("assemblies"):
            d = dest("Assemblies_Merged")
            try:
                save_merged(plan["assemblies"], d)
                generated.append(("Assemblies (merged)", d))
                log(f"  OK  Assemblies merged ({len(plan['assemblies'])} files)")
            except Exception as e:
                warn(f"  SKIP  Assemblies merged: {e}")

    finally:
        if excel:
            try:
                excel.Quit()
            except Exception:
                pass
        # Clean up filled FWO temp file
        if fwo_filled and Path(fwo_filled).exists():
            try:
                Path(fwo_filled).unlink()
            except Exception:
                pass

    return generated


# ── Printing ───────────────────────────────────────────────────────────
def _find_acrobat() -> str | None:
    for p in ACROBAT_SEARCH_PATHS:
        if os.path.isfile(p):
            return p
    return None


def _set_devmode_duplex(printer_name: str, duplex: bool) -> bool:
    if win32print is None:
        return False
    try:
        h = win32print.OpenPrinter(printer_name)
        try:
            info9 = win32print.GetPrinter(h, 9)
            dm = info9.get("pDevMode") if info9 else None
            if dm is None:
                info2 = win32print.GetPrinter(h, 2)
                dm = info2.get("pDevMode")
            if dm is None:
                return False
            dm.Duplex = 2 if duplex else 1
            win32print.SetPrinter(h, 9, {"pDevMode": dm}, 0)
            return True
        finally:
            win32print.ClosePrinter(h)
    except Exception:
        return False


def _wait_spooler_stable(printer_name: str, timeout: float = 60.0):
    JOB_STATUS_SPOOLING = 0x0004
    if not win32print:
        time.sleep(3.5)
        return
    deadline = time.time() + timeout
    while time.time() < deadline:
        try:
            h = win32print.OpenPrinter(printer_name)
            try:
                jobs = win32print.EnumJobs(h, 0, 999, 1)
            finally:
                win32print.ClosePrinter(h)
            if not any(j["Status"] & JOB_STATUS_SPOOLING for j in jobs):
                return
        except Exception:
            time.sleep(3.5)
            return
        time.sleep(0.5)


def print_documents(generated: list[tuple[str, Path]], printer_name: str):
    """Send generated PDFs to printer via Acrobat COM (preferred) or subprocess fallback."""
    if win32com is None:
        error("pywin32 not installed — cannot print.")
        return

    if win32print:
        try:
            win32print.SetDefaultPrinter(printer_name)
        except Exception:
            pass

    # Build print jobs with duplex flags
    print_jobs = []
    for title, pdf_path in generated:
        is_duplex = bool(re.search(r'^\d+_CNC_', pdf_path.stem)
                         and "CNC_Simplex" not in pdf_path.stem)
        print_jobs.append((pdf_path, is_duplex))

    # Try Acrobat COM first
    log("Attempting Acrobat COM print session...")
    try:
        acro_app = win32com.client.Dispatch("AcroExch.App")
        acro_app.Hide()

        def _print_one(pdf_path, is_duplex):
            """Dispatch a single PDF. Returns True on success, raises on COM failure."""
            _set_devmode_duplex(printer_name, is_duplex)
            pddoc = win32com.client.Dispatch("AcroExch.PDDoc")
            if not pddoc.Open(str(pdf_path)):
                error(f"  COM could not open: {pdf_path.name}")
                return False
            try:
                n_pages = pddoc.GetNumPages()
                avdoc = pddoc.OpenAVDoc("")
                if avdoc is None:
                    error(f"  Could not get AVDoc: {pdf_path.name}")
                    return False
                try:
                    ok = avdoc.PrintPages(0, n_pages - 1, 2, True, False)
                finally:
                    try: avdoc.Close(True)
                    except Exception: pass
                return bool(ok)
            finally:
                try: pddoc.Close()
                except Exception: pass

        for i, (pdf_path, is_duplex) in enumerate(print_jobs):
            mode_str = "duplex" if is_duplex else "simplex"
            # Wait for spooler before each dispatch so Acrobat doesn't get
            # slammed while it's still digesting the previous (possibly huge)
            # merged PDF. This prevents the -2147417851 "server threw an
            # exception" error we saw on the PRF after the 93-file merged
            # flats print.
            if i > 0:
                _wait_spooler_stable(printer_name, timeout=120.0)

            last_err = None
            for attempt in range(2):  # one retry on COM failure
                try:
                    if _print_one(pdf_path, is_duplex):
                        log(f"  OK  {pdf_path.name} ({mode_str})")
                    else:
                        error(f"  PrintPages failed: {pdf_path.name}")
                    last_err = None
                    break
                except Exception as e:
                    last_err = e
                    warn(f"  {pdf_path.name}: {e} (attempt {attempt + 1}/2)")
                    # Let Acrobat recover before retrying
                    _wait_spooler_stable(printer_name, timeout=120.0)
                    time.sleep(2.0)
            if last_err is not None:
                error(f"  {pdf_path.name}: giving up after retry — {last_err}")

        acro_app.Exit()
        log("Print sequence complete via Acrobat COM.")
        return

    except Exception as e:
        warn(f"Acrobat COM unavailable ({e}) — falling back to subprocess.")

    # Fallback: per-document subprocess
    acrobat_exe = _find_acrobat()
    if not acrobat_exe:
        warn("Acrobat not found — using default handler (no duplex control)")

    for i, (pdf_path, is_duplex) in enumerate(print_jobs):
        _set_devmode_duplex(printer_name, is_duplex)
        mode_str = "duplex" if is_duplex else "simplex"
        try:
            if acrobat_exe:
                subprocess.run(
                    [acrobat_exe, "/t", str(pdf_path), printer_name],
                    stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                    timeout=120)
                log(f"  OK  {pdf_path.name} via Acrobat ({mode_str})")
            else:
                os.startfile(str(pdf_path), "print")
                log(f"  OK  {pdf_path.name} via default handler")
        except Exception as e:
            error(f"  {pdf_path.name}: {e}")

        if i < len(print_jobs) - 1:
            _wait_spooler_stable(printer_name)

    log("Print sequence complete via subprocess fallback.")


# ── Summary / breakdown ───────────────────────────────────────────────
def make_breakdown(plan: dict, generated: list[tuple[str, Path]],
                   out_dir: Path | None = None) -> str:
    """Create a human-readable breakdown of what was generated and from where."""
    lines = ["=" * 60, "DOC PREP BREAKDOWN", "=" * 60, ""]

    ref = plan.get("ref_folder")
    if ref:
        lines.append(f"Main Job:      {plan['job_folder']}")
        lines.append(f"Reference Job: {ref}")
    else:
        lines.append(f"Job Folder: {plan['job_folder']}")

    prf_data = plan.get("prf_data")
    if prf_data:
        lines += [
            "",
            f"  Job No.   : {prf_data['job_no']}",
            f"  Job Name  : {prf_data['job_name']}",
            f"  Enclosure : {prf_data['enclosure']}   Qty: {prf_data['qty']}",
            f"  Model No. : {prf_data['model_no']}",
        ]

    # BOM details
    cnc_marked = plan.get("cnc_marked", 0)
    cnc_unmatched = plan.get("cnc_unmatched", [])
    bom_total = plan.get("bom_total_parts", 0)
    bom_stock = plan.get("bom_stock_count", 0)
    bom_na = plan.get("bom_na_marked", [])
    bom_stock_found = plan.get("bom_stock_found", [])
    if bom_total:
        lines += [
            "",
            "-" * 60,
            "BOM SUMMARY:",
            "-" * 60,
            f"  Total parts    : {bom_total}",
            f"  CNC marked (X) : {cnc_marked}",
            f"  N/A marked     : {len(bom_na)}",
            f"  Stock (S)      : {bom_stock}",
            f"  Unmatched      : {len(cnc_unmatched)}",
        ]
        if bom_na:
            # Support both old (str) and new (tuple) format
            if bom_na and isinstance(bom_na[0], tuple):
                lines.append(f"  N/A parts (flexibar): {', '.join(f'row {r}: {pn}' for r, pn in bom_na)}")
            else:
                lines.append(f"  N/A parts: {', '.join(bom_na)}")
        if bom_stock_found:
            if bom_stock_found and isinstance(bom_stock_found[0], tuple):
                lines.append(f"  Stock found (new): {', '.join(f'row {r}: {pn}' for r, pn in bom_stock_found)}")
            else:
                lines.append(f"  Stock found (new): {', '.join(bom_stock_found)}")
        if cnc_unmatched:
            if cnc_unmatched and isinstance(cnc_unmatched[0], tuple):
                lines.append(f"  Unmatched parts (missing CNC): {', '.join(f'row {r}: {pn}' for r, pn in cnc_unmatched)}")
            else:
                lines.append(f"  Unmatched parts: {', '.join(cnc_unmatched)}")

    lines += ["", "-" * 60, "GENERATED DOCUMENTS:", "-" * 60]

    sources = plan.get("sources", {})
    source_map = {
        "Fabrication Work Order": sources.get("fwo", ""),
        "BOM": sources.get("bom", ""),
        "PRF": sources.get("prf", ""),
        "Production Release Form": sources.get("prf", ""),
        "Electrical Pack": sources.get("pack", ""),
        "CNC": sources.get("cnc", ""),
        "PDFs_Flats": sources.get("flats", ""),
        "Assemblies": sources.get("assemblies", ""),
    }

    for i, (title, path) in enumerate(generated, 1):
        source = ""
        if ref:
            for key, src_path in source_map.items():
                if key.lower() in title.lower():
                    job_name = Path(src_path).name if src_path else "?"
                    source = f"  [from {job_name}]"
                    break
        lines.append(f"  {i:2d}. {title}{source}")
        lines.append(f"      -> {path.name}")

    excluded = plan.get("excluded_lay", [])
    if excluded:
        lines += ["", f"  Excluded -LAY files: {len(excluded)}"]

    if out_dir:
        lines += ["", "-" * 60, f"Output: {out_dir}"]

    lines += ["", "=" * 60]
    return "\n".join(lines)


# ── Main ───────────────────────────────────────────────────────────────
def run(job_number: str, ref_number: str | None = None,
        do_print: bool = False, variant: str | None = None,
        ref_variant: str | None = None,
        pack_override: str | None = None) -> tuple[Path, str]:
    """
    Main entry point.
    Returns (output_dir, breakdown_text).
    """
    if _missing:
        print("\n[WARNING] Missing dependencies:")
        for dep in _missing:
            print(f"  - {dep}")
        print("Some features may not work. Install with:")
        print(f"  pip install {' '.join(d.split()[0] for d in _missing)}\n")

    # Find job folders
    job_folder = find_job_folder(job_number)
    ref_folder = find_job_folder(ref_number) if ref_number else None

    if ref_folder:
        log(f"Vikram workflow: main={job_folder.name}, ref={ref_folder.name}")

    # Build plan
    log("Building plan...")
    plan = build_plan(job_folder, ref_folder, variant=variant, ref_variant=ref_variant,
                       pack_override=pack_override)

    # Create output directory (Fix #11: local temp + cleanup)
    job_name = safe_name(Path(plan["job_folder"]).name)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_tmp = Path(tempfile.gettempdir()) / "FoxFab_DocPrep"
    base_tmp.mkdir(parents=True, exist_ok=True)
    try:
        _cutoff = time.time() - (7 * 86400)
        _removed = 0
        for _old in base_tmp.iterdir():
            if _old.is_dir() and _old.name.startswith("DocPrep_Output_"):
                try:
                    if _old.stat().st_mtime < _cutoff:
                        shutil.rmtree(_old)
                        _removed += 1
                except Exception:
                    pass
        if _removed:
            log(f"Cleaned up {_removed} old output folders")
    except Exception:
        pass
    out_dir = base_tmp / f"DocPrep_Output_{job_name}_{ts}"

    # Generate documents
    log("Generating documents...")
    generated = generate_documents(plan, out_dir)

    if not generated:
        error("No documents were generated.")
        return out_dir, "No documents generated."

    # Create breakdown
    breakdown = make_breakdown(plan, generated, out_dir)
    print("\n" + breakdown)

    # Save breakdown to output folder
    (out_dir / "BREAKDOWN.txt").write_text(breakdown, encoding="utf-8")

    if do_print:
        log(f"Sending {len(generated)} documents to printer: {PREFERRED_PRINTER}")
        print_documents(generated, PREFERRED_PRINTER)
    else:
        log(f"Simulation complete. {len(generated)} PDFs saved to: {out_dir}")

    return out_dir, breakdown


def main():
    parser = argparse.ArgumentParser(description="FoxFab Doc Prep")
    parser.add_argument("--job", required=True, help="Job number (e.g., J16204)")
    parser.add_argument("--ref", default=None, help="Reference job number for Vikram workflow")
    parser.add_argument("--print", dest="do_print", action="store_true",
                        help="Send to printer after generating")
    parser.add_argument("--variant", default=None,
                        help="Main job variant suffix (e.g., 02 or J16227-02)")
    parser.add_argument("--ref-variant", dest="ref_variant", default=None,
                        help="Reference job variant suffix (Vikram mode)")
    parser.add_argument("--pack", default=None,
                        help="Explicit path to electrical pack PDF (overrides auto-detection)")
    parser.add_argument("--check-bom", dest="check_bom", action="store_true",
                        help="Read-only BOM vs CNC check (no modifications)")
    parser.add_argument("--jobs-root", default=None,
                        help="Override JOBS_ROOT path (e.g., for testing)")
    args = parser.parse_args()

    global JOBS_ROOT
    if args.jobs_root:
        JOBS_ROOT = Path(args.jobs_root)

    try:
        if args.check_bom:
            report, report_path = run_bom_check(args.job)
            print(f"\nReport saved: {report_path}")
        else:
            out_dir, breakdown = run(args.job, args.ref, args.do_print,
                                      variant=args.variant, ref_variant=args.ref_variant,
                                      pack_override=args.pack)
            print(f"\nOutput: {out_dir}")
    except Exception as e:
        error(f"FATAL: {e}")
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
