"""
Print Help — Quick-print electrical drawing (pages 1-2) and PRF for a new job.

Usage (called by Claude or run directly):
    python printhelp.py --job J16204
    python printhelp.py --job J16204 --variant 02
    python printhelp.py --job J16204 --print
    python printhelp.py --job J16204 --output "C:\\some\\folder"
"""

import argparse
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
from datetime import datetime
from pathlib import Path

# ── Configuration ──────────────────────────────────────────────────────
JOBS_ROOT         = Path(r"Z:\FOXFAB_DATA\ENGINEERING\2 JOBS")
PREFERRED_PRINTER = r"\\NPSVR05\FoxFab (Konica Bizhub C360i)"
EXCEL_EXTENSIONS  = {".xlsx", ".xls", ".xlsm"}
PDF_EXTENSIONS    = {".pdf", ".PDF"}

ACROBAT_SEARCH_PATHS = [
    r"C:\Program Files (x86)\Adobe\Acrobat 2017\Acrobat\Acrobat.exe",
    r"C:\Program Files\Adobe\Acrobat 2017\Acrobat\Acrobat.exe",
    r"C:\Program Files (x86)\Adobe\Acrobat DC\Acrobat\Acrobat.exe",
    r"C:\Program Files\Adobe\Acrobat DC\Acrobat\Acrobat.exe",
    r"C:\Program Files (x86)\Adobe\Acrobat 2020\Acrobat\Acrobat.exe",
    r"C:\Program Files\Adobe\Acrobat 2020\Acrobat\Acrobat.exe",
]

# ── Dependency checks ─────────────────────────────────────────────────
_missing = []

try:
    import win32print
    import win32com.client
except ImportError:
    win32print = None
    win32com = None
    _missing.append("pywin32  (pip install pywin32)")

try:
    from pypdf import PdfReader, PdfWriter
except ImportError:
    PdfReader = PdfWriter = None
    _missing.append("pypdf  (pip install pypdf)")

try:
    import openpyxl as _openpyxl
except ImportError:
    _openpyxl = None
    _missing.append("openpyxl  (pip install openpyxl)")

try:
    import xlwings as xw
except ImportError:
    xw = None
    _missing.append("xlwings  (pip install xlwings)")

# ── Logging ────────────────────────────────────────────────────────────
_log_lines: list[str] = []


def log(msg: str):
    stamp = datetime.now().strftime("%H:%M:%S")
    line = f"[{stamp}] {msg}"
    _log_lines.append(line)
    print(line)


def warn(msg: str):
    log(f"[WARN] {msg}")


def error(msg: str):
    log(f"[ERROR] {msg}")


# ── Helpers ────────────────────────────────────────────────────────────
def list_files(folder: Path) -> list[Path]:
    return sorted([p for p in folder.iterdir() if p.is_file()],
                  key=lambda p: p.name.lower())


def find_job_folder(job_number: str) -> Path:
    if not JOBS_ROOT.is_dir():
        raise RuntimeError(f"Jobs root not found: {JOBS_ROOT}")

    job_up = job_number.upper()
    matches = [d for d in JOBS_ROOT.iterdir()
               if d.is_dir() and d.name.upper().startswith(job_up)]

    if len(matches) == 0:
        raise RuntimeError(f"No job folder found matching '{job_number}' in {JOBS_ROOT}")
    if len(matches) == 1:
        log(f"Found job folder: {matches[0].name}")
        return matches[0]

    raise RuntimeError(
        f"CHOOSE_FOLDER: Multiple folders match '{job_number}':\n"
        + "\n".join(f"  {i+1}. {m.name}" for i, m in enumerate(matches))
        + "\nPlease specify which one to use.")


# ── Folder structure resolution ────────────────────────────────────────
def _has_mech_subs(p: Path) -> bool:
    return all((p / s).is_dir() for s in
               ["204 BOM", "205 CNC", "202 PDFs_Flats", "203 Assemblies"])


def _is_numbered_variant(p: Path) -> bool:
    return p.is_dir() and re.search(r"-\d{2}$", p.name) is not None


def _resolve_internal(folder: Path) -> Path | None:
    internal = folder / "Internal"
    if internal.is_dir() and _has_mech_subs(internal):
        log(f"Found Internal subfolder in {folder.name}, using it")
        return internal
    for child in sorted(folder.iterdir(), key=lambda p: p.name.lower()):
        if child.is_dir() and _has_mech_subs(child):
            return None
    return None


def get_context(job_folder: Path) -> dict:
    """Resolve folder structure and return job_root, mech_roots, variant_only."""

    # Pattern 1: user selected a numbered variant directly
    if _is_numbered_variant(job_folder) and _has_mech_subs(job_folder):
        return {"job_root": job_folder.parent.parent,
                "mech_roots": [job_folder], "variant_only": True}

    if _is_numbered_variant(job_folder):
        internal = _resolve_internal(job_folder)
        if internal:
            return {"job_root": job_folder.parent.parent,
                    "mech_roots": [internal], "variant_only": True}

    mech = job_folder / "200 Mech"
    if not mech.is_dir():
        raise RuntimeError("Missing folder: 200 Mech")

    # Pattern 2: standard direct structure
    if _has_mech_subs(mech):
        return {"job_root": job_folder, "mech_roots": [mech], "variant_only": False}

    # Pattern 3: numbered variants inside 200 Mech
    variants = []
    for child in sorted(mech.iterdir(), key=lambda p: p.name.lower()):
        if not child.is_dir():
            continue
        if _is_numbered_variant(child):
            if _has_mech_subs(child):
                variants.append(child)
            else:
                internal = _resolve_internal(child)
                if internal:
                    variants.append(internal)

    if variants:
        return {"job_root": job_folder, "mech_roots": variants, "variant_only": False}

    # Pattern 4a: named subfolders directly in 200 Mech
    internal = _resolve_internal(mech)
    if internal:
        return {"job_root": job_folder, "mech_roots": [internal], "variant_only": False}

    named_subs = []
    for child in sorted(mech.iterdir(), key=lambda p: p.name.lower()):
        if child.is_dir() and _has_mech_subs(child):
            named_subs.append(child)
    if named_subs:
        raise RuntimeError(
            f"CHOOSE_SUBFOLDER: Found subfolders with mechanical structure but no 'Internal' folder:\n"
            + "\n".join(f"  - {s.name}" for s in named_subs)
            + "\nPlease specify which subfolder to use.")

    raise RuntimeError(
        "Could not find a usable mechanical folder structure. "
        "Expected '200 Mech\\204 BOM' etc. directly, numbered variants like '*-01' "
        "inside 200 Mech, or named subfolders like 'Internal'.")


# ── PRF reading (model-aware pack matching) ────────────────────────────
def read_prf_model(prf_path: Path) -> str:
    """Read model number from PRF Excel file (cell G9). Returns empty string on failure."""
    if _openpyxl is None:
        warn("openpyxl not installed — cannot read PRF for model number.")
        return ""
    try:
        wb = _openpyxl.load_workbook(str(prf_path), data_only=True)
        ws = wb["Form"] if "Form" in wb.sheetnames else wb.worksheets[0]
        v = ws["G9"].value
        model = str(v).strip() if v is not None else ""
        if model:
            log(f"PRF model number: {model}")
        return model
    except Exception as e:
        warn(f"Failed to read PRF model number: {e}")
        return ""


# ── File matching ──────────────────────────────────────────────────────
def match_pack(folder: Path, model_no: str = "") -> Path | None:
    if not folder.is_dir():
        warn(f"Electrical drawings folder not found: {folder}")
        return None
    all_packs = [f for f in list_files(folder)
                 if f.suffix in PDF_EXTENSIONS and "PACK" in f.name.upper()]
    if not all_packs:
        warn("No PDF with 'PACK' in name found in Electrical Drawings.")
        return None
    if model_no:
        model_up = model_no.upper()
        model_matches = [f for f in all_packs if model_up in f.name.upper()]
        if len(model_matches) == 1:
            return model_matches[0]
        if model_matches:
            log(f"Multiple PACK PDFs match model '{model_no}', using first: {model_matches[0].name}")
            return model_matches[0]
    if len(all_packs) == 1:
        return all_packs[0]
    # Multiple packs, no model match — list them for the error
    names = [f.name for f in all_packs]
    raise RuntimeError(
        f"CHOOSE_PACK: Multiple PACK PDFs found and no model number to disambiguate:\n"
        + "\n".join(f"  {i+1}. {n}" for i, n in enumerate(names))
        + "\nPlease specify which one to use.")


def find_prf(prf_folder: Path, prf_name: str | None = None) -> Path | None:
    """Find PRF Excel file. If prf_name is given, match it specifically.
    If multiple PRFs exist and none specified, raise CHOOSE_PRF."""
    if not prf_folder.is_dir():
        warn(f"PRF folder not found: {prf_folder}")
        return None
    all_prfs = [f for f in list_files(prf_folder)
                if f.suffix.lower() in {e.lower() for e in EXCEL_EXTENSIONS}
                and "prf" in f.name.lower()]
    if not all_prfs:
        warn(f"No PRF file found in {prf_folder}")
        return None

    # If a specific PRF was requested, match it
    if prf_name:
        for f in all_prfs:
            if f.name == prf_name:
                return f
        warn(f"PRF '{prf_name}' not found in {prf_folder}")
        return None

    if len(all_prfs) == 1:
        return all_prfs[0]

    # Multiple PRFs — ask user which one
    names = [f.name for f in all_prfs]
    raise RuntimeError(
        f"CHOOSE_PRF: Multiple PRF files found:\n"
        + "\n".join(f"  {i+1}. {n}" for i, n in enumerate(names))
        + "\nPlease specify which PRF to use.")


# ── PDF operations ─────────────────────────────────────────────────────
def save_pdf_pages(src: Path, dest: Path, pages: tuple[int, int]):
    """Extract specific pages from a PDF. Pages are zero-indexed (0,1) = first two pages."""
    if PdfReader is None:
        raise RuntimeError("pypdf not installed. pip install pypdf")
    reader = PdfReader(str(src))
    writer = PdfWriter()
    max_page = min(pages[1] + 1, len(reader.pages))
    for i in range(pages[0], max_page):
        writer.add_page(reader.pages[i])
    with open(dest, "wb") as f:
        writer.write(f)


def excel_to_pdf(file: Path, dest: Path):
    """Export first sheet of an Excel file to PDF via COM."""
    if win32com is None:
        raise RuntimeError("pywin32 not installed. pip install pywin32")
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = None
    try:
        wb = excel.Workbooks.Open(str(file))
        ws = wb.Worksheets(1)
        ws.ExportAsFixedFormat(0, str(dest))
    finally:
        if wb is not None:
            try:
                wb.Close(False)
            except Exception:
                pass
        try:
            excel.Quit()
        except Exception:
            pass


# ── Printing ───────────────────────────────────────────────────────────
def _find_acrobat() -> str | None:
    for p in ACROBAT_SEARCH_PATHS:
        if os.path.isfile(p):
            return p
    return None


DMPAPER_LETTER  = 1   # 8.5 x 11
DMPAPER_TABLOID = 3   # 11 x 17


def _set_devmode(printer_name: str, paper_size: int = DMPAPER_LETTER) -> bool:
    """Set devmode: simplex + paper size (DMPAPER_LETTER or DMPAPER_TABLOID)."""
    if win32print is None:
        return False
    try:
        h = win32print.OpenPrinter(printer_name)
        try:
            info9 = win32print.GetPrinter(h, 9)
            dm = info9.get("pDevMode") if info9 else None
            if dm is None:
                info2 = win32print.GetPrinter(h, 2)
                dm = info2.get("pDevMode")
            if dm is None:
                return False
            dm.Duplex = 1  # simplex
            dm.PaperSize = paper_size
            win32print.SetPrinter(h, 9, {"pDevMode": dm}, 0)
            return True
        finally:
            win32print.ClosePrinter(h)
    except Exception:
        return False


def _wait_spooler_stable(printer_name: str, timeout: float = 60.0):
    JOB_STATUS_SPOOLING = 0x0004
    if not win32print:
        time.sleep(3.5)
        return
    deadline = time.time() + timeout
    while time.time() < deadline:
        try:
            h = win32print.OpenPrinter(printer_name)
            try:
                jobs = win32print.EnumJobs(h, 0, 999, 1)
            finally:
                win32print.ClosePrinter(h)
            if not any(j["Status"] & JOB_STATUS_SPOOLING for j in jobs):
                return
        except Exception:
            time.sleep(3.5)
            return
        time.sleep(0.5)


def print_documents(print_jobs: list[tuple[Path, int]], printer_name: str):
    """Send PDFs to printer via Acrobat COM. Each job is (pdf_path, paper_size).
    Paper sizes: DMPAPER_LETTER=1 for letter, DMPAPER_TABLOID=3 for 11x17."""
    if win32com is None:
        error("pywin32 not installed — cannot print.")
        return

    if win32print:
        try:
            win32print.SetDefaultPrinter(printer_name)
        except Exception:
            pass

    # Try Acrobat COM first
    log("Attempting Acrobat COM print session...")
    try:
        acro_app = win32com.client.Dispatch("AcroExch.App")
        acro_app.Hide()

        for i, (pdf_path, paper_size) in enumerate(print_jobs):
            paper_label = "tabloid" if paper_size == DMPAPER_TABLOID else "letter"
            _set_devmode(printer_name, paper_size)
            try:
                pddoc = win32com.client.Dispatch("AcroExch.PDDoc")
                if not pddoc.Open(str(pdf_path)):
                    error(f"  COM could not open: {pdf_path.name}")
                    continue
                n_pages = pddoc.GetNumPages()
                avdoc = pddoc.OpenAVDoc("")
                if avdoc is None:
                    error(f"  Could not get AVDoc: {pdf_path.name}")
                    continue
                ok = avdoc.PrintPages(0, n_pages - 1, 2, True, False)
                if ok:
                    log(f"  OK  {pdf_path.name} ({paper_label}, simplex)")
                else:
                    error(f"  PrintPages failed: {pdf_path.name}")
                avdoc.Close(True)
                pddoc.Close()
            except Exception as e:
                error(f"  {pdf_path.name}: {e}")

            if i < len(print_jobs) - 1:
                _wait_spooler_stable(printer_name)

        acro_app.Exit()
        log("Print sequence complete via Acrobat COM.")
        return

    except Exception as e:
        warn(f"Acrobat COM unavailable ({e}) — falling back to subprocess.")

    # Fallback: per-document subprocess
    acrobat_exe = _find_acrobat()
    if not acrobat_exe:
        warn("Acrobat not found — using default handler")

    for i, (pdf_path, paper_size) in enumerate(print_jobs):
        _set_devmode(printer_name, paper_size)
        try:
            if acrobat_exe:
                subprocess.run(
                    [acrobat_exe, "/t", str(pdf_path), printer_name],
                    stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                    timeout=120)
                log(f"  OK  {pdf_path.name} via Acrobat")
            else:
                os.startfile(str(pdf_path), "print")
                log(f"  OK  {pdf_path.name} via default handler")
        except Exception as e:
            error(f"  {pdf_path.name}: {e}")

        if i < len(print_jobs) - 1:
            _wait_spooler_stable(printer_name)

    log("Print sequence complete via subprocess fallback.")


# ── Main workflow ──────────────────────────────────────────────────────
def run(job_number: str, prf_name: str | None = None, do_print: bool = False,
        jobs_root: str | None = None, output: str | None = None):
    global JOBS_ROOT
    if jobs_root:
        JOBS_ROOT = Path(jobs_root)
    log(f"Print Help — Job: {job_number}" + (f" PRF: {prf_name}" if prf_name else ""))

    if _missing:
        error("Missing dependencies:\n  " + "\n  ".join(_missing))
        sys.exit(1)

    # 1. Find the job folder
    job_folder = find_job_folder(job_number)

    # 2. Resolve folder structure (only to find job_root)
    ctx = get_context(job_folder)
    job_root = Path(ctx["job_root"])

    # 3. Find the PRF (variant selection happens here — by PRF, not folder)
    prf_folder = job_root / "300 Inputs" / "302 Production Release Form"
    prf_excel = find_prf(prf_folder, prf_name=prf_name)

    # 4. Read model number from PRF for electrical pack matching
    model_no = ""
    if prf_excel:
        model_no = read_prf_model(prf_excel)

    # 5. Find the electrical drawing pack (model-aware)
    elec_folder = job_root / "100 Elec" / "102 Drawings"
    pack_pdf = match_pack(elec_folder, model_no=model_no)

    if not pack_pdf and not prf_excel:
        error("Neither electrical pack nor PRF found. Nothing to print.")
        sys.exit(1)

    # 6. Determine output directory
    if output:
        out_dir = Path(output)
    else:
        script_dir = Path(__file__).resolve().parent
        out_dir = script_dir.parent / "output"
    out_dir.mkdir(parents=True, exist_ok=True)

    parts: list[Path] = []  # individual PDFs to merge, in order: PRF first, then electrical

    # PRF — first sheet only, exported to PDF (goes first in bundle)
    prf_tmp = None
    if prf_excel:
        prf_tmp = out_dir / f"{job_number}_PRF_tmp.pdf"
        try:
            excel_to_pdf(prf_excel, prf_tmp)
            parts.append(prf_tmp)
            log(f"  OK  PRF from: {prf_excel.name}")
        except Exception as e:
            warn(f"  SKIP  PRF: {e}")
            prf_tmp = None
    else:
        warn("  SKIP  PRF — not found")

    # Electrical pack — pages 1-2 (goes second in bundle)
    elec_tmp = None
    if pack_pdf:
        elec_tmp = out_dir / f"{job_number}_Elec_tmp.pdf"
        try:
            save_pdf_pages(pack_pdf, elec_tmp, pages=(0, 1))
            parts.append(elec_tmp)
            log(f"  OK  Electrical Pack (pages 1-2) from: {pack_pdf.name}")
        except Exception as e:
            warn(f"  SKIP  Electrical Pack: {e}")
            elec_tmp = None
    else:
        warn("  SKIP  Electrical Pack — not found")

    if not parts:
        error("No documents were generated. Nothing to print.")
        sys.exit(1)

    # 7. Merge into single bundled PDF
    bundle_path = out_dir / f"{job_number}_Print_Help.pdf"
    if PdfReader is None:
        error("pypdf not installed — cannot merge PDFs.")
        sys.exit(1)

    writer = PdfWriter()
    for part in parts:
        reader = PdfReader(str(part))
        for page in reader.pages:
            writer.add_page(page)
    with open(bundle_path, "wb") as f:
        writer.write(f)
    log(f"  OK  Bundled PDF: {bundle_path.name} ({writer.pages.__len__()} pages)")

    has_prf = prf_tmp is not None and prf_tmp in parts
    has_elec = elec_tmp is not None and elec_tmp in parts

    # Clean up temp files
    for part in parts:
        try:
            part.unlink()
        except Exception:
            pass

    # 8. Summary
    print()
    print("=" * 50)
    print("PRINT HELP SUMMARY")
    print("=" * 50)
    print(f"Job:     {job_folder.name}")
    print(f"Output:  {bundle_path}")
    print()
    if has_prf:
        print(f"  1. Production Release Form (first sheet)")
    if has_elec:
        print(f"  2. Electrical Pack (pages 1-2)")
    print()

    # 9. Print if requested — send PRF (letter) and electrical (tabloid) as separate jobs
    if do_print:
        log(f"Printing to {PREFERRED_PRINTER} ...")
        jobs: list[tuple[Path, int]] = []
        if has_prf:
            prf_print = out_dir / f"{job_number}_PRF_print.pdf"
            # Re-extract PRF for printing (temp files were cleaned)
            excel_to_pdf(prf_excel, prf_print)
            jobs.append((prf_print, DMPAPER_LETTER))
        if has_elec:
            elec_print = out_dir / f"{job_number}_Elec_print.pdf"
            save_pdf_pages(pack_pdf, elec_print, pages=(0, 1))
            jobs.append((elec_print, DMPAPER_TABLOID))
        print_documents(jobs, PREFERRED_PRINTER)
        # Clean up print temp files
        for pdf, _ in jobs:
            try:
                pdf.unlink()
            except Exception:
                pass
    else:
        print("Simulation mode — PDF generated but NOT printed.")
        print(f"Re-run with --print to send to printer.")

    return {"out_dir": str(out_dir), "generated": [str(bundle_path)]}


# ── CLI ────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Print Help — quick-print electrical + PRF")
    parser.add_argument("--job", required=True, help="Job number (e.g. J16204)")
    parser.add_argument("--prf", default=None, help="Specific PRF filename (e.g. 'J16204-01 PRF.xlsx')")
    parser.add_argument("--print", dest="do_print", action="store_true",
                        help="Send to printer after generating")
    parser.add_argument("--jobs-root", default=None,
                        help="Override jobs root folder (for testing)")
    parser.add_argument("--output", default=None,
                        help="Override output directory for generated PDFs")
    args = parser.parse_args()
    try:
        run(args.job, prf_name=args.prf, do_print=args.do_print,
            jobs_root=args.jobs_root, output=args.output)
    except RuntimeError as e:
        error(str(e))
        sys.exit(1)


if __name__ == "__main__":
    main()
