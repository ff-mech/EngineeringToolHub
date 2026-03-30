"""
Engineering Tool Hub  —  app.py
FoxFab internal engineering utilities, combined into one application.

Tools:
  • Bom Filler         – marks stock parts, copies non-stock PDFs/DXFs
  • Doc Prep & Print   – builds and prints (or simulates) a manufacturing packet
  • File Logger        – tracks SolidWorks part numbers across job folders (Parts Tracker)
  • SW Batch Update    – updates SolidWorks custom properties and exports DXFs
"""

from __future__ import annotations

import os
import re
import sys
import time
import base64
import shutil
import queue
import threading
import tempfile
import traceback
import subprocess
from pathlib import Path
from datetime import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# ── splash screen — shown immediately, before heavy imports ──────────
def _make_splash() -> tuple["tk.Tk", "tk.Toplevel"]:
    """Create a minimal loading window to cover the heavy-import delay."""
    root = tk.Tk()
    root.withdraw()

    sp = tk.Toplevel(root)
    sp.overrideredirect(True)
    sp.configure(bg="#1E2B40")
    w, h = 340, 148
    sw, sh = sp.winfo_screenwidth(), sp.winfo_screenheight()
    sp.geometry(f"{w}x{h}+{(sw - w) // 2}+{(sh - h) // 2}")

    tk.Label(sp, text="Engineering Tool Hub",
             bg="#1E2B40", fg="#FFFFFF",
             font=("Segoe UI", 14, "bold")).pack(pady=(26, 2))
    tk.Label(sp, text="v1.1.0  —  FoxFab",
             bg="#1E2B40", fg="#64748B",
             font=("Segoe UI", 9)).pack()
    tk.Label(sp, text="Starting up…",
             bg="#1E2B40", fg="#93C5FD",
             font=("Segoe UI", 9)).pack(pady=(8, 0))
    bar = ttk.Progressbar(sp, mode="indeterminate", length=280)
    bar.pack(pady=(10, 0))
    bar.start(10)

    sp.lift()
    sp.attributes("-topmost", True)
    sp.update()
    return root, sp


_splash_root, _splash_win = _make_splash()

# ── optional heavy imports ────────────────────────────────────────────
try:
    import win32print
    import win32com.client
    import pythoncom
except Exception:
    win32print = None
    win32com = None
    pythoncom = None

try:
    import xlwings as xw
except Exception:
    xw = None

try:
    from pypdf import PdfReader, PdfWriter
except Exception:
    PdfReader = None
    PdfWriter = None

try:
    import fitz as _fitz   # PyMuPDF — FWO text overlay + PDF viewer
except Exception:
    _fitz = None


try:
    import openpyxl as _openpyxl   # PRF data reading (no COM needed)
except Exception:
    _openpyxl = None


# ═════════════════════════════════════════════════════════════════════
#  CONSTANTS
# ═════════════════════════════════════════════════════════════════════

APP_TITLE   = "Engineering Tool Hub"
APP_VERSION = "1.1.0"

PREFERRED_PRINTER  = r"\\NPSVR05\FoxFab (Konica Bizhub C360i)"
STOCK_PARTS_FOLDER = r"Z:\FOXFAB_DATA\ENGINEERING\0 PRODUCTS\300 Stock Parts\PDFs & Flats"

# ── Adobe Acrobat executable search paths (used for duplex-aware printing)
ACROBAT_SEARCH_PATHS = [
    r"C:\Program Files (x86)\Adobe\Acrobat 2017\Acrobat\Acrobat.exe",
    r"C:\Program Files\Adobe\Acrobat 2017\Acrobat\Acrobat.exe",
    r"C:\Program Files (x86)\Adobe\Acrobat DC\Acrobat\Acrobat.exe",
    r"C:\Program Files\Adobe\Acrobat DC\Acrobat\Acrobat.exe",
    r"C:\Program Files (x86)\Adobe\Acrobat 2020\Acrobat\Acrobat.exe",
    r"C:\Program Files\Adobe\Acrobat 2020\Acrobat\Acrobat.exe",
]
BOM_SHEET_NAME     = "FFMPL"
EXCEL_EXTENSIONS   = {".xlsx", ".xls", ".xlsm"}
PDF_EXTENSIONS     = {".pdf", ".PDF"}

# ── FWO auto-fill field positions (PDF points; Y increases downward, origin top-left)
# ── Tune these if text lands in the wrong spot after a test print.
# ── Each field has its own X so they can be nudged independently.
FWO_JOB_NO_X    = 165   # x: JOB NO. value  (slightly right of label)
FWO_JOB_NO_Y    = 145 # y: JOB NO. value  (above the box rule)
FWO_JOB_NAME_X  = 165   # x: JOB NAME value (slightly right)
FWO_JOB_NAME_Y  = 165   # y: JOB NAME value (above the box rule, nudged up)
FWO_DATE_X      = 165   # x: DATE value
FWO_DATE_Y      = 182   # y: DATE value     (above the box rule)
FWO_ENCLOSURE_X = 165   # x: ENCLOSURE value
FWO_ENCLOSURE_Y = 210   # y: ENCLOSURE value (above the box rule)
FWO_UNITS_X     = 165   # x: TOTAL UNITS value
FWO_UNITS_Y     = 245   # y: TOTAL UNITS value (above the box rule)
FWO_FONT_SIZE   = 11    # pt — match the form's body text size

# Sidebar
C_SIDEBAR   = "#1E2B40"
C_SIDEBAR_H = "#263550"
C_SIDEBAR_A = "#1F6FEB"
C_SIDEBAR_T = "#CBD5E1"     # normal text
C_SIDEBAR_M = "#64748B"     # muted text (version, status)
C_SIDEBAR_D = "#2D3E56"     # divider

# Content
C_BG        = "#F6F8FB"
C_PANEL     = "#FFFFFF"
C_BORDER    = "#D8DEE9"
C_ACCENT    = "#1F6FEB"
C_ACCENT_H  = "#1558D6"
C_TEXT      = "#1F2937"
C_SUBTLE    = "#6B7280"
C_SUCCESS   = "#16A34A"
C_ERROR     = "#DC2626"
C_WARN      = "#D97706"

# Terminal
C_TERM_BG   = "#0F1923"
C_TERM_FG   = "#E2E8F0"
C_TERM_HDR  = "#1A2535"
C_TERM_INFO = "#60A5FA"
C_TERM_OK   = "#4ADE80"
C_TERM_ERR  = "#F87171"
C_TERM_WARN = "#FCD34D"
C_TERM_MUTE = "#4B5563"

F_BODY  = ("Segoe UI", 10)
F_SMALL = ("Segoe UI", 9)
F_MONO  = ("Consolas", 10)


# ═════════════════════════════════════════════════════════════════════
#  UTILITIES
# ═════════════════════════════════════════════════════════════════════

def exe_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).resolve().parent


def safe_name(text: str) -> str:
    if not text:
        return "unknown"
    return re.sub(r'[\\/:*?"<>|]+', "_", text).strip() or "unknown"


def _find_es_exe() -> str:
    if getattr(sys, "frozen", False):
        bundled = os.path.join(sys._MEIPASS, "es.exe")
        if os.path.isfile(bundled):
            return bundled
    for candidate in [exe_dir() / "es.exe",
                      exe_dir() / "tools" / "BomFiller" / "es.exe"]:
        if candidate.exists():
            return str(candidate)
    return "es.exe"


ES_EXE = _find_es_exe()


# ═════════════════════════════════════════════════════════════════════
#  MASTER LOG
# ═════════════════════════════════════════════════════════════════════

class MasterLog:
    def __init__(self):
        self._lock = threading.Lock()
        log_dir = exe_dir() / "logs"
        log_dir.mkdir(exist_ok=True)
        stamp = datetime.now().strftime("%Y-%m-%d")
        self._path = log_dir / f"ETH_master_{stamp}.log"

    def append_section(self, tool: str, lines: list[str]):
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        block = (
            f"\n{'=' * 70}\n"
            f"  TOOL : {tool}\n"
            f"  START: {now}\n"
            f"{'=' * 70}\n"
            + "\n".join(lines)
            + "\n"
        )
        with self._lock:
            try:
                with open(self._path, "a", encoding="utf-8") as f:
                    f.write(block)
            except Exception:
                pass

    @property
    def path(self) -> Path:
        return self._path


_master_log = MasterLog()


# ═════════════════════════════════════════════════════════════════════
#  STOP MECHANISM
# ═════════════════════════════════════════════════════════════════════

_stop_event = threading.Event()


class StopRequested(Exception):
    pass


def check_stop():
    if _stop_event.is_set():
        raise StopRequested()


# ═════════════════════════════════════════════════════════════════════
#  APP CLASS
# ═════════════════════════════════════════════════════════════════════

class App:
    TOOLS = [
        ("htu",   "How to Use"),
        ("bom",   "Bom Filler"),
        ("dpp",   "Doc Prep & Print"),
        ("sw",    "SW Batch Update"),
        ("swpdf", "SW Batch PDF Export"),
        ("fl",    "File Logger"),
        ("tm",    "Training Materials"),
    ]

    def __init__(self, root: tk.Tk):
        self.root = root
        root.title(APP_TITLE)
        root.geometry("1300x820")
        root.minsize(1100, 720)
        root.configure(bg=C_BG)

        self._log_queue: queue.Queue = queue.Queue()
        self._worker_thread: threading.Thread | None = None
        self._active_key = "bom"
        self._active_term: tk.Text | None = None
        self._panel_ui: dict[str, dict] = {}   # key -> {run, stop, progress, lbl}
        self._panel_terms: dict[str, tk.Text] = {}
        self._panel_term_frames: dict[str, tk.Frame] = {}  # key -> terminal outer wrap

        self._configure_styles()
        self._build_layout()
        self._switch_tool("htu")
        self._poll_log_queue()

    # ── styles ────────────────────────────────────────────────────────

    def _configure_styles(self):
        s = ttk.Style(self.root)
        try:
            s.theme_use("vista")
        except Exception:
            try:
                s.theme_use("clam")
            except Exception:
                pass
        s.configure("ETH.Horizontal.TProgressbar",
                    troughcolor=C_BORDER, background=C_ACCENT, thickness=7)

    # ── layout ────────────────────────────────────────────────────────

    def _build_layout(self):
        outer = tk.Frame(self.root, bg=C_BG)
        outer.pack(fill="both", expand=True)

        # sidebar
        self._sidebar = tk.Frame(outer, bg=C_SIDEBAR, width=215)
        self._sidebar.pack(side="left", fill="y")
        self._sidebar.pack_propagate(False)
        self._build_sidebar()

        # 1px divider
        tk.Frame(outer, bg=C_BORDER, width=1).pack(side="left", fill="y")

        # content area
        self._content = tk.Frame(outer, bg=C_BG)
        self._content.pack(side="left", fill="both", expand=True)

        # panel frames — content is built lazily on first _switch_tool() call
        self._panels: dict[str, tk.Frame] = {}
        self._panel_built: set[str] = set()
        self._builders: dict[str, object] = {
            "htu":   self._build_htu_panel,
            "bom":   self._build_bom_panel,
            "dpp":   self._build_dpp_panel,
            "sw":    self._build_sw_panel,
            "swpdf": self._build_swpdf_panel,
            "fl":    self._build_fl_panel,
            "tm":    self._build_tm_panel,
        }
        for key, _ in self.TOOLS:
            frame = tk.Frame(self._content, bg=C_BG)
            self._panels[key] = frame

    def _build_sidebar(self):
        sb = self._sidebar

        # App title
        title_f = tk.Frame(sb, bg=C_SIDEBAR, pady=20)
        title_f.pack(fill="x")
        tk.Label(title_f, text="Engineering", bg=C_SIDEBAR,
                 fg="#FFFFFF", font=("Segoe UI", 13, "bold")).pack()
        tk.Label(title_f, text="Tool Hub", bg=C_SIDEBAR,
                 fg="#93C5FD", font=("Segoe UI", 11)).pack()
        tk.Label(title_f, text=f"v{APP_VERSION}  —  FoxFab",
                 bg=C_SIDEBAR, fg=C_SIDEBAR_M,
                 font=("Segoe UI", 8)).pack(pady=(3, 0))

        tk.Frame(sb, bg=C_SIDEBAR_D, height=1).pack(fill="x", padx=16, pady=(0, 6))

        # nav buttons
        self._nav_btns: dict[str, tk.Button] = {}
        for key, label in self.TOOLS:
            btn = tk.Button(
                sb, text=f"   {label}",
                anchor="w", bg=C_SIDEBAR, fg=C_SIDEBAR_T,
                activebackground=C_SIDEBAR_H, activeforeground="#FFFFFF",
                font=("Segoe UI", 10), bd=0, pady=10,
                cursor="hand2", relief="flat",
                command=lambda k=key: self._switch_tool(k),
            )
            btn.pack(fill="x")
            self._nav_btns[key] = btn

        # master log path label (bottom)
        tk.Frame(sb, bg=C_SIDEBAR_D, height=1).pack(
            fill="x", padx=16, side="bottom", pady=(0, 2))
        self._status_lbl = tk.Label(
            sb, text="Ready", bg=C_SIDEBAR, fg=C_SIDEBAR_M,
            font=("Segoe UI", 8), anchor="w", padx=14, wraplength=190,
            justify="left")
        self._status_lbl.pack(side="bottom", fill="x", pady=(0, 6))

    def _switch_tool(self, key: str):
        # build panel content on first visit
        if key not in self._panel_built:
            self._builders[key](self._panels[key])
            self._panel_built.add(key)

        self._active_key = key
        for k, btn in self._nav_btns.items():
            if k == key:
                btn.configure(bg=C_SIDEBAR_A, fg="#FFFFFF",
                               font=("Segoe UI", 10, "bold"))
            else:
                btn.configure(bg=C_SIDEBAR, fg=C_SIDEBAR_T,
                               font=("Segoe UI", 10))
        for k, frame in self._panels.items():
            if k == key:
                frame.pack(fill="both", expand=True)
            else:
                frame.pack_forget()
        if key in self._panel_terms:
            self._active_term = self._panel_terms[key]

    # ── shared UI builders ────────────────────────────────────────────

    def _section_header(self, parent, title: str, subtitle: str = ""):
        hdr = tk.Frame(parent, bg=C_BG)
        hdr.pack(fill="x", padx=26, pady=(20, 4))
        tk.Label(hdr, text=title, bg=C_BG, fg=C_TEXT,
                 font=("Segoe UI", 16, "bold")).pack(anchor="w")
        if subtitle:
            tk.Label(hdr, text=subtitle, bg=C_BG, fg=C_SUBTLE,
                     font=F_SMALL).pack(anchor="w", pady=(2, 0))
        tk.Frame(parent, bg=C_ACCENT, height=2).pack(
            fill="x", padx=26, pady=(4, 0))

    def _card(self, parent, title: str = "", padx=26, pady=(8, 4)):
        wrap = tk.Frame(parent, bg=C_BG)
        wrap.pack(fill="x", padx=padx, pady=pady)
        border = tk.Frame(wrap, bg=C_BORDER, bd=0)
        border.pack(fill="x")
        inner = tk.Frame(border, bg=C_PANEL)
        inner.pack(fill="x", padx=1, pady=1)
        content = tk.Frame(inner, bg=C_PANEL)
        content.pack(fill="x", padx=14, pady=10)
        if title:
            tk.Label(content, text=title, bg=C_PANEL, fg=C_TEXT,
                     font=("Segoe UI", 10, "bold")).pack(
                         anchor="w", pady=(0, 6))
        return content

    def _field_row(self, parent, label: str, var: tk.StringVar,
                   mode="file", filetypes=None, width=16):
        row = tk.Frame(parent, bg=C_PANEL)
        row.pack(fill="x", pady=3)
        tk.Label(row, text=label, bg=C_PANEL, fg=C_SUBTLE,
                 font=F_SMALL, width=width, anchor="w").pack(side="left")
        entry = tk.Entry(row, textvariable=var, font=F_BODY,
                         bg="#F8FAFC", fg=C_TEXT,
                         relief="solid", bd=1,
                         highlightbackground=C_BORDER,
                         highlightthickness=1)
        entry.pack(side="left", fill="x", expand=True, padx=(4, 6))

        def _browse():
            if mode == "dir":
                p = filedialog.askdirectory(title=f"Select {label}",
                                             mustexist=True)
            else:
                p = filedialog.askopenfilename(
                    title=f"Select {label}",
                    filetypes=filetypes or [("All files", "*.*")])
            if p:
                var.set(p)

        tk.Button(row, text="Browse", command=_browse,
                  bg=C_BG, fg=C_ACCENT, font=F_SMALL,
                  relief="solid", bd=1, padx=8, pady=2,
                  cursor="hand2").pack(side="left")
        return entry

    def _action_bar(self, parent, key: str, run_cmd, stop_cmd,
                    run_label="Run", extras=None):
        bar = tk.Frame(parent, bg=C_BG)
        bar.pack(fill="x", padx=26, pady=(6, 2))

        run_btn = tk.Button(
            bar, text=run_label,
            bg=C_ACCENT, fg="#FFFFFF",
            activebackground=C_ACCENT_H, activeforeground="#FFFFFF",
            font=("Segoe UI", 10, "bold"),
            relief="flat", padx=20, pady=7, cursor="hand2",
            command=run_cmd,
        )
        run_btn.pack(side="left")

        stop_btn = tk.Button(
            bar, text="  Stop  ",
            bg="#EF4444", fg="#FFFFFF",
            activebackground="#DC2626", activeforeground="#FFFFFF",
            font=("Segoe UI", 10, "bold"),
            relief="flat", padx=14, pady=7, cursor="hand2",
            state="disabled",
            command=stop_cmd,
        )
        stop_btn.pack(side="left", padx=(8, 0))

        if extras:
            for lbl, cmd in extras:
                tk.Button(bar, text=lbl, command=cmd,
                          bg=C_BG, fg=C_TEXT, font=F_BODY,
                          relief="solid", bd=1, padx=12, pady=6,
                          cursor="hand2").pack(side="left", padx=(8, 0))

        progress = ttk.Progressbar(
            bar, style="ETH.Horizontal.TProgressbar",
            mode="determinate", value=0, maximum=100, length=200)
        progress.pack(side="left", padx=(20, 0), pady=2)

        prog_lbl = tk.Label(bar, text="", bg=C_BG, fg=C_SUBTLE, font=F_SMALL)
        prog_lbl.pack(side="left", padx=(8, 0))

        self._panel_ui[key] = dict(
            run=run_btn, stop=stop_btn,
            progress=progress, lbl=prog_lbl)
        return bar

    def _terminal(self, parent, key: str, rows=14):
        wrap = tk.Frame(parent, bg=C_BG)
        wrap.pack(fill="both", expand=True, padx=26, pady=(4, 18))

        # header
        hdr = tk.Frame(wrap, bg=C_TERM_HDR)
        hdr.pack(fill="x")
        tk.Label(hdr, text=" Output", bg=C_TERM_HDR, fg="#94A3B8",
                 font=("Segoe UI", 8, "bold"),
                 pady=4, padx=6).pack(side="left")

        def _clear():
            t.configure(state="normal")
            t.delete("1.0", "end")
            t.configure(state="disabled")

        tk.Button(hdr, text="Clear", bg=C_TERM_HDR, fg="#94A3B8",
                  font=("Segoe UI", 8), bd=0, padx=8, pady=3,
                  activebackground="#263550", activeforeground="#FFFFFF",
                  cursor="hand2", relief="flat",
                  command=_clear).pack(side="right")

        # body
        body = tk.Frame(wrap, bg=C_TERM_BG)
        body.pack(fill="both", expand=True)

        sb = tk.Scrollbar(body, orient="vertical", bg=C_TERM_BG,
                          troughcolor=C_TERM_BG)
        sb.pack(side="right", fill="y")

        t = tk.Text(
            body, bg=C_TERM_BG, fg=C_TERM_FG,
            font=F_MONO, wrap="word", state="disabled",
            yscrollcommand=sb.set, bd=0,
            padx=10, pady=8,
            selectbackground=C_ACCENT,
            height=rows,
        )
        t.pack(side="left", fill="both", expand=True)
        sb.config(command=t.yview)

        t.tag_configure("info",    foreground=C_TERM_INFO)
        t.tag_configure("ok",      foreground=C_TERM_OK)
        t.tag_configure("error",   foreground=C_TERM_ERR)
        t.tag_configure("warn",    foreground=C_TERM_WARN)
        t.tag_configure("heading", foreground="#FFFFFF",
                         font=("Consolas", 10, "bold"))
        t.tag_configure("muted",   foreground=C_TERM_MUTE)

        self._panel_terms[key] = t
        self._panel_term_frames[key] = wrap
        return t

    # ── terminal write ────────────────────────────────────────────────

    def _term_write(self, term: tk.Text, msg: str, tag: str = ""):
        term.configure(state="normal")
        if tag:
            term.insert("end", msg + "\n", tag)
        else:
            term.insert("end", msg + "\n")
        term.see("end")
        term.configure(state="disabled")

    # ── running state ─────────────────────────────────────────────────

    def _set_running(self, key: str, running: bool):
        ui = self._panel_ui.get(key)
        if not ui:
            return
        ui["run"].configure(state="disabled" if running else "normal")
        ui["stop"].configure(state="normal" if running else "disabled")
        if running:
            ui["progress"].configure(mode="indeterminate")
            ui["progress"].start(12)
        else:
            ui["progress"].stop()
            ui["progress"].configure(mode="determinate", value=0)
            ui["lbl"].configure(text="")

    def _set_progress(self, key: str, value: int, maximum: int, label: str = ""):
        ui = self._panel_ui.get(key)
        if not ui:
            return
        ui["progress"].stop()
        ui["progress"].configure(mode="determinate",
                                  maximum=max(maximum, 1), value=value)
        if label:
            ui["lbl"].configure(text=label)

    def _set_status(self, msg: str):
        try:
            self._status_lbl.configure(text=msg)
        except Exception:
            pass

    # ── log queue polling ─────────────────────────────────────────────

    def _poll_log_queue(self):
        try:
            while True:
                item = self._log_queue.get_nowait()
                tag, payload = item
                if tag == "__done__":
                    self._dpp_print_bar.pack_forget()
                    self._on_worker_done(*payload)
                elif tag == "__progress__":
                    key, v, mx, lbl = payload
                    self._set_progress(key, v, mx, lbl)
                elif tag == "__status__":
                    self._set_status(payload)
                elif tag == "__populate_mp__":
                    self._mp_clear_all()
                    for path, duplex in payload:
                        self._mp_append_file(path, duplex)
                    if not self._mp_expanded.get():
                        self._mp_toggle()
                elif tag == "__print_confirm__":
                    count, names = payload
                    detail = f"{count} document(s): " + ",  ".join(names[:5])
                    if len(names) > 5:
                        detail += f"  … +{len(names)-5} more"
                    self._dpp_print_detail.config(text=detail)
                    self._dpp_print_bar.pack(
                        fill="x", padx=26, pady=(4, 0),
                        before=self._panel_term_frames["dpp"])
                else:
                    term = self._panel_terms.get(self._active_key)
                    if term:
                        self._term_write(term, payload, tag)
        except queue.Empty:
            pass
        self.root.after(80, self._poll_log_queue)

    def _on_worker_done(self, key: str, success: bool,
                        tool_name: str, log_lines: list[str]):
        self._set_running(key, False)
        _master_log.append_section(tool_name, log_lines)
        status = f"{tool_name} — done" if success else f"{tool_name} — errors"
        self._set_status(status)
        self._worker_thread = None

    # ── stop handler ──────────────────────────────────────────────────

    def _on_stop(self):
        _stop_event.set()
        self._log_queue.put(
            ("warn", "[USER] Stop requested — will halt after current step..."))
        self._set_status("Stopping...")
        ui = self._panel_ui.get(self._active_key)
        if ui:
            ui["stop"].configure(state="disabled")


# ═════════════════════════════════════════════════════════════════════
#  BOM FILLER PANEL
# ═════════════════════════════════════════════════════════════════════

    def _build_bom_panel(self, parent):
        self._section_header(
            parent,
            "Bom Filler",
            "Mark stock parts in the FFMPL sheet, then copy non-stock PDFs and DXFs to the target folder.")

        card = self._card(parent, "Configuration")

        self._bom_wb     = tk.StringVar()
        self._bom_target = tk.StringVar()

        self._field_row(card, "BOM Workbook", self._bom_wb,
                        mode="file",
                        filetypes=[("Excel", "*.xlsx *.xlsm *.xls"),
                                   ("All files", "*.*")])
        self._field_row(card, "Target Folder", self._bom_target, mode="dir")

        note_row = tk.Frame(card, bg=C_PANEL)
        note_row.pack(fill="x", pady=(8, 0))
        tk.Label(note_row,
                 text="The BOM workbook must be closed in Excel before running.",
                 bg=C_PANEL, fg=C_WARN, font=F_SMALL).pack(anchor="w")

        self._action_bar(parent, "bom", self._run_bom, self._on_stop,
                         run_label="  Run Bom Filler  ")

        self._terminal(parent, "bom", rows=15)

    def _run_bom(self):
        wb = self._bom_wb.get().strip()
        tgt = self._bom_target.get().strip()

        if not wb:
            messagebox.showwarning(APP_TITLE, "Please select a BOM workbook.")
            return
        if not os.path.isfile(wb):
            messagebox.showwarning(APP_TITLE, f"File not found:\n{wb}")
            return
        if not tgt:
            messagebox.showwarning(APP_TITLE, "Please select a target folder.")
            return
        if not os.path.isdir(tgt):
            messagebox.showwarning(APP_TITLE, f"Target folder not found:\n{tgt}")
            return

        confirmed = messagebox.askyesno(
            APP_TITLE,
            "Bom Filler requires the workbook to be CLOSED in Excel.\n\n"
            "Have you closed the workbook and are ready to proceed?",
        )
        if not confirmed:
            return

        self._active_key = "bom"
        _stop_event.clear()
        self._set_running("bom", True)
        self._set_status("Bom Filler — running...")

        t = threading.Thread(
            target=self._bom_worker, args=(wb, tgt), daemon=True)
        self._worker_thread = t
        t.start()

    def _bom_worker(self, wb_path: str, target: str):
        q = self._log_queue
        log_lines: list[str] = []
        success = False

        def emit(msg: str, tag: str = "info"):
            ts = datetime.now().strftime("%H:%M:%S")
            line = f"[{ts}]  {msg}"
            log_lines.append(line)
            q.put((tag, line))

        try:
            if xw is None:
                raise RuntimeError(
                    "xlwings is not installed. Run:  pip install xlwings")

            emit(f"Opening workbook: {Path(wb_path).name}", "heading")
            xw_app = xw.App(visible=False, add_book=False)
            xw_app.display_alerts = False
            xw_app.screen_updating = False

            try:
                wb = xw_app.books.open(wb_path)
            except Exception as e:
                raise RuntimeError(f"Could not open workbook: {e}")

            sheet_names = [s.name for s in wb.sheets]
            if BOM_SHEET_NAME not in sheet_names:
                wb.close(); xw_app.quit()
                raise RuntimeError(
                    f"Sheet '{BOM_SHEET_NAME}' not found.\n"
                    f"Available sheets: {sheet_names}")

            ws = wb.sheets[BOM_SHEET_NAME]
            DATA_START = 6

            last_row = ws.range(f"A{DATA_START}").end("down").row
            if last_row > 1_000_000:
                last_row = DATA_START

            raw = ws.range(f"A{DATA_START}:A{last_row}").value
            if raw is None:
                part_numbers = []
            elif not isinstance(raw, list):
                part_numbers = [raw]
            else:
                part_numbers = raw

            total = sum(1 for p in part_numbers if p and str(p).strip())
            emit(f"Found {total} part numbers to process.", "info")

            # ── PASS 1 ─────────────────────────────────────────────────
            emit("Pass 1  —  Stock Parts Check", "heading")

            stock_found = 0
            stock_flags: list[bool] = []
            done = 0

            for i, pn in enumerate(part_numbers):
                check_stop()
                if pn is None or not str(pn).strip():
                    stock_flags.append(False)
                    continue
                pn_s = str(pn).strip()
                done += 1
                q.put(("__progress__",
                       ("bom", done, total, f"Pass 1 — {done}/{total}")))

                is_stock = _bom_check_stock(pn_s)
                row = DATA_START + i
                if is_stock:
                    ws.range(f"B{row}").value = "X"
                    ws.range(f"G{row}").value = "S"
                    ws.range(f"H{row}").value = "S"
                    emit(f"  {pn_s:<28}  STOCK", "ok")
                    stock_found += 1
                    stock_flags.append(True)
                else:
                    stock_flags.append(False)

            emit(f"  {stock_found} stock  /  {done - stock_found} non-stock", "info")

            # ── PASS 2 ─────────────────────────────────────────────────
            emit("Pass 2  —  Non-Stock PDF + DXF Copy", "heading")

            tgt_path = Path(target)
            pdf_cp = pdf_ex = pdf_nf = 0
            dxf_cp = dxf_ex = dxf_nf = 0
            ns_done = 0
            ns_total = sum(1 for i, p in enumerate(part_numbers)
                           if p and str(p).strip()
                           and i < len(stock_flags) and not stock_flags[i])

            for i, pn in enumerate(part_numbers):
                check_stop()
                if pn is None or not str(pn).strip():
                    continue
                if i < len(stock_flags) and stock_flags[i]:
                    continue
                pn_s = str(pn).strip()
                ns_done += 1
                q.put(("__progress__",
                       ("bom", ns_done, max(ns_total, 1),
                        f"Pass 2 — {ns_done}/{ns_total}")))

                rev = _bom_find_revision(pn_s)
                row = DATA_START + i

                pdf_r = _bom_find_copy(pn_s, rev, tgt_path, "pdf")
                dxf_r = _bom_find_copy(pn_s, rev, tgt_path, "dxf")

                # concise status labels
                plbl = ("copied" if pdf_r is True
                        else "exists" if pdf_r is None else "not found")
                dlbl = ("copied" if dxf_r is True
                        else "exists" if dxf_r is None else "not found")

                if pdf_r is True:   pdf_cp += 1
                elif pdf_r is None: pdf_ex += 1
                else:               pdf_nf += 1

                if dxf_r is True:   dxf_cp += 1
                elif dxf_r is None: dxf_ex += 1
                else:               dxf_nf += 1

                has_issue = pdf_r is False or dxf_r is False
                tag = "warn" if has_issue else "ok"
                rv  = f" [{rev}]" if rev else ""
                emit(f"  {pn_s:<28}{rv:<6}  PDF: {plbl:<12}  DXF: {dlbl}", tag)

                if pdf_r is not False or dxf_r is not False:
                    ws.range(f"G{row}").value = "X"

            emit(f"  PDFs  copied {pdf_cp}  existed {pdf_ex}  not found {pdf_nf}",
                 "info")
            emit(f"  DXFs  copied {dxf_cp}  existed {dxf_ex}  not found {dxf_nf}",
                 "info")

            changes = stock_found + pdf_cp + pdf_ex + dxf_cp + dxf_ex
            if changes > 0:
                wb.save()
                emit(f"Workbook saved.", "ok")
            else:
                emit("No changes made.", "warn")

            wb.close()
            xw_app.quit()
            success = True
            emit("Done.", "ok")

        except StopRequested:
            emit("Stopped by user.", "warn")
        except Exception as e:
            emit(f"ERROR: {e}", "error")
            for ln in traceback.format_exc().splitlines():
                emit(ln, "error")
        finally:
            q.put(("__done__", ("bom", success, "Bom Filler", log_lines)))

    # ── CNC Column Marker ─────────────────────────────────────────────



# ═════════════════════════════════════════════════════════════════════
#  DOC PREP & PRINT PANEL
# ═════════════════════════════════════════════════════════════════════

    def _build_dpp_panel(self, parent):
        self._section_header(
            parent,
            "Doc Prep & Print",
            "Build a manufacturing print packet and send it to the printer (or save PDFs in Simulation Mode).")

        card = self._card(parent, "Configuration")

        self._dpp_folder  = tk.StringVar()
        self._dpp_printer = tk.StringVar(value=PREFERRED_PRINTER)
        self._dpp_sim     = tk.BooleanVar(value=False)
        self._dpp_plan: dict | None = None
        self._dpp_print_event = threading.Event()
        self._dpp_last_print_dir: Path | None = None   # cleaned up on next run

        self._field_row(card, "Job Folder", self._dpp_folder, mode="dir")

        # printer row
        pr_row = tk.Frame(card, bg=C_PANEL)
        pr_row.pack(fill="x", pady=3)
        tk.Label(pr_row, text="Printer", bg=C_PANEL, fg=C_SUBTLE,
                 font=F_SMALL, width=16, anchor="w").pack(side="left")
        tk.Entry(pr_row, textvariable=self._dpp_printer,
                 font=F_BODY, bg="#F8FAFC", fg=C_TEXT,
                 relief="solid", bd=1,
                 highlightbackground=C_BORDER,
                 highlightthickness=1).pack(
                     side="left", fill="x", expand=True, padx=(4, 6))
        tk.Button(pr_row, text="Detect",
                  command=self._dpp_detect_printer,
                  bg=C_BG, fg=C_ACCENT, font=F_SMALL,
                  relief="solid", bd=1, padx=8, pady=2,
                  cursor="hand2").pack(side="left")

        # simulation toggle
        sim_row = tk.Frame(card, bg=C_PANEL)
        sim_row.pack(fill="x", pady=(10, 2))
        tk.Checkbutton(
            sim_row,
            text="Simulation Mode  —  saves PDFs to folder instead of printing",
            variable=self._dpp_sim,
            bg=C_PANEL, fg=C_TEXT, font=F_BODY,
            activebackground=C_PANEL, selectcolor=C_PANEL,
        ).pack(anchor="w")

        self._action_bar(
            parent, "dpp",
            self._run_dpp, self._on_stop,
            run_label="  Build Plan  ",
            extras=[("Clear Plan", self._dpp_clear_plan),
                    ("Preview FWO", self._dpp_test_fwo_fill),
                    ("Preview BOM", self._dpp_preview_bom)],
        )

        # ── Confirm bar 1: plan review — hidden until plan is built ───────
        self._dpp_confirm_bar = tk.Frame(parent, bg="#EFF6FF",
                                         highlightbackground=C_ACCENT,
                                         highlightthickness=1)
        lbl_row = tk.Frame(self._dpp_confirm_bar, bg="#EFF6FF")
        lbl_row.pack(side="left", fill="x", expand=True, padx=(14, 8), pady=8)
        tk.Label(lbl_row, text="Plan ready — review the summary below, then confirm.",
                 bg="#EFF6FF", fg=C_ACCENT,
                 font=("Segoe UI", 10, "bold")).pack(anchor="w")
        self._dpp_confirm_detail = tk.Label(lbl_row, text="",
                                             bg="#EFF6FF", fg=C_TEXT,
                                             font=F_SMALL, justify="left")
        self._dpp_confirm_detail.pack(anchor="w")
        btn_row = tk.Frame(self._dpp_confirm_bar, bg="#EFF6FF")
        btn_row.pack(side="right", padx=14, pady=8)
        self._dpp_confirm_btn = tk.Button(
                  btn_row, text="Generate Documents",
                  bg=C_SUCCESS, fg="#FFFFFF",
                  activebackground="#15803D", activeforeground="#FFFFFF",
                  font=("Segoe UI", 10, "bold"),
                  relief="flat", padx=16, pady=6, cursor="hand2",
                  command=self._dpp_confirm)
        self._dpp_confirm_btn.pack(side="left")
        tk.Button(btn_row, text="Cancel",
                  bg=C_BG, fg=C_TEXT, font=F_BODY,
                  relief="solid", bd=1, padx=12, pady=5, cursor="hand2",
                  command=self._dpp_cancel_plan).pack(side="left", padx=(8, 0))

        # ── Confirm bar 2: print gate — hidden until PDFs are generated ───
        self._dpp_print_bar = tk.Frame(parent, bg="#FFF7ED",
                                        highlightbackground="#F59E0B",
                                        highlightthickness=1)
        lbl_row2 = tk.Frame(self._dpp_print_bar, bg="#FFF7ED")
        lbl_row2.pack(side="left", fill="x", expand=True, padx=(14, 8), pady=8)
        tk.Label(lbl_row2, text="Documents ready — confirm to send to printer.",
                 bg="#FFF7ED", fg="#92400E",
                 font=("Segoe UI", 10, "bold")).pack(anchor="w")
        self._dpp_print_detail = tk.Label(lbl_row2, text="",
                                           bg="#FFF7ED", fg=C_TEXT,
                                           font=F_SMALL, justify="left")
        self._dpp_print_detail.pack(anchor="w")
        btn_row2 = tk.Frame(self._dpp_print_bar, bg="#FFF7ED")
        btn_row2.pack(side="right", padx=14, pady=8)
        tk.Button(btn_row2, text="Send to Printer",
                  bg="#F59E0B", fg="#FFFFFF",
                  activebackground="#D97706", activeforeground="#FFFFFF",
                  font=("Segoe UI", 10, "bold"),
                  relief="flat", padx=16, pady=6, cursor="hand2",
                  command=self._dpp_confirm_print).pack(side="left")
        tk.Button(btn_row2, text="Cancel Print",
                  bg=C_BG, fg=C_TEXT, font=F_BODY,
                  relief="solid", bd=1, padx=12, pady=5, cursor="hand2",
                  command=self._dpp_cancel_print).pack(side="left", padx=(8, 0))

        # ── Manual Printing — collapsible section ─────────────────────
        self._mp_files: list[dict] = []
        self._mp_expanded = tk.BooleanVar(value=False)

        mp_toggle_bar = tk.Frame(parent, bg=C_BG)
        mp_toggle_bar.pack(fill="x", padx=26, pady=(8, 0))
        tk.Frame(mp_toggle_bar, bg=C_BORDER, height=1).pack(fill="x")
        self._mp_toggle_btn = tk.Button(
            mp_toggle_bar,
            text="▶  Manual Printing",
            bg=C_BG, fg=C_SUBTLE, font=("Segoe UI", 9, "bold"),
            relief="flat", anchor="w", padx=0, pady=4, cursor="hand2",
            command=self._mp_toggle)
        self._mp_toggle_btn.pack(anchor="w")

        self._mp_section = tk.Frame(parent, bg=C_BG)
        # (not packed yet — shown on toggle)

        # file list inner
        mp_card = self._card(self._mp_section, "", padx=26, pady=(0, 4))

        tb = tk.Frame(mp_card, bg=C_PANEL)
        tb.pack(fill="x", pady=(0, 6))
        tk.Label(tb, text="PDF Files", bg=C_PANEL, fg=C_TEXT,
                 font=("Segoe UI", 10, "bold")).pack(side="left")
        tk.Button(tb, text="Add PDFs", font=F_SMALL,
                  bg=C_ACCENT, fg="#FFFFFF", activebackground=C_ACCENT_H,
                  activeforeground="#FFFFFF", relief="flat", padx=10, pady=3,
                  cursor="hand2", command=self._mp_add_files).pack(side="right")
        tk.Button(tb, text="Clear All", font=F_SMALL,
                  bg=C_BG, fg=C_TEXT, relief="solid", bd=1, padx=8, pady=2,
                  cursor="hand2", command=self._mp_clear_all).pack(side="right", padx=(0, 6))

        # column headers
        hdr = tk.Frame(mp_card, bg=C_PANEL)
        hdr.pack(fill="x")
        tk.Label(hdr, text="File", bg=C_PANEL, fg=C_SUBTLE,
                 font=F_SMALL, anchor="w").pack(side="left", fill="x", expand=True)
        tk.Label(hdr, text="Duplex", bg=C_PANEL, fg=C_SUBTLE,
                 font=F_SMALL, width=7).pack(side="left")
        tk.Label(hdr, text="", bg=C_PANEL, width=10).pack(side="left")
        tk.Label(hdr, text="", bg=C_PANEL, width=4).pack(side="left")
        tk.Frame(mp_card, bg=C_BORDER, height=1).pack(fill="x", pady=(2, 4))

        list_wrap = tk.Frame(mp_card, bg=C_PANEL, height=180)
        list_wrap.pack(fill="x")
        list_wrap.pack_propagate(False)
        mp_canvas = tk.Canvas(list_wrap, bg=C_PANEL, highlightthickness=0)
        mp_sb = ttk.Scrollbar(list_wrap, orient="vertical", command=mp_canvas.yview)
        self._mp_inner = tk.Frame(mp_canvas, bg=C_PANEL)
        self._mp_canvas = mp_canvas
        self._mp_inner.bind("<Configure>",
            lambda *_: mp_canvas.configure(scrollregion=mp_canvas.bbox("all")))
        mp_canvas.create_window((0, 0), window=self._mp_inner, anchor="nw")
        mp_canvas.configure(yscrollcommand=mp_sb.set)
        mp_sb.pack(side="right", fill="y")
        mp_canvas.pack(side="left", fill="both", expand=True)

        self._mp_empty_lbl = tk.Label(self._mp_inner,
            text="No files added — click  Add PDFs  to get started.",
            bg=C_PANEL, fg=C_SUBTLE, font=F_SMALL)
        self._mp_empty_lbl.pack(pady=14)

        # Print All row
        pa = tk.Frame(self._mp_section, bg=C_BG)
        pa.pack(fill="x", padx=26, pady=(2, 6))
        tk.Button(pa, text="  Print All  ",
                  bg=C_SUCCESS, fg="#FFFFFF",
                  activebackground="#15803D", activeforeground="#FFFFFF",
                  font=("Segoe UI", 10, "bold"),
                  relief="flat", padx=16, pady=6, cursor="hand2",
                  command=self._mp_print_all).pack(side="left")
        tk.Label(pa, text="Sends all files in order with 3.5 s gap — uses the printer above.",
                 bg=C_BG, fg=C_SUBTLE, font=F_SMALL).pack(side="left", padx=10)

        self._terminal(parent, "dpp", rows=15)

    def _dpp_detect_printer(self):
        if win32print is None:
            messagebox.showwarning(APP_TITLE,
                "pywin32 not available — cannot detect printers.")
            return
        try:
            flags = (win32print.PRINTER_ENUM_LOCAL
                     | win32print.PRINTER_ENUM_CONNECTIONS)
            printers = sorted({p[2] for p in win32print.EnumPrinters(flags)})
            if PREFERRED_PRINTER in printers:
                self._dpp_printer.set(PREFERRED_PRINTER)
                return
            choice = self._pick_from_list("Select Printer", printers)
            if choice:
                self._dpp_printer.set(choice)
        except Exception as e:
            messagebox.showwarning(APP_TITLE,
                f"Could not enumerate printers:\n{e}")

    def _pick_from_list(self, title: str, items: list[str],
                        prompt: str = "") -> str | None:
        top = tk.Toplevel(self.root)
        top.title(title)
        top.geometry("580x380")
        top.configure(bg=C_BG)
        top.grab_set()
        top.transient(self.root)
        result: dict[str, str | None] = {"v": None}

        tk.Label(top, text=title, bg=C_BG, fg=C_TEXT,
                 font=("Segoe UI", 12, "bold")).pack(
                     pady=(16, 4), padx=16, anchor="w")
        if prompt:
            tk.Label(top, text=prompt, bg=C_BG, fg=C_SUBTLE,
                     font=F_BODY).pack(pady=(0, 4), padx=16, anchor="w")

        body = tk.Frame(top, bg=C_BORDER)
        body.pack(fill="both", expand=True, padx=16, pady=4)
        inner = tk.Frame(body, bg=C_PANEL)
        inner.pack(fill="both", expand=True, padx=1, pady=1)

        sb2 = tk.Scrollbar(inner)
        sb2.pack(side="right", fill="y")
        lb = tk.Listbox(inner, yscrollcommand=sb2.set,
                         font=F_BODY, bd=0, highlightthickness=0,
                         selectbackground=C_ACCENT,
                         selectforeground="#FFFFFF")
        for item in items:
            lb.insert("end", item)
        lb.pack(side="left", fill="both", expand=True, padx=4, pady=4)
        sb2.config(command=lb.yview)

        bar2 = tk.Frame(top, bg=C_BG)
        bar2.pack(fill="x", padx=16, pady=(4, 14))

        def _ok():
            sel = lb.curselection()
            if sel:
                result["v"] = items[sel[0]]
            top.destroy()

        tk.Button(bar2, text="OK",
                  bg=C_ACCENT, fg="#FFFFFF",
                  font=("Segoe UI", 10, "bold"),
                  relief="flat", padx=16, pady=5,
                  cursor="hand2",
                  command=_ok).pack(side="left")
        tk.Button(bar2, text="Cancel",
                  bg=C_BG, fg=C_TEXT, font=F_BODY,
                  relief="solid", bd=1, padx=12, pady=5,
                  cursor="hand2",
                  command=top.destroy).pack(side="left", padx=(8, 0))

        top.wait_window()
        return result["v"]

    def _dpp_clear_plan(self):
        self._dpp_plan = None
        if hasattr(self, "_dpp_confirm_bar"):
            self._dpp_confirm_bar.pack_forget()
        term = self._panel_terms.get("dpp")
        if term:
            self._term_write(term, "Plan cleared.", "info")

    def _run_dpp(self):
        folder  = self._dpp_folder.get().strip()
        printer = self._dpp_printer.get().strip()
        sim     = self._dpp_sim.get()

        if not folder:
            messagebox.showwarning(APP_TITLE, "Please select a job folder.")
            return
        if not os.path.isdir(folder):
            messagebox.showwarning(APP_TITLE, f"Folder not found:\n{folder}")
            return
        if not sim and not printer:
            messagebox.showwarning(APP_TITLE,
                "Please enter a printer name, or enable Simulation Mode.")
            return

        self._active_key = "dpp"
        term = self._panel_terms.get("dpp")

        # Build plan on main thread (dialogs may appear for variant/file selection)
        try:
            if term:
                self._term_write(term, f"Building plan for: {folder}", "info")
            plan = dpp_build_plan(Path(folder), self)
        except Exception as e:
            if term:
                self._term_write(term, f"Plan error: {e}", "error")
            return

        # Write summary to terminal
        summary = dpp_make_summary(plan, printer, sim)
        if term:
            self._term_write(term, summary, "info")

        # Store plan + params and show inline confirm bar
        self._dpp_plan    = plan
        self._dpp_pending_printer = printer
        self._dpp_pending_sim     = sim

        cnc_d = sum(1 for p in plan["cnc"]
                    if _dpp_classify_cnc(p) == "duplex")
        cnc_s = len(plan["cnc"]) - cnc_d
        detail = (f"{len(plan['cnc'])} CNC  •  {len(plan['flats'])} Flats  •  "
                  f"{len(plan['assemblies'])} Assemblies"
                  + (f"  •  Simulation (save to folder)" if sim else f"  •  {printer}"))
        self._dpp_confirm_detail.config(text=detail)
        self._dpp_confirm_btn.config(
            text="  Start Simulation  " if sim else "  Generate Documents  ")
        self._dpp_confirm_bar.pack(fill="x", padx=26, pady=(4, 0),
                                    before=self._panel_term_frames["dpp"])

    def _dpp_confirm(self):
        plan    = self._dpp_plan
        printer = getattr(self, "_dpp_pending_printer", "")
        sim     = getattr(self, "_dpp_pending_sim", False)
        if plan is None:
            return

        # Resolve CNC folder on the main thread so we can prompt if missing
        bom_path = plan.get("bom")
        if bom_path:
            cnc_folder = _cnc_find_cnc_folder_for_bom(Path(bom_path))
            if cnc_folder is None:
                chosen = filedialog.askdirectory(
                    title="205 CNC folder not found — select it manually",
                    mustexist=True,
                )
                cnc_folder = Path(chosen) if chosen else None
            plan["cnc_mark_folder"] = cnc_folder   # None = skip marking

        self._dpp_confirm_bar.pack_forget()
        _stop_event.clear()
        self._dpp_print_event.clear()
        self._set_running("dpp", True)
        self._set_status("Doc Prep — running...")
        t = threading.Thread(
            target=self._dpp_worker,
            args=(plan, printer, sim),
            daemon=True,
        )
        self._worker_thread = t
        t.start()

    def _dpp_cancel_plan(self):
        self._dpp_plan = None
        self._dpp_confirm_bar.pack_forget()
        term = self._panel_terms.get("dpp")
        if term:
            self._term_write(term, "Plan cancelled.", "warn")

    def _dpp_confirm_print(self):
        """User clicked 'Send to Printer' on the print confirm bar."""
        self._dpp_print_bar.pack_forget()
        self._dpp_print_event.set()

    def _dpp_cancel_print(self):
        """User clicked 'Cancel Print' — stop the worker."""
        self._dpp_print_bar.pack_forget()
        _stop_event.set()   # worker will raise StopRequested on next check_stop()

    def _dpp_test_fwo_fill(self):
        """Fill the FWO with PRF data and open the result — for coordinate tuning."""
        folder = self._dpp_folder.get().strip()
        if not folder or not os.path.isdir(folder):
            messagebox.showwarning(APP_TITLE,
                "Select a job folder first so the FWO and PRF can be located.")
            return
        term = self._panel_terms.get("dpp")
        try:
            if term:
                self._term_write(term, "Preview FWO: building plan...", "info")
            plan = dpp_build_plan(Path(folder), self)
            prf_data = plan.get("prf_data")
            fwo_path = plan.get("fwo")
            if not fwo_path:
                messagebox.showwarning(APP_TITLE, "No FWO PDF found in the job folder.")
                return
            if not prf_data:
                messagebox.showwarning(APP_TITLE, "No PRF data found — cannot fill FWO.")
                return
            logs_dir = exe_dir() / "logs"
            logs_dir.mkdir(exist_ok=True)
            out = logs_dir / "FWO_preview.pdf"
            filled = _dpp_fill_fwo(fwo_path, prf_data)
            shutil.copy2(str(filled), str(out))
            Path(filled).unlink(missing_ok=True)
            if term:
                self._term_write(term, f"FWO preview saved: {out}", "ok")
            os.startfile(str(out))
        except Exception as e:
            if term:
                self._term_write(term, f"FWO preview error: {e}", "error")
            else:
                messagebox.showerror(APP_TITLE, str(e))

    def _dpp_preview_bom(self):
        """Run CNC marking, then export the BOM to PDF and open it for review."""
        folder = self._dpp_folder.get().strip()
        if not folder or not os.path.isdir(folder):
            messagebox.showwarning(APP_TITLE,
                "Select a job folder first so the BOM can be located.")
            return
        term = self._panel_terms.get("dpp")

        def log(msg, tag="info"):
            if term:
                self._term_write(term, msg, tag)

        try:
            log("Preview BOM: locating plan...", "info")
            plan = dpp_build_plan(Path(folder), self)
            bom_path = plan.get("bom")
            if not bom_path:
                messagebox.showwarning(APP_TITLE, "No BOM Excel file found in the job folder.")
                return

            # ── Run CNC marking first ─────────────────────────────────
            cnc_folder = _cnc_find_cnc_folder_for_bom(bom_path)
            if cnc_folder is None:
                chosen = filedialog.askdirectory(
                    title="205 CNC folder not found — select it manually",
                    mustexist=True,
                )
                cnc_folder = Path(chosen) if chosen else None

            if cnc_folder:
                log(f"Preview BOM: running CNC marker on {cnc_folder.name}...", "info")
                try:
                    matched, unmatched = _cnc_do_mark(bom_path, cnc_folder, log)
                    log(f"CNC: marked {matched} row(s).", "ok")
                    if unmatched:
                        log(f"CNC: {len(unmatched)} row(s) with no CNC file:", "warn")
                        for pn in unmatched:
                            log(f"  - {pn}", "warn")
                except Exception as e:
                    log(f"CNC marking failed ({e}) — exporting BOM without marks.", "warn")
            else:
                log("Preview BOM: CNC folder skipped.", "warn")

            # ── Export BOM to PDF ─────────────────────────────────────
            logs_dir = exe_dir() / "logs"
            logs_dir.mkdir(exist_ok=True)
            out = logs_dir / "BOM_preview.pdf"
            log(f"Preview BOM: exporting {bom_path.name}...", "info")
            excel = _dpp_get_excel()
            try:
                wb = excel.Workbooks.Open(str(bom_path))
                active = wb.ActiveSheet
                ps = active.PageSetup
                ps.Orientation = 2
                ps.Zoom = False
                ps.FitToPagesWide = 1
                ps.FitToPagesTall = False
                active.ExportAsFixedFormat(0, str(out))
                wb.Close(False)
            finally:
                try:
                    excel.Quit()
                except Exception:
                    pass
            log(f"BOM preview saved: {out}", "ok")
            os.startfile(str(out))
        except Exception as e:
            log(f"BOM preview error: {e}", "error")

    def _dpp_worker(self, plan: dict, printer_name: str, simulation: bool):
        q = self._log_queue
        log_lines: list[str] = []
        success = False

        def emit(msg: str, tag: str = "info"):
            ts = datetime.now().strftime("%H:%M:%S")
            line = f"[{ts}]  {msg}"
            log_lines.append(line)
            q.put((tag, line))

        if pythoncom:
            pythoncom.CoInitialize()

        fwo_filled: Path | None = None
        out_dir:   Path | None = None
        try:
            # ── Clean up previous print run's temp folder (safe now — spooling is done)
            prev = self._dpp_last_print_dir
            if prev and prev.exists():
                try:
                    shutil.rmtree(str(prev), ignore_errors=True)
                except Exception:
                    pass
                self._dpp_last_print_dir = None

            # ── Determine output directory ────────────────────────────
            job_name = safe_name(Path(plan["job_folder"]).name)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            if simulation:
                out_dir = exe_dir() / f"Simulated_Print_Output_{job_name}_{ts}"
                out_dir.mkdir(parents=True, exist_ok=True)
                emit(f"Simulation output: {out_dir}", "info")
            else:
                out_dir = Path(tempfile.mkdtemp(prefix="ETH_print_"))

            # ── Fill FWO with PRF data ─────────────────────────────────
            prf_data = plan.get("prf_data")
            if prf_data and plan.get("fwo"):
                try:
                    fwo_filled = _dpp_fill_fwo(plan["fwo"], prf_data)
                    plan["fwo_filled"] = fwo_filled
                    emit(f"FWO filled: {prf_data['job_no']}  {prf_data['enclosure']}", "ok")
                except Exception as e:
                    emit(f"FWO fill skipped ({e}) — using blank original", "warn")
                    plan["fwo_filled"] = None
            else:
                plan["fwo_filled"] = None

            # ── CNC Column Marker — runs before BOM is exported to PDF ──
            cnc_folder = plan.get("cnc_mark_folder")
            if cnc_folder is None:
                emit("CNC Column Marker: no CNC folder selected — skipping.", "warn")
            else:
                bom_path = Path(plan["bom"])
                emit(f"CNC Column Marker: scanning {cnc_folder.name}...", "heading")
                try:
                    matched, unmatched = _cnc_do_mark(bom_path, cnc_folder, emit)
                    emit(f"CNC: marked {matched} row(s).", "ok")
                    if unmatched:
                        emit(f"CNC: {len(unmatched)} row(s) with no CNC file:", "warn")
                        for pn in unmatched:
                            emit(f"  - {pn}", "warn")
                except Exception as e:
                    emit(f"CNC Column Marker failed ({e}) — continuing without marks.", "warn")

            # ── Phase 1: Generate all PDFs ─────────────────────────────
            sections = dpp_build_sections(plan, out_dir)
            gen_total = len(sections)

            for idx, (title, func) in enumerate(sections):
                check_stop()
                q.put(("__progress__",
                       ("dpp", idx, gen_total, f"{idx+1}/{gen_total}: {title}")))
                emit(f"  ▶  {title}", "heading")
                try:
                    func()
                    emit(f"     ✓  {title}", "ok")
                except StopRequested:
                    raise
                except Exception as e:
                    emit(f"     ✗  {title}: {e}", "error")
                    emit("     Skipping...", "warn")

            # ── Phase 2: Print (print mode only) ──────────────────────
            if not simulation:
                pdf_files = sorted(out_dir.glob("*.pdf"))
                print_total = len(pdf_files)
                emit(f"All {print_total} documents generated. Waiting for print confirmation...", "info")

                # Populate Manual Printing list with generated PDFs + duplex flags
                mp_jobs = [
                    (pdf, bool(re.search(r'^\d+_CNC_', pdf.stem)
                               and "CNC_Simplex" not in pdf.stem))
                    for pdf in pdf_files
                ]
                q.put(("__populate_mp__", mp_jobs))

                q.put(("__print_confirm__",
                       (print_total, [p.stem for p in pdf_files])))

                # Pause here until the user clicks "Send to Printer" or Stop
                while not self._dpp_print_event.is_set():
                    check_stop()
                    time.sleep(0.15)

                emit("Print confirmed — sending to printer...", "info")
                if win32print and printer_name:
                    try:
                        win32print.SetDefaultPrinter(printer_name)
                    except Exception:
                        pass

                # Build ordered job list with duplex flag per file
                print_jobs = [
                    (pdf, bool(re.search(r'^\d+_CNC_', pdf.stem)
                               and "CNC_Simplex" not in pdf.stem))
                    for pdf in pdf_files
                ]

                for i, (pdf, _) in enumerate(print_jobs):
                    emit(f"  ▶  {pdf.name}", "heading")
                    q.put(("__progress__",
                           ("dpp", gen_total + i, gen_total + print_total,
                            f"Printing {i+1}/{print_total}: {pdf.stem}")))

                # Try single-instance Acrobat COM session first
                emit("Opening Acrobat COM session...", "info")
                com_ok = _dpp_acrobat_print_all_com(print_jobs, printer_name, emit)

                if not com_ok:
                    # Fallback: per-document subprocess (one Acrobat per file)
                    emit("COM unavailable — falling back to subprocess per document", "warn")
                    acrobat_exe = _find_acrobat()
                    if not acrobat_exe:
                        emit("Acrobat not found — using default handler (no duplex control)", "warn")
                    for i, (pdf, is_duplex) in enumerate(print_jobs):
                        check_stop()
                        try:
                            status = _dpp_acrobat_print(pdf, printer_name, is_duplex, acrobat_exe)
                            emit(f"     ✓  {status}", "ok")
                        except Exception as e:
                            emit(f"     ✗  {e}", "error")
                        if i < print_total - 1:
                            _dpp_wait_spooler_stable(printer_name)

                q.put(("__progress__",
                       ("dpp", gen_total + print_total, gen_total + print_total, "Done")))
                emit("Print sequence complete.", "ok")
            else:
                q.put(("__progress__", ("dpp", gen_total, gen_total, "Done")))
                emit(f"Simulation complete. Output: {out_dir}", "ok")

            success = True

        except StopRequested:
            emit("Stopped by user.", "warn")
        except Exception as e:
            emit(f"FATAL: {e}", "error")
            for ln in traceback.format_exc().splitlines():
                emit(ln, "error")
        finally:
            if fwo_filled and fwo_filled.exists():
                try:
                    fwo_filled.unlink()
                except Exception:
                    pass
            # Defer temp print dir cleanup to next run so the spooler can finish reading.
            # Simulation folders are kept permanently (user reviews them).
            if not simulation and out_dir and out_dir.exists():
                self._dpp_last_print_dir = out_dir
            if pythoncom:
                try:
                    pythoncom.CoUninitialize()
                except Exception:
                    pass
            q.put(("__done__", ("dpp", success, "Doc Prep & Print", log_lines)))


# ═════════════════════════════════════════════════════════════════════
#  FILE LOGGER PANEL
# ═════════════════════════════════════════════════════════════════════

    def _build_fl_panel(self, parent):
        self._section_header(
            parent,
            "File Logger",
            "Track SolidWorks part files across job folders — find gaps, orphans, and next available numbers.")

        script_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "tools", "File Logger", "parts_tracker.py")

        # ── Launch card ───────────────────────────────────────────────
        card = self._card(parent, "Launch")

        desc = (
            "Parts Tracker is a standalone window that scans your network job folders, "
            "maintains a local database of SolidWorks part numbers, and helps you find "
            "the next available number in any category."
        )
        tk.Label(card, text=desc, bg=C_PANEL, fg=C_SUBTLE,
                 font=F_BODY, wraplength=680, justify="left",
                 anchor="w").pack(fill="x", pady=(0, 12))

        btn_row = tk.Frame(card, bg=C_PANEL)
        btn_row.pack(fill="x")

        self._fl_status_var = tk.StringVar(value="")
        self._fl_proc: subprocess.Popen | None = None

        def _launch():
            # If already running, bring focus by doing nothing (process manages its own window)
            if self._fl_proc is not None and self._fl_proc.poll() is None:
                self._fl_status_var.set("Already running.")
                return
            if not os.path.isfile(script_path):
                self._fl_status_var.set(f"Script not found: {script_path}")
                return
            try:
                self._fl_proc = subprocess.Popen(
                    [sys.executable, script_path],
                    cwd=os.path.dirname(script_path),
                    creationflags=subprocess.CREATE_NO_WINDOW if hasattr(subprocess, "CREATE_NO_WINDOW") else 0,
                )
                self._fl_status_var.set("Launched.")
            except Exception as exc:
                self._fl_status_var.set(f"Error: {exc}")

        tk.Button(btn_row, text="  Launch Parts Tracker  ",
                  font=("Segoe UI", 10, "bold"),
                  bg=C_ACCENT, fg="white",
                  activebackground=C_ACCENT_H, activeforeground="white",
                  relief="flat", cursor="hand2", pady=6,
                  command=_launch).pack(side="left")

        tk.Label(btn_row, textvariable=self._fl_status_var,
                 bg=C_PANEL, fg=C_SUBTLE, font=F_SMALL).pack(side="left", padx=12)

        # ── Features card ─────────────────────────────────────────────
        feat_card = self._card(parent, "Features")
        for bullet in [
            "Scans .sldprt and .sldasm files across all job folders on the network drive.",
            "Per-user prefix filtering — see only your own parts by default.",
            "Gap-aware next-number suggestions — reuses the lowest available number first.",
            "Orphan detection: identifies part files not tracked in any job folder.",
            "Archive classification: separates parts inside archive folders from active ones.",
            "Live filesystem watcher refreshes automatically when files change on disk.",
        ]:
            row = tk.Frame(feat_card, bg=C_PANEL)
            row.pack(fill="x", pady=2)
            tk.Label(row, text="•", bg=C_PANEL, fg=C_ACCENT,
                     font=F_BODY).pack(side="left", padx=(0, 8))
            tk.Label(row, text=bullet, bg=C_PANEL, fg=C_TEXT,
                     font=F_BODY, anchor="w").pack(side="left")

        # ── Requirements card ─────────────────────────────────────────
        req_card = self._card(parent, "Requirements")
        for req, note in [
            ("Python 3.10+",  "with PyQt6, openpyxl, and requests installed"),
            ("Everything",    "HTTP Server must be enabled on port 8080 (voidtools.com)"),
            ("Network drive", r"Z:\FOXFAB_DATA\ENGINEERING\2 JOBS must be accessible"),
        ]:
            row = tk.Frame(req_card, bg=C_PANEL)
            row.pack(fill="x", pady=3)
            tk.Label(row, text=req, bg=C_PANEL, fg=C_TEXT,
                     font=("Segoe UI", 10, "bold"), width=16,
                     anchor="w").pack(side="left")
            tk.Label(row, text=note, bg=C_PANEL, fg=C_SUBTLE,
                     font=F_BODY, anchor="w").pack(side="left")


# ═════════════════════════════════════════════════════════════════════
#  SW BATCH UPDATE PANEL
# ═════════════════════════════════════════════════════════════════════

    def _build_sw_panel(self, parent):
        self._section_header(
            parent,
            "SW Batch Update",
            "Run the SolidWorks macro directly to update custom properties and export DXFs.")

        macro_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "tools", "SolidworksBatchUpdate", "SoldworksBatchUpdate.swp")

        # ── How to Run card ───────────────────────────────────────
        card = self._card(parent, "How to Run")

        steps = [
            ("1", "Open SolidWorks 2023."),
            ("2", "Go to  Tools  →  Macro  →  Run..."),
            ("3", "Navigate to the macro file shown below and click Open."),
            ("4", "The macro dialog will appear — fill in your options and click Run."),
        ]
        for num, text in steps:
            row = tk.Frame(card, bg=C_PANEL)
            row.pack(fill="x", pady=3)
            tk.Label(row, text=num, bg=C_ACCENT, fg="white",
                     font=("Segoe UI", 9, "bold"),
                     width=2, anchor="center").pack(side="left", padx=(0, 10))
            tk.Label(row, text=text, bg=C_PANEL, fg=C_TEXT,
                     font=F_BODY, anchor="w").pack(side="left")

        # ── Macro Location card ───────────────────────────────────
        loc_card = self._card(parent, "Macro File Location")

        path_row = tk.Frame(loc_card, bg=C_PANEL)
        path_row.pack(fill="x")

        path_var = tk.StringVar(value=macro_path)
        path_entry = tk.Entry(path_row, textvariable=path_var, font=F_BODY,
                              bg="#F8FAFC", fg=C_TEXT,
                              relief="solid", bd=1, state="readonly")
        path_entry.pack(side="left", fill="x", expand=True, padx=(0, 8))

        def _open_folder():
            folder = os.path.dirname(macro_path)
            if os.path.isdir(folder):
                os.startfile(folder)

        tk.Button(path_row, text="Open Folder", font=F_BODY,
                  bg=C_PANEL, fg=C_ACCENT, relief="solid", bd=1,
                  cursor="hand2", command=_open_folder).pack(side="left")

        # ── What it does card ─────────────────────────────────────
        info_card = self._card(parent, "What the Macro Does")
        for bullet in [
            "Updates  DrawnBy  on all part/assembly files in a chosen folder.",
            "Updates  DwgDrawnBy  and re-stamps all standard drawing properties.",
            "Exports sheet metal flat-pattern DXFs to a chosen destination folder.",
            "Optionally skips files whose name starts with  003-.",
        ]:
            row = tk.Frame(info_card, bg=C_PANEL)
            row.pack(fill="x", pady=2)
            tk.Label(row, text="•", bg=C_PANEL, fg=C_ACCENT,
                     font=F_BODY).pack(side="left", padx=(0, 8))
            tk.Label(row, text=bullet, bg=C_PANEL, fg=C_TEXT,
                     font=F_BODY, anchor="w").pack(side="left")

    # ── How to Use panel ──────────────────────────────────────────────

    def _build_htu_panel(self, parent):
        self._section_header(
            parent,
            "How to Use",
            "Engineering Tool Hub — user manual.")

        pdf_path = Path(exe_dir()) / "Engineering_Tool_Hub.pdf"

        if not pdf_path.is_file() or _fitz is None:
            card = self._card(parent, "User Manual")
            msg = ("Engineering_Tool_Hub.pdf not found."
                   if not pdf_path.is_file()
                   else "PyMuPDF (fitz) is not available.")
            tk.Label(card, text=msg, bg=C_PANEL, fg="#EF4444",
                     font=F_BODY).pack(anchor="w")
            return

        # state
        doc = _fitz.open(str(pdf_path))
        total_pages = len(doc)
        state = {"page": 0, "photo": None}

        # ── nav bar ───────────────────────────────────────────────────
        nav = tk.Frame(parent, bg=C_BG, pady=4)
        nav.pack(fill="x", padx=26)

        btn_prev = tk.Button(nav, text="◀  Prev", font=F_BODY,
                             bg=C_PANEL, fg=C_ACCENT, relief="solid", bd=1,
                             cursor="hand2", width=9)
        btn_prev.pack(side="left")

        page_lbl = tk.Label(nav, text=f"Page 1 / {total_pages}",
                            bg=C_BG, fg=C_TEXT, font=F_BODY, width=16)
        page_lbl.pack(side="left", padx=12)

        btn_next = tk.Button(nav, text="Next  ▶", font=F_BODY,
                             bg=C_PANEL, fg=C_ACCENT, relief="solid", bd=1,
                             cursor="hand2", width=9)
        btn_next.pack(side="left")

        # ── scrollable canvas ─────────────────────────────────────────
        canvas_frame = tk.Frame(parent, bg=C_BG)
        canvas_frame.pack(fill="both", expand=True, padx=26, pady=(4, 12))

        vbar = tk.Scrollbar(canvas_frame, orient="vertical")
        vbar.pack(side="right", fill="y")

        canvas = tk.Canvas(canvas_frame, bg="#F1F5F9",
                           highlightthickness=0,
                           yscrollcommand=vbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        vbar.config(command=canvas.yview)

        canvas.bind("<MouseWheel>",
                    lambda e: canvas.yview_scroll(-1 * (e.delta // 120), "units"))

        img_id = canvas.create_image(0, 0, anchor="nw")

        def _render(page_idx: int):
            page = doc[page_idx]
            pix = page.get_pixmap(matrix=_fitz.Matrix(1.5, 1.5))
            png_b64 = base64.b64encode(pix.tobytes("png"))
            photo = tk.PhotoImage(data=png_b64)
            state["photo"] = photo           # keep reference alive
            canvas.itemconfig(img_id, image=photo)
            canvas.config(scrollregion=(0, 0, pix.width, pix.height))
            canvas.yview_moveto(0)
            page_lbl.config(text=f"Page {page_idx + 1} / {total_pages}")
            btn_prev.config(state="normal" if page_idx > 0 else "disabled")
            btn_next.config(state="normal" if page_idx < total_pages - 1 else "disabled")

        def _go_prev():
            if state["page"] > 0:
                state["page"] -= 1
                _render(state["page"])

        def _go_next():
            if state["page"] < total_pages - 1:
                state["page"] += 1
                _render(state["page"])

        btn_prev.config(command=_go_prev)
        btn_next.config(command=_go_next)

        _render(0)

    # ── SW Batch PDF Export panel ──────────────────────────────────────

    def _build_swpdf_panel(self, parent):
        self._section_header(
            parent,
            "SW Batch PDF Export",
            "Use SolidWorks Task Scheduler to batch-export drawings as PDFs.")

        # ── How to Use card ───────────────────────────────────────────
        card = self._card(parent, "How to Use Task Scheduler")

        steps = [
            ("1", "Open SolidWorks Task Scheduler  (Start → SolidWorks Tools → Task Scheduler)."),
            ("2", "Click  Export Files  in the left panel."),
            ("3", "Under Task Details, set Output Format to  Adobe Portable Document Format (*.pdf)."),
            ("4", "Click  Add Files  or  Add Folder  and select the drawings you want to export."),
            ("5", "Under Output Folder, choose where the PDFs should be saved."),
            ("6", "Set Schedule to  Run Now  (or pick a future time)."),
            ("7", "Click  Add Task, then  Run  (or wait for the scheduled time)."),
        ]
        for num, text in steps:
            row = tk.Frame(card, bg=C_PANEL)
            row.pack(fill="x", pady=3)
            tk.Label(row, text=num, bg=C_ACCENT, fg="white",
                     font=("Segoe UI", 9, "bold"),
                     width=2, anchor="center").pack(side="left", padx=(0, 10))
            tk.Label(row, text=text, bg=C_PANEL, fg=C_TEXT,
                     font=F_BODY, anchor="w", wraplength=820,
                     justify="left").pack(side="left", fill="x", expand=True)

        # ── Tips card ─────────────────────────────────────────────────
        tips_card = self._card(parent, "Tips")
        for bullet in [
            "SolidWorks does NOT need to be open — Task Scheduler runs independently.",
            "Use  Add Folder  to queue an entire folder of drawings at once.",
            "Enable  Include sub-folders  to recurse into sub-directories.",
            "Paper size and orientation are taken from the drawing sheet format, not set here.",
        ]:
            row = tk.Frame(tips_card, bg=C_PANEL)
            row.pack(fill="x", pady=2)
            tk.Label(row, text="•", bg=C_PANEL, fg=C_ACCENT,
                     font=F_BODY).pack(side="left", padx=(0, 8))
            tk.Label(row, text=bullet, bg=C_PANEL, fg=C_TEXT,
                     font=F_BODY, anchor="w").pack(side="left")

    # ── Training Materials panel ───────────────────────────────────────

    def _build_tm_panel(self, parent):
        self._section_header(
            parent,
            "Training Materials",
            "Print reference documents and open design guidelines.")

        pkg_dir  = Path(exe_dir()) / "tools" / "EngineeringDesignPackage"
        tips_doc = Path(exe_dir()) / "tools" / "FoxFab_Design_Tips.docx"

        # ── Engineering Design Package card ───────────────────────────
        pkg_card = self._card(parent, "Engineering Design Package")

        pdfs = sorted(pkg_dir.glob("*.pdf")) if pkg_dir.is_dir() else []

        if pdfs:
            desc = tk.Label(pkg_card,
                            text=(f"{len(pdfs)} reference documents will be sent to "
                                  f"{PREFERRED_PRINTER.split(chr(92))[-1]}  —  double-sided."),
                            bg=C_PANEL, fg=C_SUBTLE, font=F_BODY, anchor="w", wraplength=760)
            desc.pack(fill="x", pady=(0, 8))

            for pdf in pdfs:
                row = tk.Frame(pkg_card, bg=C_PANEL)
                row.pack(fill="x", pady=1)
                tk.Label(row, text="•", bg=C_PANEL, fg=C_ACCENT,
                         font=F_BODY).pack(side="left", padx=(0, 8))
                tk.Label(row, text=pdf.name, bg=C_PANEL, fg=C_TEXT,
                         font=F_BODY, anchor="w").pack(side="left")

            status_lbl = tk.Label(pkg_card, text="", bg=C_PANEL, fg=C_SUBTLE,
                                  font=F_BODY, anchor="w")
            status_lbl.pack(fill="x", pady=(8, 0))

            print_btn = tk.Button(pkg_card, text="Print All — Double-Sided",
                                  font=("Segoe UI", 10, "bold"),
                                  bg=C_ACCENT, fg="white", relief="flat",
                                  cursor="hand2", padx=14, pady=6)
            print_btn.pack(anchor="w", pady=(6, 0))

            def _start_print():
                print_btn.config(state="disabled")
                status_lbl.config(text="Starting print job…", fg=C_SUBTLE)

                def _worker():
                    acrobat = _find_acrobat()
                    printer = PREFERRED_PRINTER
                    _dpp_set_devmode_duplex(printer, True)
                    total = len(pdfs)
                    for i, pdf in enumerate(pdfs, 1):
                        parent.after(0, lambda n=pdf.name, idx=i:
                                     status_lbl.config(
                                         text=f"Printing {idx}/{total}: {n}…",
                                         fg=C_SUBTLE))
                        try:
                            _dpp_acrobat_print(pdf, printer, True, acrobat)
                        except Exception as exc:
                            parent.after(0, lambda e=str(exc):
                                         status_lbl.config(
                                             text=f"Error: {e}", fg="#EF4444"))
                            parent.after(0, lambda: print_btn.config(state="normal"))
                            return
                    parent.after(0, lambda:
                                 status_lbl.config(
                                     text=f"Done — {total} document(s) sent to printer.",
                                     fg="#16A34A"))
                    parent.after(0, lambda: print_btn.config(state="normal"))

                t = threading.Thread(target=_worker, daemon=True)
                t.start()

            print_btn.config(command=_start_print)
        else:
            tk.Label(pkg_card,
                     text="No PDFs found in tools/EngineeringDesignPackage/.",
                     bg=C_PANEL, fg="#EF4444", font=F_BODY).pack(anchor="w")

        # ── Design Reference card ─────────────────────────────────────
        ref_card = self._card(parent, "Design Reference")

        tk.Label(ref_card,
                 text="Open the FoxFab design tips and best practices document.",
                 bg=C_PANEL, fg=C_SUBTLE, font=F_BODY, anchor="w").pack(fill="x", pady=(0, 8))

        if tips_doc.is_file():
            tk.Button(ref_card, text="Open FoxFab Design Tips",
                      font=("Segoe UI", 10, "bold"),
                      bg=C_ACCENT, fg="white", relief="flat",
                      cursor="hand2", padx=14, pady=6,
                      command=lambda: os.startfile(str(tips_doc))
                      ).pack(anchor="w")
        else:
            tk.Label(ref_card,
                     text="FoxFab_Design_Tips.docx not found in tools/.",
                     bg=C_PANEL, fg="#EF4444", font=F_BODY).pack(anchor="w")


# ═════════════════════════════════════════════════════════════════════
#  MANUAL PRINTING PANEL
    def _mp_toggle(self):
        if self._mp_expanded.get():
            self._mp_section.pack_forget()
            self._mp_expanded.set(False)
            self._mp_toggle_btn.config(text="▶  Manual Printing")
        else:
            self._mp_section.pack(fill="x", before=self._panel_term_frames["dpp"])
            self._mp_expanded.set(True)
            self._mp_toggle_btn.config(text="▼  Manual Printing")

    def _mp_add_files(self):
        paths = filedialog.askopenfilenames(
            title="Select PDF files to print",
            filetypes=[("PDF files", "*.pdf *.PDF"), ("All files", "*.*")])
        for p in paths:
            self._mp_append_file(Path(p))

    def _mp_append_file(self, path: Path, duplex: bool = False):
        duplex_var = tk.BooleanVar(value=duplex)
        row = tk.Frame(self._mp_inner, bg=C_PANEL)
        row.pack(fill="x", pady=1)

        tk.Label(row, text=path.name, bg=C_PANEL, fg=C_TEXT,
                 font=F_BODY, anchor="w",
                 wraplength=500, justify="left").pack(
                     side="left", fill="x", expand=True, padx=(0, 8))

        tk.Checkbutton(row, variable=duplex_var, text="Duplex",
                       bg=C_PANEL, fg=C_TEXT, font=F_SMALL,
                       activebackground=C_PANEL, selectcolor=C_PANEL,
                       width=7).pack(side="left")

        entry = {"path": path, "duplex_var": duplex_var, "row": row}
        self._mp_files.append(entry)

        tk.Button(row, text="Print", font=F_SMALL,
                  bg=C_ACCENT, fg="#FFFFFF", activebackground=C_ACCENT_H,
                  activeforeground="#FFFFFF", relief="flat", padx=10, pady=3,
                  cursor="hand2",
                  command=lambda e=entry: self._mp_print_one(e)).pack(side="left", padx=(4, 4))

        tk.Button(row, text="✕", font=F_SMALL,
                  bg=C_BG, fg="#EF4444", relief="flat", padx=6, pady=3,
                  cursor="hand2",
                  command=lambda e=entry: self._mp_remove_file(e)).pack(side="left")

        self._mp_empty_lbl.pack_forget()
        self._mp_canvas.update_idletasks()
        self._mp_canvas.configure(scrollregion=self._mp_canvas.bbox("all"))

    def _mp_remove_file(self, entry: dict):
        entry["row"].destroy()
        self._mp_files = [e for e in self._mp_files if e is not entry]
        if not self._mp_files:
            self._mp_empty_lbl.pack(pady=14)
        self._mp_canvas.update_idletasks()
        self._mp_canvas.configure(scrollregion=self._mp_canvas.bbox("all"))

    def _mp_clear_all(self):
        for e in self._mp_files:
            e["row"].destroy()
        self._mp_files.clear()
        self._mp_empty_lbl.pack(pady=14)
        self._mp_canvas.update_idletasks()
        self._mp_canvas.configure(scrollregion=self._mp_canvas.bbox("all"))

    def _mp_print_one(self, entry: dict):
        # Capture values on the main thread before handing off
        printer = self._dpp_printer.get().strip()
        path    = entry["path"]
        duplex  = entry["duplex_var"].get()
        acrobat = _find_acrobat()
        self._active_key = "dpp"
        term = self._panel_terms.get("dpp")
        if term:
            self._term_write(term,
                f"Manual print: {path.name}  ({'duplex' if duplex else 'simplex'})", "info")
        q = self._log_queue

        def _do_print():
            try:
                if win32print and printer:
                    try:
                        win32print.SetDefaultPrinter(printer)
                    except Exception:
                        pass
                status = _dpp_acrobat_print(path, printer, duplex, acrobat)
                q.put(("ok", f"  ✓  {status}"))
            except Exception as e:
                q.put(("error", f"  ✗  {e}"))

        threading.Thread(target=_do_print, daemon=True).start()

    def _mp_print_all(self):
        if not self._mp_files:
            messagebox.showwarning(APP_TITLE, "No files in the list.")
            return
        printer = self._dpp_printer.get().strip()
        files   = list(self._mp_files)
        self._active_key = "dpp"

        def worker():
            if pythoncom:
                pythoncom.CoInitialize()
            q     = self._log_queue
            jobs  = [(e["path"], e["duplex_var"].get()) for e in files]
            total = len(jobs)

            def emit(msg, tag="info"):
                q.put((tag, msg))

            for i, (path, duplex) in enumerate(jobs):
                emit(f"[{i+1}/{total}]  {path.name}  ({'duplex' if duplex else 'simplex'})")

            com_ok = _dpp_acrobat_print_all_com(jobs, printer, emit)

            if not com_ok:
                emit("COM unavailable — falling back to subprocess per document", "warn")
                acrobat = _find_acrobat()
                for i, (path, duplex) in enumerate(jobs):
                    try:
                        status = _dpp_acrobat_print(path, printer, duplex, acrobat)
                        emit(f"  ✓  {status}", "ok")
                    except Exception as e:
                        emit(f"  ✗  {e}", "error")
                    if i < total - 1:
                        _dpp_wait_spooler_stable(printer)

            emit("Print All complete.", "ok")

        threading.Thread(target=worker, daemon=True).start()


# ═════════════════════════════════════════════════════════════════════
#  BOM FILLER  —  module-level helpers
# ═════════════════════════════════════════════════════════════════════

def _run_es(args: list[str]) -> str:
    try:
        r = subprocess.run([ES_EXE] + args,
                           capture_output=True, text=True, timeout=10)
        return r.stdout.strip()
    except Exception:
        return ""


def _bom_check_stock(pn: str) -> bool:
    out = _run_es(["-path", STOCK_PARTS_FOLDER, pn])
    return bool([l for l in out.splitlines() if l.strip()])


def _bom_find_revision(pn: str) -> str:
    out = _run_es([pn])
    highest = ""
    for line in out.splitlines():
        stem = Path(line.strip()).stem
        if not stem.upper().startswith(pn.upper()):
            continue
        suffix = stem[len(pn):].strip()
        m = re.match(r'^[-_\s]?r([A-Za-z])$', suffix)
        if m:
            r = m.group(1).upper()
            if r > highest:
                highest = r
    return f"r{highest}" if highest else ""


def _bom_find_copy(pn: str, rev: str, target: Path, ext: str):
    out = _run_es([f"ext:{ext}", pn])
    if not out:
        return False
    expected = ({f"{pn} {rev}".upper(), f"{pn}-{rev}".upper()}
                if rev else {pn.upper()})
    ext_up = f".{ext.upper()}"
    for line in out.splitlines():
        p = Path(line.strip())
        if p.stem.upper() in expected and p.suffix.upper() == ext_up:
            if p.exists():
                dest = target / p.name
                if dest.exists():
                    return None
                shutil.copy2(str(p), str(dest))
                return True
    return False


# ═════════════════════════════════════════════════════════════════════
#  DOC PREP & PRINT  —  module-level helpers
# ═════════════════════════════════════════════════════════════════════

def _dpp_get_excel():
    if win32com is None:
        raise RuntimeError("pywin32 required. pip install pywin32")
    app = win32com.client.DispatchEx("Excel.Application")
    app.Visible = False
    app.DisplayAlerts = False
    return app


def _dpp_classify_cnc(p: Path) -> str:
    base = p.stem
    if re.match(r"^[Jj]", base) or re.match(r"^\d{3}-", base):
        return "duplex"
    return "simplex"


def _dpp_list_files(folder: Path) -> list[Path]:
    return sorted([p for p in folder.iterdir() if p.is_file()],
                  key=lambda p: p.name.lower())


def _dpp_match_fwo(folder: Path) -> Path:
    for f in _dpp_list_files(folder):
        if (f.suffix in PDF_EXTENSIONS
                and f.stem == "Fabrication Work Order - Standard v1.0"):
            return f
    raise RuntimeError("Fabrication Work Order PDF not found in 300 Inputs.")


def _dpp_revision_letter(p: Path) -> str:
    """Extract the revision letter from a filename like 'J12345 rB BOM.xlsx'.
    Matches ' rA', ' rB', etc. (case-insensitive). Returns '' if not found."""
    m = re.search(r'\br([A-Za-z])\b', p.stem)
    return m.group(1).upper() if m else ""


def _dpp_match_excel(folder: Path, token: str, title: str,
                     app: "App") -> Path:
    matches = [f for f in _dpp_list_files(folder)
               if f.suffix.lower() in {e.lower() for e in EXCEL_EXTENSIONS}
               and token.lower() in f.name.lower()]
    if not matches:
        raise RuntimeError(f"No Excel file containing '{token}' found in {title}.")
    if len(matches) == 1:
        return matches[0]
    # Auto-select latest revision (rA < rB < rC ...) when revision info is present
    with_rev = [(f, _dpp_revision_letter(f)) for f in matches]
    if any(rev for _, rev in with_rev):
        return max(with_rev, key=lambda x: x[1])[0]
    # No revision info — fall back to user prompt
    choice = app._pick_from_list(f"Select {title}", [f.name for f in matches])
    if not choice:
        raise RuntimeError(f"{title} selection cancelled.")
    return next(f for f in matches if f.name == choice)


def _dpp_match_pack(folder: Path, app: "App", model_no: str = "") -> Path:
    """Find the PACK PDF. Tries model number first, falls back to any PACK, prompts if ambiguous."""
    all_packs = [f for f in _dpp_list_files(folder)
                 if f.suffix in PDF_EXTENSIONS and "PACK" in f.name.upper()]
    if not all_packs:
        raise RuntimeError("No PDF with 'PACK' in name found in Electrical Drawings.")

    if model_no:
        model_up = model_no.upper()
        model_matches = [f for f in all_packs if model_up in f.name.upper()]
        if len(model_matches) == 1:
            return model_matches[0]
        if len(model_matches) > 1:
            choice = app._pick_from_list(
                "Select Electrical Drawing Pack",
                [f.name for f in model_matches],
                prompt=f"Multiple PACK PDFs match model '{model_no}':")
            if not choice:
                raise RuntimeError("Pack PDF selection cancelled.")
            return next(f for f in model_matches if f.name == choice)
        # No model match — fall through to any PACK, with a prompt
        choice = app._pick_from_list(
            "Select Electrical Drawing Pack",
            [f.name for f in all_packs],
            prompt=f"No PACK PDF found matching model '{model_no}'. Select manually:")
        if not choice:
            raise RuntimeError("Pack PDF selection cancelled.")
        return next(f for f in all_packs if f.name == choice)

    if len(all_packs) == 1:
        return all_packs[0]
    choice = app._pick_from_list("Select Electrical Drawing Pack",
                                  [f.name for f in all_packs])
    if not choice:
        raise RuntimeError("Pack PDF selection cancelled.")
    return next(f for f in all_packs if f.name == choice)


def _dpp_find_variant_prf(prf_folder: Path, mechs: list, app: "App") -> Path:
    """Find the PRF Excel file. When one variant is selected, prefers a PRF with its suffix."""
    all_prfs = [f for f in _dpp_list_files(prf_folder)
                if f.suffix.lower() in {e.lower() for e in EXCEL_EXTENSIONS}
                and "prf" in f.name.lower()]
    if not all_prfs:
        raise RuntimeError(f"No PRF file found in {prf_folder}")

    # Extract variant suffix from the first mech name (e.g. "J15302-01" → "-01")
    variant_suffix = ""
    if len(mechs) == 1:
        m = re.search(r"(-\d{2})$", mechs[0].name)
        if m:
            variant_suffix = m.group(1)

    if variant_suffix:
        variant_prfs = [f for f in all_prfs if variant_suffix in f.name]
        if len(variant_prfs) == 1:
            return variant_prfs[0]
        if len(variant_prfs) > 1:
            choice = app._pick_from_list(
                f"Select PRF for variant {variant_suffix}",
                [f.name for f in variant_prfs])
            if not choice:
                raise RuntimeError("PRF selection cancelled.")
            return next(f for f in variant_prfs if f.name == choice)
        # No variant-specific PRF found
        if len(all_prfs) == 1:
            return all_prfs[0]   # only one PRF — use it
        choice = app._pick_from_list(
            f"Select PRF for variant {variant_suffix}",
            [f.name for f in all_prfs],
            prompt=f"No PRF found with '{variant_suffix}' in name. Select manually:")
        if not choice:
            raise RuntimeError("PRF selection cancelled.")
        return next(f for f in all_prfs if f.name == choice)

    # No variant suffix — standard match
    if len(all_prfs) == 1:
        return all_prfs[0]
    choice = app._pick_from_list("Select Production Release Form",
                                  [f.name for f in all_prfs])
    if not choice:
        raise RuntimeError("PRF selection cancelled.")
    return next(f for f in all_prfs if f.name == choice)


def _dpp_read_prf(prf_path: Path) -> dict:
    """Read key fields from the PRF using openpyxl (no COM needed)."""
    if _openpyxl is None:
        raise RuntimeError("openpyxl required for PRF auto-fill. pip install openpyxl")
    wb = _openpyxl.load_workbook(str(prf_path), data_only=True)
    ws = wb["Form"] if "Form" in wb.sheetnames else wb.worksheets[0]

    def cell(ref: str) -> str:
        v = ws[ref].value
        return str(v).strip() if v is not None else ""

    job_no   = cell("C4")
    model_no = cell("G9")
    job_name = cell("C8")
    size     = cell("G18")
    material = cell("G19")
    rating   = cell("G20")
    qty      = cell("G22")

    mat_lc = material.lower()
    if "aluminum" in mat_lc or "aluminium" in mat_lc:
        mat_abbr = "ALU"
    elif "stainless" in mat_lc:
        mat_abbr = "SS"
    else:
        mat_abbr = material

    rat_lc = rating.lower().strip()
    if rat_lc in ("type 3r", "type3r"):
        rat_abbr = "N3R"
    else:
        rat_abbr = rating

    enclosure = " ".join(p for p in [size, mat_abbr, rat_abbr] if p)

    return {
        "job_no":    job_no,
        "model_no":  model_no,
        "job_name":  job_name,
        "enclosure": enclosure,
        "qty":       qty,
        # raw fields for summary
        "size":      size,
        "material":  material,
        "rating":    rating,
    }


def _dpp_fill_fwo(fwo_path: Path, prf_data: dict) -> Path:
    """Overlay PRF data onto the FWO PDF. Returns path to a filled temp PDF.
    Adjust FWO_* constants at the top of the file if text position is off."""
    if _fitz is None:
        raise RuntimeError("PyMuPDF required for FWO auto-fill. pip install pymupdf")
    today = datetime.now().strftime("%B %d, %Y")   # e.g. March 26, 2026
    fields = [
        (FWO_JOB_NO_X,    FWO_JOB_NO_Y,    prf_data["job_no"]),
        (FWO_JOB_NAME_X,  FWO_JOB_NAME_Y,  prf_data["job_name"]),
        (FWO_DATE_X,      FWO_DATE_Y,       today),
        (FWO_ENCLOSURE_X, FWO_ENCLOSURE_Y,  prf_data["enclosure"]),
        (FWO_UNITS_X,     FWO_UNITS_Y,      prf_data["qty"]),
    ]
    doc = _fitz.open(str(fwo_path))
    page = doc[0]
    for x, y, text in fields:
        if text:
            page.insert_text(
                _fitz.Point(x, y), text,
                fontsize=FWO_FONT_SIZE,
                fontname="helv",
                color=(0, 0, 0),
            )
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix="_FWO_filled.pdf")
    tmp_path = Path(tmp.name)
    tmp.close()
    doc.save(str(tmp_path))
    doc.close()
    return tmp_path


def _dpp_match_cnc(folder: Path) -> list[Path]:
    pdfs = [f for f in _dpp_list_files(folder) if f.suffix in PDF_EXTENSIONS]
    if not pdfs:
        raise RuntimeError("No PDFs found in CNC folder.")
    return sorted(pdfs, key=lambda p: p.name.lower())


def _dpp_match_flats(folder: Path) -> list[Path]:
    pdfs = [f for f in _dpp_list_files(folder) if f.suffix in PDF_EXTENSIONS]
    if not pdfs:
        raise RuntimeError("No PDFs found in PDFs_Flats folder.")
    return sorted(pdfs, key=lambda p: p.name.lower())


def _dpp_match_assemblies(folder: Path) -> tuple[list[Path], list[Path]]:
    pdfs, lay = [], []
    for f in _dpp_list_files(folder):
        if f.suffix not in PDF_EXTENSIONS:
            continue
        if f.stem.endswith("-LAY"):
            lay.append(f)
        else:
            pdfs.append(f)
    if not pdfs:
        raise RuntimeError("No printable PDFs found in Assemblies folder.")
    return (sorted(pdfs, key=lambda p: p.name.lower()),
            sorted(lay, key=lambda p: p.name.lower()))


def _dpp_get_context(job_folder: Path) -> dict:
    def is_variant(p: Path):
        return p.is_dir() and re.search(r"-\d{2}$", p.name)

    def has_mech_subs(p: Path):
        return all((p / s).is_dir() for s in
                   ["204 BOM", "205 CNC", "202 PDFs_Flats", "203 Assemblies"])

    if is_variant(job_folder) and has_mech_subs(job_folder):
        return {"job_root": job_folder.parent.parent,
                "mech_roots": [job_folder],
                "variant_only": True}

    mech = job_folder / "200 Mech"
    if not mech.is_dir():
        raise RuntimeError("Missing folder: 200 Mech")

    if has_mech_subs(mech):
        return {"job_root": job_folder,
                "mech_roots": [mech],
                "variant_only": False}

    variants = [c for c in sorted(mech.iterdir(), key=lambda p: p.name.lower())
                if is_variant(c) and has_mech_subs(c)]
    if variants:
        return {"job_root": job_folder,
                "mech_roots": variants,
                "variant_only": False}

    raise RuntimeError(
        "Could not find a usable mechanical folder structure. "
        "Expected '200 Mech\\204 BOM' etc. directly, or numbered "
        "variants like '*-01' inside 200 Mech.")


def _find_acrobat() -> str | None:
    """Return the path to Acrobat.exe, or None if not found."""
    for p in ACROBAT_SEARCH_PATHS:
        if os.path.isfile(p):
            return p
    return None


def _dpp_set_devmode_duplex(printer_name: str, duplex: bool) -> bool:
    """Set per-user duplex preference for the printer via PRINTER_INFO_9.
    Level 9 requires only PRINTER_ACCESS_USE — no admin rights needed.
    Returns True on success, False if it could not be applied."""
    if win32print is None:
        return False
    try:
        h = win32print.OpenPrinter(printer_name)
        try:
            # Prefer per-user DEVMODE (level 9); fall back to global (level 2)
            info9 = win32print.GetPrinter(h, 9)
            dm = info9.get("pDevMode") if info9 else None
            if dm is None:
                info2 = win32print.GetPrinter(h, 2)
                dm = info2.get("pDevMode")
            if dm is None:
                return False
            dm.Duplex = 2 if duplex else 1   # 2 = DMDUP_VERTICAL (long-edge), 1 = simplex
            win32print.SetPrinter(h, 9, {"pDevMode": dm}, 0)
            return True
        finally:
            win32print.ClosePrinter(h)
    except Exception:
        return False


def _dpp_wait_spooler_stable(printer_name: str, timeout: float = 60.0):
    """
    Block until the printer has no jobs in JOB_STATUS_SPOOLING state, then return.

    Called after each Acrobat print submission so that large PDFs (which take
    longer to spool) cannot be overtaken in the queue by the next document.
    Falls back to a 3.5-second sleep if win32print is unavailable or the
    printer cannot be opened.
    """
    JOB_STATUS_SPOOLING = 0x0004
    POLL_INTERVAL       = 0.5   # seconds between queue checks

    if not win32print:
        time.sleep(3.5)
        return

    deadline = time.time() + timeout
    while time.time() < deadline:
        try:
            h = win32print.OpenPrinter(printer_name)
            try:
                jobs = win32print.EnumJobs(h, 0, 999, 1)
            finally:
                win32print.ClosePrinter(h)
            if not any(j["Status"] & JOB_STATUS_SPOOLING for j in jobs):
                return   # all jobs fully spooled — safe to send next document
        except Exception:
            time.sleep(3.5)
            return
        time.sleep(POLL_INTERVAL)
    # Timeout reached — proceed anyway rather than blocking forever


def _dpp_acrobat_print(pdf_path: Path, printer_name: str, duplex: bool,
                        acrobat_exe: str | None) -> str:
    """Print one PDF via Acrobat with the specified duplex setting.
    Sets per-user DEVMODE then launches Acrobat /t for silent printing.
    Returns a status string for the terminal."""
    devmode_ok = _dpp_set_devmode_duplex(printer_name, duplex)
    mode_str = ("duplex" if duplex else "simplex") + ("" if devmode_ok else " [devmode failed]")

    if acrobat_exe:
        try:
            # subprocess.run is BLOCKING — Acrobat must fully close before we return.
            # This guarantees print ORDER (no race to the spooler) and correct DUPLEX
            # (next DEVMODE change only happens after the previous Acrobat instance exits).
            subprocess.run(
                [acrobat_exe, "/t", str(pdf_path), printer_name],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                timeout=120,
            )
        except subprocess.TimeoutExpired:
            pass   # Acrobat hung — job was likely submitted; move on
        return f"sent via Acrobat ({mode_str})"
    else:
        os.startfile(str(pdf_path), "print")
        return f"sent via default handler (no duplex control)"


def _dpp_acrobat_print_all_com(
    print_jobs: list[tuple[Path, bool]],
    printer_name: str,
    emit,
) -> bool:
    """
    Print all PDFs through a SINGLE Acrobat COM instance that stays open for the
    entire run.  Eliminates the 3-5 s Acrobat startup cost per document.

    print_jobs : list of (pdf_path, is_duplex)
    emit       : callable(msg, tag) for terminal output
    Returns True  if all jobs were dispatched via COM.
    Returns False if COM is unavailable so the caller can fall back to subprocess.
    """
    if win32com is None:
        return False

    try:
        acro_app = win32com.client.Dispatch("AcroExch.App")
        acro_app.Hide()
    except Exception as e:
        emit(f"Acrobat COM init failed ({e}) — falling back to subprocess", "warn")
        return False

    try:
        for i, (pdf_path, is_duplex) in enumerate(print_jobs):
            devmode_ok = _dpp_set_devmode_duplex(printer_name, is_duplex)
            mode_str = ("duplex" if is_duplex else "simplex") + (
                "" if devmode_ok else " [devmode failed]")

            pddoc = None
            avdoc = None
            try:
                pddoc = win32com.client.Dispatch("AcroExch.PDDoc")
                if not pddoc.Open(str(pdf_path)):
                    emit(f"  ✗  COM could not open: {pdf_path.name}", "error")
                    continue

                n_pages = pddoc.GetNumPages()
                avdoc   = pddoc.OpenAVDoc("")
                if avdoc is None:
                    emit(f"  ✗  Could not get AVDoc: {pdf_path.name}", "error")
                    continue

                # PrintPages(nFirstPage, nLastPage, nPrintFlags, bAnnotations, bShrinkToFit)
                # nPrintFlags = 2  →  kAVPrintSilent: no dialog, no progress bar
                ok = avdoc.PrintPages(0, n_pages - 1, 2, True, False)
                if ok:
                    emit(f"  ✓  sent via Acrobat COM ({mode_str})", "ok")
                else:
                    emit(f"  ✗  PrintPages returned False for {pdf_path.name}", "error")

            except Exception as e:
                emit(f"  ✗  {pdf_path.name}: {e}", "error")
            finally:
                if avdoc:
                    try:
                        avdoc.Close(True)
                    except Exception:
                        pass
                if pddoc:
                    try:
                        pddoc.Close()
                    except Exception:
                        pass

            # Small breath between docs — no spooler poll needed because
            # COM PrintPages submits jobs sequentially through one instance,
            # so queue order is already guaranteed by call order.
            if i < len(print_jobs) - 1:
                time.sleep(0.2)

    finally:
        try:
            acro_app.Exit()
        except Exception:
            pass

    return True


def dpp_build_plan(job_folder: Path, app: "App") -> dict:
    ctx = _dpp_get_context(job_folder)
    base = Path(ctx["job_root"])
    mechs = ctx["mech_roots"]

    # If multiple variants found and user didn't select one directly, let them choose
    if len(mechs) > 1 and not ctx["variant_only"]:
        choice = app._pick_from_list(
            "Select Mechanical Variant",
            [m.name for m in mechs],
            prompt="Choose which variant to print:")
        if choice is None:
            raise RuntimeError("Variant selection cancelled.")
        mechs = [m for m in mechs if m.name == choice]

    plan: dict = {
        "job_folder": str(job_folder),
        "base": str(base),
        "mech_roots": [str(m) for m in mechs],
        "variant_only": ctx["variant_only"],
    }

    plan["fwo"] = _dpp_match_fwo(base / "300 Inputs")
    plan["bom"] = _dpp_match_excel(mechs[0] / "204 BOM", "BOM", "BOM", app)

    # Variant-aware PRF: prefer a PRF named with the variant suffix (e.g. -01)
    prf_folder = base / "300 Inputs" / "302 Production Release Form"
    plan["prf"] = _dpp_find_variant_prf(prf_folder, mechs, app)

    # Read PRF data to get model number for electrical drawing matching
    prf_data: dict | None = None
    model_no = ""
    if _openpyxl is not None:
        try:
            prf_data = _dpp_read_prf(plan["prf"])
            model_no = prf_data.get("model_no", "")
        except Exception:
            prf_data = None  # non-fatal — will fill FWO manually

    plan["prf_data"]  = prf_data
    plan["pack"]      = _dpp_match_pack(
        base / "100 Elec" / "102 Drawings", app, model_no=model_no)

    cnc, flats, assemblies, excluded_lay = [], [], [], []
    for m in mechs:
        cnc.extend(_dpp_match_cnc(m / "205 CNC"))
        flats.extend(_dpp_match_flats(m / "202 PDFs_Flats"))
        a, ex = _dpp_match_assemblies(m / "203 Assemblies")
        assemblies.extend(a)
        excluded_lay.extend(ex)

    plan["cnc"]              = sorted(cnc, key=lambda p: p.name.lower())
    plan["flats"]            = sorted(flats, key=lambda p: p.name.lower())
    plan["assemblies"]       = sorted(assemblies, key=lambda p: p.name.lower())
    plan["assemblies_excl"]  = sorted(excluded_lay, key=lambda p: p.name.lower())
    return plan


def dpp_make_summary(plan: dict, printer: str, simulation: bool) -> str:
    cnc_d = sum(1 for p in plan["cnc"] if _dpp_classify_cnc(p) == "duplex")
    cnc_s = len(plan["cnc"]) - cnc_d
    mode  = "SIMULATION MODE (save PDFs)" if simulation else f"PRINT  »  {printer}"
    lines = [
        f"Mode: {mode}",
        f"Job Folder: {plan['job_folder']}",
        "",
    ]

    # PRF data block
    prf_data = plan.get("prf_data")
    if prf_data:
        lines += [
            f"  Job No.    : {prf_data['job_no']}",
            f"  Job Name   : {prf_data['job_name']}",
            f"  Enclosure  : {prf_data['enclosure']}   Qty: {prf_data['qty']}",
            f"  Model No.  : {prf_data['model_no']}",
            "",
        ]

    lines += [
        f"  Fabrication Work Order : {plan['fwo'].name}"
        + (" (auto-filled)" if prf_data else ""),
        f"  BOM                    : {plan['bom'].name}",
        f"  CNC files              : {len(plan['cnc'])} PDFs"
        f"  ({cnc_d} duplex, {cnc_s} simplex merged)",
        f"  PDFs_Flats             : {len(plan['flats'])} PDFs (merged)",
        f"  Production Release Form: {plan['prf'].name}",
        f"  Electrical Pack        : {plan['pack'].name} (pages 1-2)",
        f"  Assemblies             : {len(plan['assemblies'])} PDFs (merged,"
        f" excl. {len(plan['assemblies_excl'])} -LAY files)",
    ]
    return "\n".join(lines)


def _dpp_merge_pdfs(files: list[Path], label: str) -> Path:
    if PdfReader is None:
        raise RuntimeError("pypdf not installed. pip install pypdf")
    writer = PdfWriter()
    for pdf in files:
        try:
            reader = PdfReader(str(pdf))
            for page in reader.pages:
                writer.add_page(page)
        except Exception as e:
            raise RuntimeError(f"Could not merge {pdf.name} into {label}: {e}") from e
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=f"_{label}.pdf")
    tmp_path = Path(tmp.name)
    tmp.close()
    with open(tmp_path, "wb") as f:
        writer.write(f)
    return tmp_path



def _dpp_sim_save_pdf(src: Path, dest: Path, pages=None):
    if pages is not None:
        reader = PdfReader(str(src))
        writer = PdfWriter()
        for i in range(pages[0], pages[1] + 1):
            writer.add_page(reader.pages[i])
        with open(dest, "wb") as f:
            writer.write(f)
    else:
        shutil.copy2(str(src), str(dest))


def _dpp_sim_excel_to_pdf(file: Path, dest: Path,
                           first_sheet_only: bool, excel=None):
    created = excel is None
    if created:
        excel = _dpp_get_excel()
    wb = None
    try:
        wb = excel.Workbooks.Open(str(file))
        if first_sheet_only:
            ws = wb.Worksheets(1)
            ws.ExportAsFixedFormat(0, str(dest))
        else:
            active = wb.ActiveSheet
            ps = active.PageSetup
            ps.Orientation = 2
            ps.Zoom = False
            ps.FitToPagesWide = 1
            ps.FitToPagesTall = False
            active.ExportAsFixedFormat(0, str(dest))
    finally:
        if wb is not None:
            try:
                wb.Close(False)
            except Exception:
                pass
        if created and excel:
            try:
                excel.Quit()
            except Exception:
                pass


def _dpp_sim_save_merged(files: list[Path], dest: Path):
    merged = _dpp_merge_pdfs(files, dest.stem)
    try:
        shutil.copy2(str(merged), str(dest))
    finally:
        if merged.exists():
            try:
                merged.unlink()
            except Exception:
                pass


def dpp_build_sections(plan: dict, out_dir: Path) -> list:
    """Build all documents as PDFs into out_dir. Used for both simulation and print modes."""
    sections = []
    n = [0]

    def dest(label: str) -> Path:
        n[0] += 1
        return out_dir / f"{n[0]:02d}_{safe_name(label)}.pdf"

    excel_holder = [None]

    def get_xl():
        if excel_holder[0] is None:
            excel_holder[0] = _dpp_get_excel()
        return excel_holder[0]

    def quit_xl():
        if excel_holder[0]:
            try:
                excel_holder[0].Quit()
            except Exception:
                pass
            excel_holder[0] = None

    _fwo_src = plan.get("fwo_filled") or plan["fwo"]
    sections.append(("Fabrication Work Order",
        lambda _f=_fwo_src: _dpp_sim_save_pdf(_f, dest("Fabrication_Work_Order"))))

    sections.append(("BOM",
        lambda: _dpp_sim_excel_to_pdf(
            plan["bom"], dest("BOM"),
            first_sheet_only=False, excel=get_xl())))

    for pdf in plan["cnc"]:
        p = pdf
        if _dpp_classify_cnc(p) == "duplex":
            sections.append((f"CNC (duplex): {p.name}",
                lambda _p=p: _dpp_sim_save_pdf(_p, dest(f"CNC_{_p.stem}"))))

    cnc_simplex = [p for p in plan["cnc"] if _dpp_classify_cnc(p) == "simplex"]
    if cnc_simplex:
        sections.append(("CNC Simplex (merged)",
            lambda: _dpp_sim_save_merged(cnc_simplex, dest("CNC_Simplex_Merged"))))

    sections.append(("PDFs_Flats (merged)",
        lambda: _dpp_sim_save_merged(plan["flats"], dest("PDFs_Flats_Merged"))))

    sections.append(("Production Release Form",
        lambda: _dpp_sim_excel_to_pdf(
            plan["prf"], dest("Production_Release_Form"),
            first_sheet_only=True, excel=get_xl())))

    sections.append(("Electrical Pack (pages 1-2)",
        lambda: _dpp_sim_save_pdf(
            plan["pack"], dest("Electrical_Pack_Pages_1_2"), pages=(0, 1))))

    sections.append(("Assemblies (merged)",
        lambda: _dpp_sim_save_merged(plan["assemblies"], dest("Assemblies_Merged"))))

    sections.append(("Cleanup Excel", quit_xl))

    return sections


# ═════════════════════════════════════════════════════════════════════
#  CNC COLUMN MARKER HELPERS
# ═════════════════════════════════════════════════════════════════════

# Revision suffix patterns:  "rB"  or  " rB"  at end of string
_CNC_REV_RE = re.compile(r"\s*r[A-Za-z]$", re.IGNORECASE)

# Recognised prefix families that appear directly in CNC filenames
_CNC_DIRECT_PREFIXES = re.compile(
    r"^(240|250|200|295|210|220|230|260|270|280|290)-(\d+)")

# Bare digit segment (no prefix): prepend "240-"
_CNC_BARE_DIGITS = re.compile(r"^\d+$")

# J-prefix GALV files
_CNC_J_PREFIX = re.compile(r"^J\d+", re.IGNORECASE)

# DRAWING NUMBER line in GALV PDF text
_CNC_DRAWING_NO_RE = re.compile(
    r"DRAWING\s+NUMBER\s*:\s*([0-9]{3}-[0-9]+(?:\s*r[A-Za-z])?)", re.IGNORECASE)


def _cnc_strip_rev(s: str) -> str:
    """Remove trailing revision suffix (e.g. ' rB', 'rC') from a part number string."""
    return _CNC_REV_RE.sub("", s).strip()


def _cnc_parts_from_rest(prefix: str, rest: str) -> list[str]:
    """
    Given the prefix family (e.g. '240') and the remainder of a CNC filename
    after the leading NNN-XXXXX segment, extract additional part numbers.

    Segments separated by '_' that are all-digit become additional part numbers
    under the same prefix.  Non-digit segments end the scan.
    """
    parts: list[str] = []
    for seg in rest.split("_"):
        seg = seg.strip()
        if not seg:
            continue
        if _CNC_BARE_DIGITS.match(seg):
            parts.append(f"{prefix}-{seg}")
        else:
            break   # stop at first non-digit token (description text)
    return parts


def _cnc_extract_galv_parts(pdf_path: Path) -> list[str]:
    """
    Open a GALV-style PDF (J##### prefix) with PyMuPDF and return all part
    numbers found on 'DRAWING NUMBER: NNN-XXXXX' lines, stripped of revisions.
    Returns [] if fitz is unavailable or the file can't be read.
    """
    if _fitz is None:
        return []
    parts: list[str] = []
    try:
        doc = _fitz.open(str(pdf_path))
        for page in doc:
            text = page.get_text()
            for m in _CNC_DRAWING_NO_RE.finditer(text):
                raw = m.group(1).strip()
                pn = _cnc_strip_rev(raw)
                if pn and pn not in parts:
                    parts.append(pn)
        doc.close()
    except Exception:
        pass
    return parts


def _cnc_parse_filename(pdf_path: Path, emit=None) -> list[str]:
    """
    Parse a CNC PDF filename and return a list of clean part numbers (no revisions).

    Rules (in order):
      1. Skip files containing 'Merged' (caller should already filter these).
      2. J-prefix → open PDF, extract DRAWING NUMBER lines.
      3. NNN-XXXXX direct prefix → parse stem; additional _DIGIT segments = more PNs.
      4. Bare digits → prepend '240-'.
      5. Unrecognised → skip with a warning.
    """
    stem = pdf_path.stem   # filename without extension

    # Strip trailing revision from the stem itself before analysing
    stem_clean = _CNC_REV_RE.sub("", stem).strip()

    # Rule 2 — GALV / J-prefix
    if _CNC_J_PREFIX.match(stem_clean):
        parts = _cnc_extract_galv_parts(pdf_path)
        if emit:
            if parts:
                emit(f"  {pdf_path.name}  →  GALV  {parts}", "info")
            else:
                emit(f"  {pdf_path.name}  →  GALV  (no DRAWING NUMBER lines found)", "warn")
        return parts

    # Rule 3 — direct prefix NNN-XXXXX[_more]
    m = _CNC_DIRECT_PREFIXES.match(stem_clean)
    if m:
        prefix = m.group(1)   # e.g. "240"
        first_num = m.group(2)
        first_pn = f"{prefix}-{first_num}"
        rest = stem_clean[m.end():]   # everything after the first NNN-XXXXX
        extra = _cnc_parts_from_rest(prefix, rest)
        parts = [first_pn] + extra
        if emit:
            emit(f"  {pdf_path.name}  →  {parts}", "info")
        return parts

    # Rule 4 — bare digits
    if _CNC_BARE_DIGITS.match(stem_clean):
        pn = f"240-{stem_clean}"
        if emit:
            emit(f"  {pdf_path.name}  →  {[pn]}", "info")
        return [pn]

    # Rule 5 — unrecognised
    if emit:
        emit(f"  {pdf_path.name}  →  unrecognised pattern — skipped", "warn")
    return []


def _cnc_do_mark(bom_path: Path, cnc_folder: Path, emit) -> tuple[int, list[str]]:
    """
    Open the BOM workbook via xlwings, scan cnc_folder for PDFs, and write 'X'
    to column H for every matching non-S row.

    Returns (matched_count, unmatched_part_numbers).
    Raises RuntimeError on hard failures (missing sheet, xlwings unavailable, etc.).
    The caller is responsible for xlwings COM initialisation if needed.
    """
    if xw is None:
        raise RuntimeError("xlwings is not installed. Run:  pip install xlwings")

    DATA_START = 6

    xw_app = xw.App(visible=False, add_book=False)
    xw_app.display_alerts = False
    xw_app.screen_updating = False
    try:
        try:
            wb = xw_app.books.open(str(bom_path))
        except Exception as e:
            raise RuntimeError(f"Could not open BOM workbook: {e}")

        sheet_names = [s.name for s in wb.sheets]
        if BOM_SHEET_NAME not in sheet_names:
            raise RuntimeError(
                f"Sheet '{BOM_SHEET_NAME}' not found. Available: {sheet_names}")

        ws = wb.sheets[BOM_SHEET_NAME]

        last_row = ws.range(f"A{DATA_START}").end("down").row
        if last_row > 1_000_000:
            last_row = DATA_START

        col_a = ws.range(f"A{DATA_START}:A{last_row}").value
        col_h = ws.range(f"H{DATA_START}:H{last_row}").value
        if not isinstance(col_a, list):
            col_a = [col_a]
        if not isinstance(col_h, list):
            col_h = [col_h]

        # Build part-number → file mapping from CNC folder
        cnc_pdfs = [
            f for f in cnc_folder.iterdir()
            if f.is_file()
            and f.suffix.lower() == ".pdf"
            and "merged" not in f.name.lower()
        ]
        emit(f"Found {len(cnc_pdfs)} CNC PDFs in {cnc_folder.name}", "info")

        cnc_parts: dict[str, Path] = {}
        for pdf in cnc_pdfs:
            for pn in _cnc_parse_filename(pdf, emit):
                cnc_parts[pn] = pdf

        emit(f"Resolved {len(cnc_parts)} unique part numbers from CNC folder", "info")
        emit("Matching BOM rows...", "heading")

        matched: list[int] = []
        unmatched: list[str] = []

        for i, pn_raw in enumerate(col_a):
            if not pn_raw or not str(pn_raw).strip():
                continue
            pn_str = str(pn_raw).strip()
            row_idx = DATA_START + i

            h_val = col_h[i] if i < len(col_h) else None
            if h_val and str(h_val).strip().upper() == "S":
                continue

            pn_clean = _cnc_strip_rev(pn_str)
            if pn_clean in cnc_parts:
                ws.range(f"H{row_idx}").value = "X"
                matched.append(row_idx)
                emit(f"  row {row_idx}  {pn_str:<28}  → X  ({cnc_parts[pn_clean].name})", "ok")
            else:
                unmatched.append(pn_str)

        wb.save()
        wb.close()
        return len(matched), unmatched

    finally:
        try:
            xw_app.quit()
        except Exception:
            pass


def _cnc_find_cnc_folder_for_bom(bom_path: Path) -> Path | None:
    """
    Locate the 205 CNC folder that sits alongside the BOM.
    Structure:  <variant>/204 BOM/<bom_file>
                <variant>/205 CNC/          ← what we want
    So we look for a '205' sibling of bom_path.parent.parent.
    """
    variant = bom_path.parent.parent
    try:
        return next(
            (d for d in variant.iterdir()
             if d.is_dir() and d.name.lower().startswith("205")),
            None,
        )
    except Exception:
        return None


# ═════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═════════════════════════════════════════════════════════════════════


def main():
    global _splash_root, _splash_win

    # Reuse the root that was created for the splash
    root = _splash_root

    # DPI awareness on Windows
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass

    # Build the app while the splash is still visible
    App(root)

    # Dismiss splash and reveal the fully-built main window
    if _splash_win and _splash_win.winfo_exists():
        _splash_win.destroy()
    _splash_win = None
    root.attributes("-topmost", False)
    root.deiconify()
    root.mainloop()


if __name__ == "__main__":
    main()
